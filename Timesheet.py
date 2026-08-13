# # """
# # Timesheet Auto-Filler - Streamlit App
# # =======================================
# # Upload SOW & Resource Details + Replicon Dump files, and this app will
# # automatically generate one filled Timesheet per employee.

# # Features:
# # - Upload SOW & Resource Details (.xlsx)
# # - Upload Replicon Dump (.xlsx)
# # - Timesheet template: use the BUILT-IN default template (hardcoded/embedded
# #   in this file as base64) OR upload your own custom template (.xlsx)
# # - Generates one filled timesheet per employee found in the SOW file
# # - Download each file individually, or all together as a ZIP

# # Run with:
# #     streamlit run timesheet_app.py
# # """

# # import streamlit as st
# # import pandas as pd
# # import openpyxl
# # import re
# # import io
# # import base64
# # import calendar
# # import zipfile
# # from datetime import datetime
# # from openpyxl.worksheet.protection import SheetProtection
# # from openpyxl.styles import Protection

# # st.set_page_config(page_title="Timesheet Auto-Filler", layout="wide", page_icon="📋")

# # # =====================================================================
# # # 1. HARDCODED DEFAULT TIMESHEET TEMPLATE (base64-embedded .xlsx)
# # #    This lets the app work out-of-the-box with no template upload needed.
# # #    A second option below lets the user upload a custom template instead.
# # # =====================================================================
# # DEFAULT_TEMPLATE_B64 = """
# # UEsDBBQABgAIAAAAIQCVZpbtkAEAAIIFAAATAAgCW0NvbnRlbnRfVHlwZXNdLnhtbCCiBAIooAACAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACslM9O4zAQxu8r7TtEvqLEhcMKoaYc+CMh7QIS7AMM9qSx6tiWZ4D27Zm4gBDqtqzoJVZiz/f9PJmZ6ely8NUTZnIxtOqwmagKg4nWhXmr/t5f1seqIoZgwceArVohqdPZzx/T+1VCqiQ6UKt65nSiNZkeB6AmJgyy08U8AMtrnusEZgFz1EeTyS9tYmAMXPOooWbTc+zg0XN1sZTPa5KMnlR1tj44erUKUvLOAAupfgr2k0v96tBIZDlDvUt0IBhKb3QYd/5t8Bp3I6nJzmJ1C5mvYRAMvfT6OebFQ4yLZrvIBsrYdc6gjeZxkAw0lDKCpR6RB9+UtRnAhTfuLf7lMOmyHO4ZZLxfEd7BwfK/UZfn9xGKzA5D4pVH2nfai+gu5x4y2jvO0hl7B/iovYPDgDdnvZTInpPwrrvNX+r2DzJYYNC/4QH9VejiF0AGqteF3xgPRE56oPSoHzXem3RTsYvjbY6JZGZk/P8rvw2FMbpOIoSZHX7NUebNt3OM40SzaDd46zJBZy8AAAD//wMAUEsDBBQABgAIAAAAIQAxHYnNIgEAAN4CAAALAAgCX3JlbHMvLnJlbHMgogQCKKAAAgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAArJLRSgMxEEXfBf8h5L0721VEpNu+iFBQEKkfME1mt6FJJiRR2783rRZdqEXQx2Tu3Jy5k8ls46x4pZgM+1aOq1oK8oq18X0rnxd3o2spUkav0bKnVm4pydn0/GzyRBZzaUorE5IoLj61cpVzuAFIakUOU8WBfKl0HB3mcow9BFRr7Amaur6C+N1DTgeeYq5bGef6QorFNpSX/+INjjJqzAiKI41CLGQxmzKLWGDsKbdSs3os12mvqAq1hONAzQ9AzqjIibtcKXbAXWfUbsymhroZTgrKYkqmCPYJWlySHZI8HHDvd7W57/gU0fj3EX1g3bJ6ceTzkS18gh8UX/lsLLxxXC+Z16dYLv+ThTaZvCZ9emEYwoEIBr9y+g4AAP//AwBQSwMEFAAGAAgAAAAhAD4Ms3q3AwAAeQkAAA8AAAB4bC93b3JrYm9vay54bWysVm1vozgQ/n7S/QfEdxcbDAHUdBXedJXaVZVm2zup0soFp0EBnDNOk6ra/75jAknbnE657kUJxvbw+JmZZ4acf9nWlfHMZVuKZmySM2wavMlFUTZPY/PbLEO+abSKNQWrRMPH5gtvzS8Xv/92vhFy+SjE0gCAph2bC6VWoWW1+YLXrD0TK97AzlzImimYyierXUnOinbBuaory8bYs2pWNuYOIZSnYIj5vMx5IvJ1zRu1A5G8Ygrot4ty1Q5odX4KXM3kcr1CuahXAPFYVqV66UBNo87Dy6dGSPZYgdtb4hpbCV8PfgTDxR5Ogq2jo+oyl6IVc3UG0NaO9JH/BFuEvAvB9jgGpyFRS/LnUudwz0p6n2Tl7bG8AxjBv4xGQFqdVkII3ifR3D0327w4n5cVv9tJ12Cr1VdW60xVplGxVqVFqXgxNkcwFRv+bkGuV9G6rGDXwcQhpnWxl/ONNAo+Z+tKzUDIA/zYtLHtYKwtQRiTSnHZMMVj0SjQYe/Xr2quw44XAhRuTPnf61JyKCzQF/gKV5aH7LG9YWphrGU1NuPw4VsL7j+81KyqyiV7SMSmqQTU2MMbcbLjSvgP8mS59tkCp3fEdvcfAwD8ZDhI8EZJA+4vkytIwy17hqRA6ou+Zi8h6v73V4c6MfGdBI2CNEEUTwLkByRCmZf4qR+7JKWTH+CF9MJcsLVa9InWmGOTQlaPtq7ZdtghOFyXxeH8V9x/kB4/XIa9H9pT3dLuSr5pD5LQU2N7XzaF2IxNRLSQX95PN93mfVmoBTTRgNpgslv7g5dPC2BMiAvy0X1DMxubry5xUpplPoojb4Jo4HrIx06E/HSSxK7tZI7ndoysN5S65gnUutFoOsHf6oZKoEvrUUcX7mWoz5CXRadta3gsZ1UOAtdDZxgQbAfaa75VV63qRtBWCfQIpGSEA4pw6riI+oGNfOrYKKaJnbqjNEkjoDcU9P/QAjuJh8NbRbNcMKlmkuVLeBdN+TxiLShp5xDwfUs2cv0IO0CRZiRDlAQYRZFHkZtkjjsiSZy62YGsdn/+yQbkW93TnKk1FKeuy24e6mvWr+4X57uFPk/vii6cJjru/dP/ZngL3lf8ROPs7kTD+Ov17PpE26t09v0+O9V4ch0lk9PtJ9Pp5K9Z+udwhPWPAbU+JDwhNMBOOkGOE1NER9kI+Rl2kUNHNHZplBI8OiS82uTPn8u3Ta1BkfHb/wh9M9L51+Bh/wfKaLnqt+CN0TXOjrim39XXHu3iJwAAAP//AwBQSwMEFAAGAAgAAAAhAJIHlOwEAQAAPwMAABoACAF4bC9fcmVscy93b3JrYm9vay54bWwucmVscyCiBAEooAABAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAKySy2rEMAxF94X+g9G+cTJ9UIZxZtFSmG2bfoBwlDhMYgdbfeTva1I6ycCQbrIxSML3Hom72393rfgkHxpnFWRJCoKsdmVjawXvxcvNI4jAaEtsnSUFAwXY59dXu1dqkeOnYJo+iKhigwLD3G+lDNpQhyFxPdk4qZzvkGPpa9mjPmJNcpOmD9LPNSA/0xSHUoE/lLcgiqGPzv9ru6pqND07/dGR5QsWMvDQxgVEgb4mVvBbJ5ER5GX7zZr2HM9Ck/tYyvHNlhiyNRm+nD8GQ8QTx6kV5DhZhLlfE0Zjq58MNnaCObWWLnK3aigMeirf2MfMz7Mxb//ByLPY5z8AAAD//wMAUEsDBBQABgAIAAAAIQDDG6yZDw8AAG1IAAAYAAAAeGwvd29ya3NoZWV0cy9zaGVldDEueG1spFxrb9vIFf1eoP9BIBaFg40t8SXJqq2FJVGSgc1mkTi724+KTMdCJNOV5Dxa9L/3zMy94uXMOKbpoq3iM/femXvmdUgOefbLt8269SXf7lbF3XkQnnSCVn63LK5Xd5/Ogw9X0+N+0NrtF3fXi3Vxl58H3/Nd8Mvw7387+1psP+9u83zfQoS73Xlwu9/fD9rt3fI23yx2J8V9foeSm2K7Wezx5/ZTe3e/zRfX2mmzbkedTre9WazuAhNhsK0To7i5WS3zSbF82OR3exNkm68Xe7R/d7u633G0zbJOuM1i+/nh/nhZbO4R4uNqvdp/10GD1mY5uPx0V2wXH9fI+1uYLJatb1v8N8L/Yq5G405Nm9VyW+yKm/0JIrdNm930T9un7cXyEMnNv1aYMGlv8y8r1YFlqKhZk8L0ECsqg8UNg3UPwRRd28HD6vo8+G+UJtHkojM97l10kuPk9AJj7LTbO06TbBpPo36/34n+FwzPrlfoYZVVa5vfnAejcHDZ7QTt4ZkeQH+s8q878e/WfvHxfb7Ol/sclYRBS43Pj0XxWRleAuog5E4bqJC7f+ugF+Hgr+kk7CT9tNdVsduH4PLfXNFUD+bft63r/GbxsN6/K77O89Wn270K3/pPvi34TzRAj4zB9fdJvltiqMLkJE5VHctijYD4/9ZmpeYchtrim2nz6np/q2Zh1D1N4zBKg9byYbcvNn9SAbkbR/SQdsTvV1N+etKLumn3CT90hvbDL/mFaP0PKkrIAb/kEHVq1YT265rwyzX1Trpx1H8qtS454pcd41o19sgRv+TYP4nr1IhlTjcVv2VT67B5So74Zccfdl/bdL8eZ5PFfjE82xZfW5j7GAe7+4VaSaMBgqlBhfRNt9QZVhQGo8EOEyKOZzQmmBFLVfNI+egGoM4d0C/Dzln7C2bAkizGxiLS40+5TGwgs4GpDcxsYG6AWM87tP7ABAamk0JygkHkn1KchXI7DzBID1mEVhZsoSahzsIAiHxwiaouGVuwy9QGZjYwN4BZTmRaaFmTtJTbeYABfWhjbKXFFoe0DNA/9FZmgNMDMDWAWY4UEzMDqIXgUE1SrWZOJhinB5P0YNKWmYLQJpkqNwxDDItDBV0rVTLRY8Z0oUFErgYQuRpA5EpB9Gaig8wJCZ2hiJnjyaR7WJifnJg8OFUg5Cb57Vm5HUwO/UhIUnakg0wNIrLz1NS3epJMQj3Vo05UslzpRoy5Jt2o3KxuPLVSJRPRjYSUeWSE6Hmke2lqEJGqDcwN4E49JWKttTWssaIoN2tFsRdGNjn0mQHCnuY2STud6DRMQqsHMrIqh+3UICI5G5gbwE0Om0WT5JQbugk/5XppL5gHm0N6BomUmvoyTNKkZ60SGbmIzAwiMqMYeraZ+UeI3l8qQ1BNmCbJaT87O2ttH5dGh/QI4vx6ccfq8IydRIIEiQw5jEiRILf7Qs/GX2dwaj87RXtrKI3KFFV954FYMclILJmEyIyMVyQzMpAnI58GqTHd1Mqv5E+5LIwJKhsycZDMQaYcqKIFrIE6Y6MUQ/lm+P7Dm6NZmAxmSfLqrH3j0UBzdpD7U/jIDhhaGgZpxYCMmjOXEVpwKZUX+eUZLhZ489DRjGE5V+2tkY2kUgitPWbCRlhLykj24sRGckAYURPL/Su0VvUZ+VWMImv+zNlIyojIWncu2cizIDRUUaFRL7ESvTfDycVVdjTvvn7z9rer+dE8+cdic//P8NXrkPs+6YahM5coBMnF/VaF0jTeLrb5dWCuU8cYRONEma70ReeH33/P3h1dZX9dHY3C5HUwwX+CV1zPn9nkt+z95OJf1Q1yws0td7+MoV6p4hgq5zKGsJ5Dcdl3c4KSUu9cMuQKnrChdtN+2CsjLymjMB2MErVegBTIDjDxM/pcTzNFtTU3xxwNo04NU5dqQy7WtC/Dq/mHdz4GjaaLJYMESQYJkgwaCF3IS+acm1MuQ5cMeRj0a0ZE/fHli7pKU2rDYdBwpnNVXJUrjt47x+xXh6vpu0vPWDMVV5giSDJFkGSKWiyZIkgyRZCHqYYCE+KqFlPOCkl+dZh6f3H14Z2HKwohRxVBkisDxZIrspJcEVRaXXJmHq4aStiQBOpTo8q+GGG/Wlx9+M3DFFUsmSJIMmWgClNkJZkiSI4qgjxMNdTDSgnXmX/WbjlmvzpMYb/xMGUqrsw/giRTBMkxRS2WTBEkmSLIZQoqt5G41n41Vir7qo/9KkwN5QYZdewN8uqDb3ukSJIyhgRlDAnKuAmCMoYEZQx5KGso1qGb6wwuWy2N2a/O4HpMS1CMClmmObEkiyBJFjVakkWQJIsgD1kNrwMicx3w1E5oq8Yx+71ANVCICld0m1JyRZDkitosuSJIrO/cQg9X1gWDujCocxWIi6VaA8u+9mW/xqqBAlSYoosEyRRBkilqsWSKIDmqCPIw1fAKIDLK+MlRZct+9nuBauAQYi9kSHJF0l1yRW2WXBEkuSLIw1VDNa+eINXYCyNbt7NfY9XAASRT1BbJlKva2VEyRY6SKYI8TDVU7erivQ5Ttmpnv8aqgQNIpqgtkilXtbOjZIoc5Ur1qGqPGqp27fe0ahA3o831DfvVWtX9asEV7RxUUuWKdraSVLmina08g6qhaI/qifbIFu3s9yK14Op2DivJcnU7W0myXN3OVh6yGur2qJ5uj2zdzn4vUQtUtZyDBEmuXOXOdUuuXOXOVi5X6hZbo0ekyq/GHLSVu64Pfo3VAgcQTDEkmCJIXg2ylWCKIbGuM+RhqqFgj+sJ9th+HsR+L1ALHEJyRa2RXLl6nR0lV65eZysPVw31elxPr8fO43jya6wWuGLJFMWUTLlqnR0lU+QoRxVBHqaUOm3wQFE9A6gx/2JbrbNfY7XAASRT1BbJlIEq84+sJFMESaYI8jDVUK3jRn0tpmy1zn7N1QJHkFRRYyRVrlhnR0mVK9bZykNVQ7GOB0W1qLLFOvu9RC1wDEmWq9fJqjKu3LvsHEuOq0f1Oq47m83Aeno9tvW6rq/mDvjIEwkOIblyFTtZVbhy77NzLMnVo4odt3macVXvPnts32fX9b1ILbiCnWPKKegKdraSU5BiSaYI8kzBhoIdvVVrCtqCnf1eohZcvc5RJVeuXmcryZWr19nKw1VDvY5Hk7W4svU6+zVXC65a55iSKVets5VkylXrbOUyhSevzU7+1VPrsa3WdX0155/3iQQHECsVQ4IpguRKxVaCKYbE/GPIw1RDta4fLT99XZPYap39mqsFjiCpcsU6WVWocm+ucyxJ1aM313Efs9mgkmJ9eDk9wgOpo+ytOQnx0+gnnFHovHp1HvVfB8FrPLP/WZ6JSGwJr1tRc6g99nyCY0gKXRVPVhUK3XvuHEtS+KiKV5ezjU7kShWvKBwleN8iUIQ9zuep4TMCn+KQSWIrfd2ml+kMDiH5dLU+WVX4dLU+x5J8Pqr19amXBldF2g85Y00tT3sYXuOneI3xxFCN07jkVb3GcTi4TrFTOrj0o5Mk0m/CfnqN0jciM4bKU/BThsoTazg/pre3VC6BBGkSK8cbk4aSX/udB2nZxWOGygV64kIZQ9joyyPv1hiclkZ8AmZGEO6y1zwsxzH0OaRqzg2Ve2L0rXkHyNwaJqgcnBMHyRxk6iAzB5kTgud1alRUTvE3FNOJEZ2V5htINt9GMnIrbaYOMnOQOSG+5nsUbhI//XJFYqRhF/tROXCcFxHISLwnwn7lDMlcaOpCMxeaM6SnVrVTpBZFI+o+PVVLDu7HVDqFINkrBMkF1YGmHKu0mrnQnCE9TasvTUiV+Iwc1NJm5cBQ+TxnwlB5LjBzoakLzVxoTlBPHyis5iD123NyMDJH9kNqINENDpI5yNRBZg4yJ8QzOVKpqcrmN3jdQ0fCWzuV6WLfgGEjMV0YEtOFoXJDmbqOM9dxXnGsdpMUPi/L04gBeWQ7NVBP9pwDZa7VlCGs8+USY92HmZVGh9OZDHmmlLz5+JzhaDbsXmWXdF5FIiOZpwNl0AFqdgo2pgSVfjMHmRPiG6NSMjwnJ3MrrTLFDFTpKQfKUgeaEiQzMEYlMicbXwZSADwnA1cAQABpdkk5mveZ7DPvbCSWb4bE1SxBMicTXOZkEF9OUhU8Jyez4ffLs9njlKBSeU5cKHOhqQvNXGjOkKtsUikNKAck/Ow3UNNnbsaYyPTuKgZneUFWg0WlzcSLtJt8+ykf5+v1rrUsHtQr11jqh2cHmA7op91BZpSzVTJJ4kGGKx6EdUoSlOgV2I6W9AdqQ3R9xmmEejTNlg/Wb/j4SsYpWoCZ40bD6ggfb9vi3iAzJ/nsVsd9lGgR4JScokRrASfTDjL15TNJQpToUen4IFNceXt4i+GDB46+EvjgMZyvBBzgeZWvBL2A5ye+khQlPt4mMXrbHFKzWx3BBxc0nmgRfHBMxlcCrnHew1cCrnG4wVcCrvEo39dqcI1H156SED44WOyLBh+cT/WVgGvzfpaTKbg2i5ZTAq7N5u2UgGscUvO1DT54QcdXAh+8WuIrAdd4ZcJXAq7xCoOvBFzjQLyvBFzjgLpbMuoNxl4PhPLho/5g7I1zOhj72B+FnYF6bdBTcxiixDvSUZJ5S0ZYBnwzYBQPxv5VCNPMV3c6GHu5BeleHJRrxtvliqk+SHF3vVKfUVmszQtqe3wDhj+UgReSBrgn0prid46FEOY37x7WeWv//R5fR8m/4dMuO/MJkutvN+Z7G6377arY4oMq+pWfM/UlmIf1Yvj23dFPeEnqPOBXLILX9Ld+jSBQd8fIFPdxdC3qho6vdfhICL6g8MdivcKv+gAMr/tq2a8WUUOv8+Vqs1gHrcV6XXwdrRd3n82XF26Lr5d39w/7N0hj8QkpqcsPgNl2W2wlmCvgarVXH4UZXf7668Xo16w1f4t3kPDlHFUGeLVeq6/GtG6Lh+2utcEnPFof81b/JG0V29YaNZwguPkCCb1zKL6P0u+k0WTU6x73o3H/OMEXUY5xe+viOBn38a7BRSfsjrrq+yhEUqi+1XD4N6PREA8FCI1AX5UMB8D3SO6R9ZvF9tMKLK7zG2ye+MYH3mmjz5uof++Le41C630s9vgsCf91i8/65EgcXzgJWjdFsec/ME/ahw8FDf8PAAD//wMAUEsDBBQABgAIAAAAIQD2YLRBuAcAABEiAAATAAAAeGwvdGhlbWUvdGhlbWUxLnhtbOxazY8btxW/B8j/QMxd1szoe2E50Kc39u564ZVd5EhJlIZeznBAUrsrFAEK59RLgQJp0UuB3nooigZogAa55I8xYCNN/4g8ckaa4YqKvf5AkmJ3LzPU7z3+5r3HxzePc/eTq5ihCyIk5UnXC+74HiLJjM9psux6TybjSttDUuFkjhlPSNdbE+l9cu/jj+7iAxWRmCCQT+QB7nqRUulBtSpnMIzlHZ6SBH5bcBFjBbdiWZ0LfAl6Y1YNfb9ZjTFNPJTgGNQ+WizojKCJVund2ygfMbhNlNQDMybOtGpiSRjs/DzQCLmWAybQBWZdD+aZ88sJuVIeYlgq+KHr+ebPq967W8UHuRBTe2RLcmPzl8vlAvPz0MwpltPtpP4obNeDrX4DYGoXN2rr/60+A8CzGTxpxqWsM2g0/XaYY0ug7NKhu9MKaja+pL+2wznoNPth3dJvQJn++u4zjjujYcPCG1CGb+zge37Y79QsvAFl+OYOvj7qtcKRhTegiNHkfBfdbLXbzRy9hSw4O3TCO82m3xrm8AIF0bCNLj3FgidqX6zF+BkXYwBoIMOKJkitU7LAM4jiXqq4REMqU4bXHkpxwiUM+2EQQOjV/XD7byyODwguSWtewETuDGk+SM4ETVXXewBavRLk5TffvHj+9Yvn/3nxxRcvnv8LHdFlpDJVltwhTpZluR/+/sf//fV36L///tsPX/7JjZdl/Kt//v7Vt9/9lHpYaoUpXv75q1dff/XyL3/4/h9fOrT3BJ6W4RMaE4lOyCV6zGN4QGMKmz+ZiptJTCJMLQkcgW6H6pGKLODJGjMXrk9sEz4VkGVcwPurZxbXs0isFHXM/DCKLeAx56zPhdMAD/VcJQtPVsnSPblYlXGPMb5wzT3AieXg0SqF9EpdKgcRsWieMpwovCQJUUj/xs8JcTzdZ5Radj2mM8ElXyj0GUV9TJ0mmdCpFUiF0CGNwS9rF0FwtWWb46eoz5nrqYfkwkbCssDMQX5CmGXG+3ilcOxSOcExKxv8CKvIRfJsLWZl3Egq8PSSMI5GcyKlS+aRgOctOf0hhsTmdPsxW8c2Uih67tJ5hDkvI4f8fBDhOHVypklUxn4qzyFEMTrlygU/5vYK0ffgB5zsdfdTSix3vz4RPIEEV6ZUBIj+ZSUcvrxPuL0e12yBiSvL9ERsZdeeoM7o6K+WVmgfEcLwJZ4Tgp586mDQ56ll84L0gwiyyiFxBdYDbMeqvk+IhDJJ1zW7KfKISitkz8iS7+FzvL6WeNY4ibHYp/kEvG6F7lTAYnRQeMRm52XgCYXyD+LFaZRHEnSUgnu0T+tphK29S99Ld7yuheW/N1ljsC6f3XRdggy5sQwk9je2zQQza4IiYCaYoiNXugURy/2FiN5XjdjKKbewF23hBiiMrHonpsnrip8TLAS//Hlqnw9W9bgVv0u9sy+vHF6rcvbhfoW1zRCvklMC28lu4rotbW5LG+//vrTZt5ZvC5rbgua2oHG9gn2QgqaoYaC8KVo9pvET7+37LChjZ2rNyJE0rR8JrzXzMQyanpRpTG77gGkEl/p5YAILtxTYyCDB1W+ois4inEJ/KDBdzKXMVS8lSrmEtpEZNv1Uck23aT6t4mM+z9qdpr/kZyaUWBXjfgMaT9k4tKpUhm628kHNb0PdsF2aVuuGgJa9CYnSZDaJmoNEazP4GhK6c/Z+WHQcLNpa/cZVO6YAaluvwHs3grf1rteoZ4ygIwc1+lz7KXP1xrvaOe/V0/uMycoRAK3FXU93NNe9j6efLgu1N/C0RcI4JQsrm4TxlSnwZARvw3l0lvvuPxVwN/V1p3CpRU+bYrMaChqt9ofwtU4i13IDS8qZgiXoEtZ4CIvOQzOcdr0F9I3hMk4heKR+98JsCYcvMyWyFf82qSUVUg2xjDKLm6yT+SemigjEaNz19PNvw4ElJolk5DqwdH+p5EK94H5p5MDrtpfJYkFmquz30oi2dHYLKT5LFs5fjfjbg7UkX4G7z6L5JZqylXiMIcQarUB7d04lHB8EmavnFM7DtpmsiL9rO1Oe/a1DriIfY5ZGON9Sytk8g5sNZUvH3G1tULrLnxkMumvC6VLvsO+87b5+r9aWK/bHTrFpWmlFb5vubPrhdvkSq2IXtVhluft6zu1skh0EqnObePe9v0StmMyiphnv5mGdtPNRm9p7rAhKu09zj922m4TTEm+79YPc9ajVO8SmsDSBbw7Oy2fbfPoMkscQThFXLDvtZgncmdIyPRXGt1M+X+eXTGaJJvO5LkqzVP6YLBCdX3W90FU55ofHeTXAEkCbmhdW2FbQWe3Zgnqzy0WzBbsVzsrYa/WqLbyV2ByzboVNa9FFW11tTtR1rW5m1g7LntqkYWMpuNq1IrTJBYbSOTvMzXIv5JkrlVfacIVWgna93/qNXn0QNgYVv90YVeq1ul9pN3q1Sq/RqAWjRuAP++HnQE9FcdDIvnwYw2kQW+ffP5jxnW8g4s2B150Zj6vcfONQNd4330AE4f5vIMCRQCscBfWwFw4qg2HQrNTDYbPSbtV6lUHYHIY92LSb497nHrow4KA/HI7HjbDSHACu7vcalV6/Nqg026N+OA5G9aEP4Hz7uYK3GJ1zc1vApeF170cAAAD//wMAUEsDBBQABgAIAAAAIQBowXh5+wcAADNSAAANAAAAeGwvc3R5bGVzLnhtbOxcWY/bNhB+L9D/IAh99eqw5WNrO83uxkCANAiQLVCg6YNWlr1CdBgSvbET9L93qMMiLVOiTmvR7kNiyTL5zT2coTh/c3Bs4cX0A8tzF6JyI4uC6Rre2nK3C/GPx9VgKgoB0t21bnuuuRCPZiC+Wf780zxAR9v8/GyaSIAh3GAhPiO0u5WkwHg2HT248XamC99sPN/REVz6WynY+aa+DvCPHFtSZXksObrlitEIt47BM4ij+1/3u4HhOTsdWU+WbaFjOJYoOMbt+63r+fqTDVAPykg3hIMy9lXh4CeThHcz8ziW4XuBt0E3MK7kbTaWYWbhzqSZpBvpSDBytZEUTZJVivaDX3GkkeSbLxYWn7icu3tn5aBAMLy9i0Ccp1tC9M37Ndwcj0Qhksq9twY+/fXLQJHh7+/1F8FxnC/CEf5+/U2UlnMpHnA533huOu4YWICZe/vV9b65K/xVNBl+ajkPvgsvug13FDyG4dmeLyBQCpgrvOPqjhk98XaHvED4qPu+9w0/u9Edyz5G36n4RqhM8cOOBaINUUXTRP8+4aeuNmEIkqJQxggbpTChbtYBN0+cDMnoUnTtTiiF+gsKZtn2yTwmYB74xnIOngSZvruCCyH+/Hjcgb664PQilQufK3h66+tHRdX4fxB4trXGKLb3GSt5iu9Z7to8mGC4YLfYJAmscBUhKsDFmAZ8PbKwnxjIN+poNptMFPw3mc6GeKZuASgAYDYdTUbyZKSpYzX0FF1A6JzSM0LLsjoUeQDi8fw1xO3E1w9l0KPo3nJumxsEdPnW9hn/j7wdptJDCILbcr629K3n6jbWpuQX5C8h4ENsX4joGWJz4sDP1RBPQc3A9StAUgIITUKLE3ANHXIzZGYLSMryvGXEXMNHClWsT4kYS2Lun8r2gStlda9RzD3T0tZ8WZuqV9UKrirIsyDSK/fQPw2PFztkfnUh6EtpQDnp20VaOIfjdcd15iCifp1hXjflkTE04ogbFm1upC2tlX2ksFfKWwcMkQm/dqM/Xxr0ybpzli3NOLD/dSCsXLB14L/HoJz4wOmDO3YOHYSNfhLeyHqmmbhankFVVz+cMzWe6jXkCLoju307PveajVRUrqWP7a/DGtLc1mMVh4rSpc389V9dLWlwvVIGSlxbhlK1Ydr2Z1xT/nND9SYPG6IvCatl3KTBLUr8Efot8ceoNB1dACkSOVo0NjHsdAgV6/LjCofNaQLWr9UUFXxMUSnQDY1/Lei7nX3EDVHc6oyvgJL06i6s2qfXb21r6zpm0q3Vk0vh2fOt7zAQbqRiVxz2fg4bNnEseHC/z/CG/YBHqBwgSoULvfIz7nUkTgIQYCBtgAEITIPS3LjPH1kU0JRYFDWadk7ex73zZPqrcNMGocNNa7QBGm9Gzfw8na4tlVymEKMDH1IW4+0NlEEzmMKvCZEn+OR7yDRQtNMG3FSOseMtOcgysPHDWgDb/u70Y8H2jK+4IRy2zKUM/yjBs2icvG4aLweL3lkuwMxxJbmxoLTKlNIRDv5lrADvjugikLGMPqOyVwBECRS2xtF+oh8Cxdu74uwp361VzlPK6Gbq6fk9GKGdLGKyylA1RHVIDCumt5FklI+wVGLQBqSwLlyYxxKpSoE/7y7NZoXUfJbVThtYKUKebLnSPlBErgwnh8NcC5eaSpjJTlke/zzNuxzeKGnNeBP7CmNn2Evj5mIdn7HAIvNiXp813ytl8CDCBCAkFMSiOcujlhASqTSnOl4Rc7jSbxBx+7WJmKcNYm6/YNE85itqs/DN13eP5iEsOeH6WN6CmqXa7fiLdBGbcLwMViL1o9LuAl4zvWsGTRGziHQtH0DNKblCtpIplPWrVMOqMCkVVSsn7BJTUXJRsvkCFVPOl2fn9ZhKCxQmmOz6+YpgCpZLnXKmJBbKuDpcrDEFm1n20+voVnhZ2hdWqgYrmaT4lfiY5qoBuVXZooDBVJlXrvJqQcjt1H3Am0G5Da1uwRQkq82BwXE06Y+y9CwrJx77bRBj9GppfstJfaWJjNr+mpqrz8sUfp+yH7VP2U+R5Lr1GAVq1FgPoWbhnalmnaWSXEXXTDAoV3StUNzLenxGXyxnbFbZrmLpowIV2W0XnFRwSYU3j+zRnNHKgrmSL1O4IOoGVAMum7jU76iWr7BTkLiVuYA95G6lmpC4NbM7SNy9gu4gNdFiqCko7u5Md1yBE0ea365V18Sy9YKrm73SD7uv1pAroU58TTSWv66zA+Ya23doz17HaV0dfZ29PlcHX2fvRA92o6hthDuupS1rf0y3O7+6yHDb9GGUV+3DLj4KEH+IrrI24461FQbnD5q4TozfUoD3EoiXH6hXH04vMQj4rLCF+BHvuLaJ7OVpb9lwOtOF1x5gzPWBfpECri8cZZUe4XThxRl83tJ4MtM0bazJmhydt3TxcCkpHB3/CzMjfLZe+C7HiRyw2rW50fc2ejx9uRDTz7+ba2vvQMYRP/XJevFQOMRCTD9/wGcVKWO8PQx62R8COFwI/hf2vrUQf7y7m8we3q3UwVS+mw5GQ1MbzLS7h4E2ur97eFjNZFW+/4c44a/G+X7hgYSwlV4Z3QY2nALox8TG4D+n9xYicRHBDze3AWwS+0wdy281RR6shrIyGI316WA6HmqDlaaoD+PR3TttpRHYtYrnAMqSokQnCmLw2i2yHNO23ERWiYTIuyAkuMwhQkokIaWnPS7/BQAA//8DAFBLAwQUAAYACAAAACEAjqCrVJIBAACqAwAAFAAAAHhsL3NoYXJlZFN0cmluZ3MueG1sfJNRT+QgFIXfN9n/cNN3p6MPZmPaGmxxprEtDUWNjziDMyQFukDN+u+3xs0mwugj59zvcHIJ2fUfNcKrsE4anSfnq3UCQu/MXupDntyz27NfCTjP9Z6PRos8eRMuuS5+/sic87Cw2uXJ0fvpKk3d7igUdyszCb04L8Yq7pejPaRusoLv3VEIr8b0Yr2+TBWXOoGdmbXPk4vlllnL37Mo/wtF5mSR+YLVLR62GDM4g4FRxPCmLmEg97Ssu02W+iJL3yc/pntEWYdpKJeo3+C27moYFgj1hGLoGYaGVV8kAG77FXSoxeFASzq2DUU0WTlGt9bV1bfxlDRR/BNGUf3KvK8riiLQzepZ2BPG4Ln1UHEv4gYEsN6f9gby+C8zwm7qpkE3DYbtsvoBIn9rZuvCJtXyXLH2FG2vZPVDzSL9kdA7uKWkDYHPZUKX4hbRu1BlhKEGvidbo/0R+M7LV+mlcOBmpbh9C7M6ruLFlnw6CCW1BCrcZLSTz6MIyUEeNPezjXG8QafBdPlrxV8AAAD//wMAUEsDBBQABgAIAAAAIQDfLqcEcwEAANIFAAAQAAAAeGwvY2FsY0NoYWluLnhtbHSU627CIBiG/y/ZPRD+T+Qwd4jVBJJ5A9sFkIq2SUtNaZbt7scW+Zgc/phIH5/yva+w3X+NA/o0s+sn22C6WmNkbDsde3tu8Mf728MzRm7R9qiHyZoGfxuH97v7u22rh1Z1urfIG6xrcLcsl1dCXNuZUbvVdDHWPzlN86gX/3U+E3eZjT66zphlHAhbrzdk9AK827ZobvBBPGLU+01gNPx+kuu6pOK6HleA9C+O5IGyhFS05txkZP6Wp+J+FM1/WyYl9en9TRR3/lJxZiTzVRTSUDQYwMl8YiUSDJEM+dwmrMAQSMXKpGQ8nYiF3FJnSqoKKVnoKO4zJJw6c7LSEUs7kiwknDqDIc5eJiXLkueVjgpkpSMwwOy80hEPhkiGhJOJwAAT8TIpefaf5+XzogpkpSMwxLeXScnT5CWvdFQgK+cIDJCSqHTEszZFpSMwwEQVUors/oGV5E4T2TkSlXMU12/uOvXfQOA23v0AAAD//wMAUEsDBBQABgAIAAAAIQAwIYPQ9gAAAEUBAAAZAAAAZG9jTWV0YWRhdGEvTGFiZWxJbmZvLnhtbFTQ3WqEMBAF4FcJuY+JrjGuqAu9K2yhrxCTcQ3kZzFTaSl998a79u4wMB9zZrx9Bk8O2LNLcaJ1JSiBaJJ18THRD1xZT0lGHa32KcJEvyDT2zwav/jB6wX83WUkBYl5OIcT3RCfA+fZbBB0roIze8ppxcqkwNO6OgO8EY3gwT3vp/AGqK1GTf+yxNmJfoPqhGmvLeukFKxtl5b1jeyZUqZW0irVL+rnvFgvHspCTUkA3FKJ77s7nIcH2FLAIbyenup0o0Fqdl1F8dZuYVcJlkl7uRTNtLari2dSRIj44jBPtHxkh5CO0y+ZzyP/337+BQAA//8DAFBLAwQUAAYACAAAACEA0B2j6k4BAAB/AgAAEQAIAWRvY1Byb3BzL2NvcmUueG1sIKIEASigAAEAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAjJJfS8MwFMXfBb9DybNt2k2nlLYDlT05EJz45+2S3G1xTRqSzG7f3rTdamUKPibn3F/OuSSb7mQZfKKxolI5SaKYBKhYxYVa5eR5MQtvSGAdKA5lpTAne7RkWpyfZUynrDL4aCqNxgm0gScpmzKdk7VzOqXUsjVKsJF3KC8uKyPB+aNZUQ1sAyukozieUIkOODigDTDUPZEckJz1SL01ZQvgjGKJEpWzNIkS+u11aKT9daBVBk4p3F77Toe4QzZnndi7d1b0xrquo3rcxvD5E/o6f3hqq4ZCNbtiSIqMs5QZBFeZYg5lKTZgPrYKLoI3kBpK2GwzOvA0+yzBurlf/VIgv93/PXZq9a+15bonkQc+btqVOyov47v7xYwUo3g0CeObMIkXyXUaX6WX4/cmyY/5Jn53IQ95/k+ceOiAeAQUGT35MsUXAAAA//8DAFBLAwQUAAYACAAAACEAb3gbEpEBAAAaAwAAEAAIAWRvY1Byb3BzL2FwcC54bWwgogQBKKAAAQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACckkFv2zAMhe8D9h8M3Rs53VAMgayiSDf0sGEBkrZnVqZjobIkiKyR7NdPtpHU2XbajXx8ePpESd0eOlf0mMgGX4nlohQFehNq6/eVeNx9u/oiCmLwNbjgsRJHJHGrP35QmxQiJrZIRY7wVImWOa6kJNNiB7TIY58nTUgdcG7TXoamsQbvg3nr0LO8LssbiQdGX2N9Fc+BYkpc9fy/oXUwAx897Y4xA2t1F6OzBjjfUv+wJgUKDRdfDwadkvOhynRbNG/J8lGXSs5btTXgcJ2DdQOOUMl3QT0gDEvbgE2kVc+rHg2HVJD9ldd2LYoXIBxwKtFDsuA5Yw22qRlrF4mTfg7plVpEJiWzYRLHcu6d1/azXo6GXFwah4AJJA8uEXeWHdLPZgOJ/0G8nBOPDBPvhLMd+KYz53zjlfNJf2SvQxfBH/Ua4h47662SJ0l9t/6VHuMu3APjaa+Xotq2kLDOT3He+1lQD3mlyQ0h6xb8HuuT5+/B8Auepq+ulzeL8lOZH3imKfn+qfVvAAAA//8DAFBLAQItABQABgAIAAAAIQCVZpbtkAEAAIIFAAATAAAAAAAAAAAAAAAAAAAAAABbQ29udGVudF9UeXBlc10ueG1sUEsBAi0AFAAGAAgAAAAhADEdic0iAQAA3gIAAAsAAAAAAAAAAAAAAAAAyQMAAF9yZWxzLy5yZWxzUEsBAi0AFAAGAAgAAAAhAD4Ms3q3AwAAeQkAAA8AAAAAAAAAAAAAAAAAHAcAAHhsL3dvcmtib29rLnhtbFBLAQItABQABgAIAAAAIQCSB5TsBAEAAD8DAAAaAAAAAAAAAAAAAAAAAAALAAB4bC9fcmVscy93b3JrYm9vay54bWwucmVsc1BLAQItABQABgAIAAAAIQDDG6yZDw8AAG1IAAAYAAAAAAAAAAAAAAAAAEQNAAB4bC93b3Jrc2hlZXRzL3NoZWV0MS54bWxQSwECLQAUAAYACAAAACEA9mC0QbgHAAARIgAAEwAAAAAAAAAAAAAAAACJHAAAeGwvdGhlbWUvdGhlbWUxLnhtbFBLAQItABQABgAIAAAAIQBowXh5+wcAADNSAAANAAAAAAAAAAAAAAAAAHIkAAB4bC9zdHlsZXMueG1sUEsBAi0AFAAGAAgAAAAhAI6gq1SSAQAAqgMAABQAAAAAAAAAAAAAAAAAmCwAAHhsL3NoYXJlZFN0cmluZ3MueG1sUEsBAi0AFAAGAAgAAAAhAN8upwRzAQAA0gUAABAAAAAAAAAAAAAAAAAAXC4AAHhsL2NhbGNDaGFpbi54bWxQSwECLQAUAAYACAAAACEAMCGD0PYAAABFAQAAGQAAAAAAAAAAAAAAAAD9LwAAZG9jTWV0YWRhdGEvTGFiZWxJbmZvLnhtbFBLAQItABQABgAIAAAAIQDQHaPqTgEAAH8CAAARAAAAAAAAAAAAAAAAACoxAABkb2NQcm9wcy9jb3JlLnhtbFBLAQItABQABgAIAAAAIQBveBsSkQEAABoDAAAQAAAAAAAAAAAAAAAAAK8zAABkb2NQcm9wcy9hcHAueG1sUEsFBgAAAAAMAAwABQMAAHY2AAAAAA=="""

# # def get_default_template_bytes():
# #     """Decode the hardcoded default template back into raw .xlsx bytes."""
# #     return base64.b64decode(DEFAULT_TEMPLATE_B64.strip())


# # # =====================================================================
# # # 2. HELPER FUNCTIONS
# # # =====================================================================

# # def norm_name(name):
# #     """Normalize a name to a comparable token set regardless of 'First Last'
# #     vs 'Last, First' formatting differences between source files."""
# #     parts = re.split(r"[,\s]+", str(name).strip().lower())
# #     parts = [p for p in parts if p]
# #     return " ".join(sorted(parts))


# # def name_tokens(name):
# #     parts = re.split(r"[,\s]+", str(name).strip().lower())
# #     return {p for p in parts if p}


# # def find_matching_rep_key(sow_name, rep_name_keys):
# #     """Match a SOW resource name to a Replicon name key even when one side
# #     has an extra middle name (e.g. SOW 'Kalyani Ghaytadkar' vs Replicon
# #     'Ghaytadkar, Kalyani Popatrao'). A match is any rep key whose token set
# #     fully contains the SOW name's tokens, or vice versa."""
# #     sow_tok = name_tokens(sow_name)
# #     best_match = None
# #     for rep_key in rep_name_keys:
# #         rep_tok = set(rep_key.split())
# #         if sow_tok.issubset(rep_tok) or rep_tok.issubset(sow_tok):
# #             # Prefer the closest-length match if multiple candidates exist
# #             if best_match is None or abs(len(rep_tok) - len(sow_tok)) < abs(len(set(best_match.split())) - len(sow_tok)):
# #                 best_match = rep_key
# #     return best_match

# # def get_field(row, candidates, default=""):
# #     """
# #     Flexible lookup of a value from a SOW row, regardless of small
# #     header-naming differences (spacing, trailing colon, casing) between
# #     what the code expects and what the uploaded SOW file actually has.

# #     `candidates` is a list of possible header names to try, in priority
# #     order. Returns `default` if none of them are found or all are blank.
# #     """
# #     def normalize(s):
# #         return re.sub(r"\s+", " ", str(s).strip().lower().rstrip(":"))

# #     normalized_map = {normalize(col): col for col in row.index}

# #     for cand in candidates:
# #         key = normalize(cand)
# #         if key in normalized_map:
# #             val = row[normalized_map[key]]
# #             if pd.notna(val) and str(val).strip() != "":
# #                 return val
# #     return default


# # # def load_sow(file):
# # #     """Load the SOW & Resource Details file into a DataFrame."""
# # #     df = pd.read_excel(file)
# # #     df["name_key"] = df["Resource name"].apply(norm_name)
# # #     return df
# # def load_sow(file):
# #     df = pd.read_excel(file)
# #     name_col = None
# #     for candidate in ["Resource name", "Employee Name", "Name"]:
# #         for col in df.columns:
# #             if re.sub(r"\s+", " ", str(col).strip().lower()) == candidate.lower():
# #                 name_col = col
# #                 break
# #         if name_col:
# #             break
# #     if name_col is None:
# #         raise ValueError("Could not find a resource/employee name column in the SOW file.")
# #     df["name_key"] = df[name_col].apply(norm_name)
# #     return df

# # # def load_replicon(file):
# # #     """Load the Replicon dump (Expenditure Details Report) into a DataFrame.
# # #     Row 1 is an instructions banner, row 2 holds the real headers."""
# # #     df = pd.read_excel(file, sheet_name="Expenditure Details Report", header=1)
# # #     df = df.dropna(subset=["Employee Name/Supplier Name"])
# # #     df["name_key"] = df["Employee Name/Supplier Name"].apply(norm_name)
# # #     df["Item Date"] = pd.to_datetime(df["Item Date"], format="%d-%b-%Y")
# # #     df["Quantity"] = pd.to_numeric(df["Quantity"], errors="coerce").fillna(0)
# # #     return df

# # def load_replicon(file):
# #     """Load the Replicon dump into a DataFrame. Sheet is 'Sheet1' with
# #     headers on row 1. Employee name, date, and hours columns are renamed
# #     to the internal standard names used throughout the rest of the app."""
# #     df = pd.read_excel(file, sheet_name="Sheet1")
# #     df = df.dropna(subset=["Employee Name"])
# #     df["name_key"] = df["Employee Name"].apply(norm_name)
# #     df["Item Date"] = pd.to_datetime(df["Expenditure Item Date"])
# #     df["Quantity"] = pd.to_numeric(df["QTT"], errors="coerce").fillna(0)
# #     return df


# # def build_daily_hours(rep_df):
# #     """Sum hours per employee per date (nets out negative correction rows),
# #     and track which (employee, date) pairs had a correction so we can flag
# #     them in the Remark column."""
# #     daily = rep_df.groupby(["name_key", "Item Date"])["Quantity"].sum().reset_index()
# #     neg = rep_df[rep_df["Quantity"] < 0][["name_key", "Item Date"]]
# #     neg_set = set(zip(neg["name_key"], neg["Item Date"]))
# #     return daily, neg_set


# # def determine_work_from(role):
# #     """Determine WFH vs WFO from the Partner Emp. Role string.

# #     Rule (per business input): the role code contains a segment like
# #     'T&M_13' — the number right after 'T&M_'. Only the LAST digit of that
# #     number matters (the leading digit, e.g. the '1' in '13', is ignored):
# #         - last digit 3  -> WFO (work from office)
# #         - last digit 2  -> WFH (work from home)
# #     Falls back to a keyword check ("onsite" -> WFO) and then defaults to
# #     WFH if the pattern can't be parsed, so nothing breaks on unexpected
# #     role formats.
# #     """
# #     role_str = str(role)

# #     # Explicit keyword override, if present
# #     if "onsite" in role_str.lower():
# #         return "WFO"

# #     match = re.search(r"T&M[_\s]*(\d+)", role_str)
# #     if match:
# #         last_digit = match.group(1)[-1]
# #         if last_digit == "3":
# #             return "WFO"
# #         elif last_digit == "2":
# #             return "WFH"

# #     # Unrecognized pattern - default to WFH, but this should be reviewed
# #     return "WFH"

# # def fill_timesheet_for_employee(template_bytes, res_row, daily_hours, neg_set, month, year, matched_rep_key):
# #     """Fill one copy of the timesheet template for a single employee and
# #     return it as an in-memory .xlsx (BytesIO)."""
# #     wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
# #     ws = wb["Sheet1"]

# #     name_key = matched_rep_key  # the Replicon-side key matched to this SOW resource

# #     # ---- Pull all resource/PO/SOW fields from the SOW & Resource Details
# #     # ---- file, tolerant of header-name variations in that file.
# #     emp_name = get_field(res_row, ["Resource name", "Employee Name", "Name"])
# #     cid = get_field(res_row, ["CID", "C.I.D", "C.I.D.", "Cid"])
# #     role = get_field(res_row, ["Partner Emp. role", "Partner Emp Role", "PARTNER EMP. ROLE", "Emp Role", "Role"])
# #     domain = get_field(res_row, ["Domain"])
# #     po_number = get_field(res_row, ["PO Number", "PO Number:", "PO No", "PO No."])
# #     po_start = get_field(res_row, ["PO Start Date", "PO Start Date:", "PO Start"])
# #     po_end = get_field(res_row, ["PO End Date", "PO End Date:", "PO End"])
# #     sow_number = get_field(res_row, ["SOW number", "SOW Number", "SOW Number:", "SOW No"])

# #     work_from_value = determine_work_from(role)

# #     # ---- Header fields ----
# #     ws["D4"] = emp_name
# #     ws["D5"] = cid
# #     ws["D6"] = role
# #     ws["D7"] = domain
# #     if po_number != "":
# #         ws["D8"] = po_number
# #     if po_start != "":
# #         ws["D9"] = po_start
# #     if po_end != "":
# #         ws["D10"] = po_end
# #     ws["D11"] = sow_number
# #     ws["H4"] = calendar.month_name[month]
# #     ws["H6"] = year

# #     # ---- Daily rows (row 14 = day 1) ----
# #     emp_hours = daily_hours[daily_hours["name_key"] == name_key].set_index("Item Date")["Quantity"]
# #     days_in_month = calendar.monthrange(year, month)[1]
# #     date_number_format = ws["B14"].number_format  # capture template's date format before overwriting

# #     proj_start_raw = get_field(res_row, ["Start date", "Start Date", "Project Start Date"], None)
# #     proj_end_raw = get_field(res_row, ["End date", "End Date", "Project End Date"], None)

# #     month_start = datetime(year, month, 1)
# #     month_end = datetime(year, month, days_in_month)

# #     proj_start = pd.to_datetime(proj_start_raw) if pd.notna(proj_start_raw) else month_start
# #     proj_end = pd.to_datetime(proj_end_raw) if pd.notna(proj_end_raw) else month_end

# #     effective_start = max(proj_start, month_start)
# #     effective_end = min(proj_end, month_end)


    
# #     for day in range(1, days_in_month + 1):
# #         row = 13 + day
# #         this_date = datetime(year, month, day)

# #         if this_date < effective_start or this_date > effective_end:
# #             for col in range(2, 9):
# #                 ws.cell(row=row, column=col, value=None)
# #             continue


        
# #         hours = emp_hours.get(this_date, 0)
# #         remark = "Correction adjusted" if (name_key, this_date) in neg_set else ""
# #         is_weekend = this_date.weekday() >= 5  # 5 = Saturday, 6 = Sunday

# #         date_cell = ws.cell(row=row, column=2, value=this_date)
# #         date_cell.number_format = date_number_format
# #         ws.cell(row=row, column=3, value=this_date.strftime("%A").upper())

# #         if is_weekend and hours == 0:
# #             ws.cell(row=row, column=7, value=None)   # blank instead of 0
# #         else:
# #             ws.cell(row=row, column=7, value=float(hours))

# #         ws.cell(row=row, column=8, value=remark)

# #     # ---- Dropdown for Work From column (F) — WFO / WFH ----
# #     from openpyxl.worksheet.datavalidation import DataValidation
# #     # dv = DataValidation(type="list", formula1='"WFO,WFH"', allow_blank=True)
# #     dv = DataValidation(type="list", formula1='"WFO,WFH,On Leave"', allow_blank=True)
# #     dv.error = "Please select WFO or WFH"
# #     dv.errorTitle = "Invalid entry"
# #     dv.prompt = "Select WFO or WFH"
# #     dv.promptTitle = "Work From"
# #     ws.add_data_validation(dv)
# #     dv.add(f"F14:F{13 + days_in_month}")
# #     # ---- Dropdown for Activity column (D) — with custom entry allowed ----
# #     dv_activity = DataValidation(
# #         type="list",
# #         formula1='"On Leave,Sick Leave,Public Holiday,Project work"',
# #         allow_blank=True,
# #         showErrorMessage=False,  # allows typing a custom value beyond the 3 options
# #     )
# #     dv_activity.prompt = "Select an activity or type your own"
# #     dv_activity.promptTitle = "Activity"
# #     ws.add_data_validation(dv_activity)
# #     dv_activity.add(f"D14:D{13 + days_in_month}")
# #     # ---- Dropdown for Month (H4) ----
# #     dv_month = DataValidation(
# #         type="list",
# #         formula1='"January,February,March,April,May,June,July,August,September,October,November,December"',
# #         allow_blank=True,
# #     )
# #     dv_month.error = "Please select a valid month"
# #     dv_month.errorTitle = "Invalid entry"
# #     dv_month.prompt = "Select a month"
# #     dv_month.promptTitle = "Month"
# #     ws.add_data_validation(dv_month)
# #     dv_month.add("H4")

# #     # ---- Dropdown for Year (H6) ----
# #     dv_year = DataValidation(
# #         type="list",
# #         formula1='"2024,2025,2026,2027,2028,2029,2030,2031"',
# #         allow_blank=True,
# #     )
# #     dv_year.error = "Please select a valid year"
# #     dv_year.errorTitle = "Invalid entry"
# #     dv_year.prompt = "Select a year"
# #     dv_year.promptTitle = "Year"
# #     ws.add_data_validation(dv_year)
# #     dv_year.add("H6")
# #     # Blank out any leftover template rows beyond this month's day count
# #     for row in range(14 + days_in_month, 45):
# #         for col in range(2, 9):
# #             ws.cell(row=row, column=col, value=None)
    
# #     # ---- Sign-off block ----
# #     ws["C52"] = emp_name
# #     ws["F52"] = get_field(res_row, ["Capgemini Resposible", "Capgemini Responsible"])
# #     ws["F56"] = get_field(res_row, ["EGA Resposible", "EGA Responsible"])
# #     # ---- Lock only these specific cells; everything else stays editable ----
# #     # locked_cells = ["D6", "D7", "D8", "D9", "D10", "D11"]

# #     # # Unlock all cells first
# #     # for row in ws.iter_rows():
# #     #     for cell in row:
# #     #         cell.protection = Protection(locked=False)

# #     # # Lock only the required ones
# #     # for coord in locked_cells:
# #     #     ws[coord].protection = Protection(locked=True)

# #     # # Enable sheet protection so locking actually takes effect
# #     # ws.protection.sheet = True
# #     # ws.protection.password = "yourpassword"  # optional, remove this line if no password needed
# #     # ---- Freeze specific fields so employees can't edit them ----
# #     # from openpyxl.styles import Protection

# #     # locked_cells = ["D6", "D7", "D8", "D9", "D10", "D11"]

# #     # for row in ws.iter_rows():
# #     #     for cell in row:
# #     #         cell.protection = Protection(locked=False)

# #     # for coord in locked_cells:
# #     #     ws[coord].protection = Protection(locked=True)

# #     # ws.protection.sheet = True
# #     # ws.protection.enable()

# #     # out = io.BytesIO()
# #     # # wb.save(out)

# #     # # out = io.BytesIO()
# #     # ---- Lock everything EXCEPT Activity (D) and Work From (F) daily cells ----
# #     from openpyxl.styles import Protection

# #     editable_cols = {4, 6, 7}  # column D = Activity, column F = Work From
# #     editable_rows = set(range(14, 14 + days_in_month))  # daily table rows only

# #     for row in ws.iter_rows():
# #         for cell in row:
# #             is_editable = (cell.row in editable_rows) and (cell.column in editable_cols)
# #             cell.protection = Protection(locked=not is_editable)
# #     ws["D5"].protection = Protection(locked=False)   # ← new line: unlocks CID

# #     ws.protection.sheet = True
# #     ws.protection.enable()

# #     out = io.BytesIO()
# #     # wb.save(out)
# #     wb.save(out)
# #     out.seek(0)
# #     return out.getvalue()


# # # =====================================================================
# # # 3. STREAMLIT UI
# # # =====================================================================
# # # =====================================================================
# # # 3. STREAMLIT UI
# # # =====================================================================

# # st.title("📋 Timesheet Auto-Filler")
# # st.write(
# #     "Upload **SOW & Resource Details** and **Replicon Dump** files below. "
# #     "The app will automatically generate a filled timesheet for **every employee** "
# #     "found in the SOW file, and save each one with that employee's name."
# # )

# # col1, col2 = st.columns(2)
# # with col1:
# #     sow_file = st.file_uploader("1️⃣ Upload SOW & Resource Details (.xlsx)", type=["xlsx"])
# # with col2:
# #     replicon_file = st.file_uploader("2️⃣ Upload Replicon Dump (.xlsx)", type=["xlsx"])

# # st.subheader("3️⃣ Timesheet Template")
# # template_option = st.radio(
# #     "Choose which timesheet template to fill:",
# #     ["Use built-in default template (hardcoded in app)", "Upload a custom template"],
# #     index=0,
# # )

# # custom_template_file = None
# # if template_option == "Upload a custom template":
# #     custom_template_file = st.file_uploader(
# #         "Upload Timesheet Template (.xlsx)", type=["xlsx"], key="template_upload"
# #     )
# # else:
# #     st.caption("✅ Using the built-in default timesheet template embedded in this app.")

# # st.divider()
# # generate_btn = st.button("🚀 Generate Timesheets", type="primary", use_container_width=True)

# # if generate_btn:
# #     if not sow_file or not replicon_file:
# #         st.error("Please upload both the SOW & Resource Details and Replicon Dump files.")
# #     elif template_option == "Upload a custom template" and not custom_template_file:
# #         st.error("Please upload a custom timesheet template, or switch to the built-in default.")
# #     else:
# #         try:
# #             with st.spinner("Reading files..."):
# #                 sow_df = load_sow(sow_file)
# #                 rep_df = load_replicon(replicon_file)
# #                 daily_hours, neg_set = build_daily_hours(rep_df)

# #                 # Auto-detect month/year from the Replicon dump's dates
# #                 month = int(rep_df["Item Date"].dt.month.mode()[0])
# #                 year = int(rep_df["Item Date"].dt.year.mode()[0])

# #                 template_bytes = (
# #                     custom_template_file.read() if custom_template_file else get_default_template_bytes()
# #                 )

# #             st.success(
# #                 f"Detected period: **{calendar.month_name[month]} {year}**. "
# #                 f"Found **{len(sow_df)}** resource(s) in the SOW file."
# #             )

# #             generated_files = {}
# #             unmatched_employees = []
# #             rep_name_keys = daily_hours["name_key"].unique().tolist()
# #             progress = st.progress(0.0)
# #             for i, (_, res_row) in enumerate(sow_df.iterrows()):
# #                 # matched_key = find_matching_rep_key(res_row["Resource name"], rep_name_keys)
# #                 matched_key = find_matching_rep_key(get_field(res_row, ["Resource name", "Employee Name", "Name"]), rep_name_keys)
# #                 # if matched_key is None:
# #                 #     unmatched_employees.append(res_row["Resource name"])
# #                 if matched_key is None:
# #                     unmatched_employees.append(get_field(res_row, ["Resource name", "Employee Name", "Name"]))
# #                 data = fill_timesheet_for_employee(
# #                     template_bytes, res_row, daily_hours, neg_set, month, year, matched_key
# #                 )
# #                 # emp_name_clean = str(res_row.get("Resource name", f"Employee_{i+1}")).strip().replace(" ", "_")
# #                 emp_name_clean = str(get_field(res_row, ["Resource name", "Employee Name", "Name"], f"Employee_{i+1}")).strip().replace(" ", "_")
# #                 fname = f"Timesheet_{calendar.month_name[month]}_{year}_{emp_name_clean}.xlsx"
# #                 generated_files[fname] = data
# #                 progress.progress((i + 1) / len(sow_df))

# #             if unmatched_employees:
# #                 st.warning(
# #                     "⚠️ No matching hours found in the Replicon dump for: "
# #                     + ", ".join(unmatched_employees)
# #                     + ". Their timesheet was generated with all days marked as Leave — "
# #                     "double check the name spelling in both files."
# #                 )


# #             st.subheader("✅ Generated Timesheets")
# #             for fname, data in generated_files.items():
# #                 st.download_button(
# #                     label=f"⬇️ Download {fname}",
# #                     data=data,
# #                     file_name=fname,
# #                     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
# #                     key=fname,
# #                 )

# #             # Zip download of all files together
# #             zip_buffer = io.BytesIO()
# #             with zipfile.ZipFile(zip_buffer, "w") as zf:
# #                 for fname, data in generated_files.items():
# #                     zf.writestr(fname, data)
# #             zip_buffer.seek(0)

# #             st.download_button(
# #                 "📦 Download All Timesheets as ZIP",
# #                 data=zip_buffer.getvalue(),
# #                 file_name=f"All_Timesheets_{calendar.month_name[month]}_{year}.zip",
# #                 mime="application/zip",
# #                 use_container_width=True,
# #             )

# #         except Exception as e:
# #             st.error(f"Something went wrong: {e}")
# #             st.exception(e)

# # st.divider()
# # with st.expander("ℹ️ How field mapping works"):
# #     st.markdown(
# #         """
# #         - **Partner Emp. Name / CID / Role / Domain / PO Number / SOW Number** ← from SOW & Resource Details file
# #         - **Daily Billable Hours** ← summed per employee per date from the Replicon dump
# #           (negative correction rows are netted automatically)
# #         - **Activity** ← SOW name, filled only on days with hours > 0
# #         - **Work From** ← "WFO" if role contains "onsite", otherwise "WFH"
# #         - **Days with 0 hours** ← marked as "Leave"
# #         - **Capgemini/EGA Responsible, sign-off Name** ← from SOW & Resource Details file 
# #         """
# #     )





# """
# Timesheet Auto-Filler - Streamlit App
# =======================================
# Upload SOW & Resource Details + Replicon Dump files, and this app will
# automatically generate one filled Timesheet per employee.

# Features:
# - Upload SOW & Resource Details (.xlsx)
# - Upload Replicon Dump (.xlsx)
# - Timesheet template: use the BUILT-IN default template (hardcoded/embedded
#   in this file as base64) OR upload your own custom template (.xlsx)
# - Generates one filled timesheet per employee found in the SOW file
# - Download each file individually, or all together as a ZIP

# Run with:
#     streamlit run timesheet_app.py
# """

# import streamlit as st
# import pandas as pd
# import openpyxl
# import re
# import io
# import base64
# import calendar
# import zipfile
# from datetime import datetime
# from openpyxl.worksheet.protection import SheetProtection
# from openpyxl.styles import Protection
# from openpyxl.styles import Protection, Alignment

# st.set_page_config(page_title="Timesheet Auto-Filler", layout="wide", page_icon="📋")
# st.markdown("""
# <style>
#     @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

#     html, body, [class*="css"] {
#         font-family: 'Inter', -apple-system, sans-serif !important;
#     }

#     /* Overall background */
#     .stApp {
#         background: linear-gradient(-45deg, #f4f7fb, #eaf0f8, #e6eefc, #eef4fc);
#         background-size: 400% 400%;
#         animation: gradientShift 18s ease infinite;
#     }
#     @keyframes gradientShift {
#         0% { background-position: 0% 50%; }
#         50% { background-position: 100% 50%; }
#         100% { background-position: 0% 50%; }
#     }

#     /* Top header bar */
#     header[data-testid="stHeader"] {
#         background: transparent;
#     }

#     div.block-container {
#         padding-top: 0.5rem;
#         padding-bottom: 1rem;
#         padding-left: 2rem;
#         padding-right: 2rem;
#         max-width: 100%;
#     }

#     div[data-testid="stVerticalBlock"] > div {
#         gap: 0.5rem;
#     }

#     h1 {
#         color: #0b3d91;
#         font-weight: 700;
#         letter-spacing: -0.5px;
#     }
#     h2, h3 {
#         color: #14395e;
#         display: flex;
#         align-items: center;
#         gap: 0.5rem;
#         margin-top: 0.2rem;
#         margin-bottom: 0.2rem;
#         font-size: 1.1rem;
#     }

#     /* Hero banner */
#     .hero-banner {
#         display: flex;
#         align-items: center;
#         gap: 1rem;
#         background: linear-gradient(135deg, #0b3d91 0%, #1a56c4 55%, #3b7ce0 100%);
#         border-radius: 14px;
#         padding: 1rem 1.4rem;
#         margin-bottom: 0.8rem;
#         box-shadow: 0 6px 18px rgba(11, 61, 145, 0.3);
#         position: relative;
#         overflow: hidden;
#     }
#     .hero-banner::after {
#         content: "";
#         position: absolute;
#         top: -40px;
#         right: -40px;
#         width: 180px;
#         height: 180px;
#         background: rgba(255,255,255,0.08);
#         border-radius: 50%;
#     }
#     .hero-banner::before {
#         content: "";
#         position: absolute;
#         bottom: -60px;
#         right: 80px;
#         width: 140px;
#         height: 140px;
#         background: rgba(255,255,255,0.06);
#         border-radius: 50%;
#     }
#     .hero-icon {
#         font-size: 1.8rem;
#         background: rgba(255,255,255,0.15);
#         border-radius: 10px;
#         padding: 0.5rem 0.65rem;
#         box-shadow: inset 0 0 0 1px rgba(255,255,255,0.25);
#         z-index: 1;
#     }
#     .hero-text { z-index: 1; }
#     .hero-text h1 {
#         color: #ffffff !important;
#         margin: 0 0 0.2rem 0;
#         font-size: 1.3rem;
#     }
#     .hero-text p {
#         color: #dce8fb;
#         font-size: 0.82rem;
#         margin: 0;
#         line-height: 1.4;
#     }
#     .hero-text p b { color: #ffffff; }

#     /* Step tracker */
#     .step-tracker {
#         display: flex;
#         align-items: center;
#         justify-content: center;
#         margin: 0.8rem 0 1rem 0;
#     }
#     .step {
#         display: flex;
#         flex-direction: column;
#         align-items: center;
#         gap: 0.3rem;
#     }
#     .step-num {
#         width: 24px;
#         height: 24px;
#         border-radius: 50%;
#         display: flex;
#         align-items: center;
#         justify-content: center;
#         background: #d7e3f5;
#         color: #6f89ad;
#         font-weight: 700;
#         font-size: 0.75rem;
#         transition: all 0.3s ease;
#     }
#     .step.active .step-num {
#         background: linear-gradient(135deg, #1a56c4, #0b3d91);
#         color: #ffffff;
#         box-shadow: 0 4px 12px rgba(11,61,145,0.35);
#     }
#     .step-label {
#         font-size: 0.68rem;
#         color: #5c7699;
#         font-weight: 600;
#     }
#     .step.active .step-label { color: #0b3d91; }
#     .step-line {
#         width: 45px;
#         height: 2px;
#         background: #d7e3f5;
#         margin: 0 0.6rem;
#         margin-bottom: 1rem;
#     }

#     /* Card-style containers */
#     [data-testid="stVerticalBlockBorderWrapper"],
#     div.stFileUploader, div[data-testid="stExpander"] {
#         background: #ffffff;
#         border-radius: 10px;
#         padding: 0.6rem;
#         box-shadow: 0 3px 10px rgba(11, 61, 145, 0.07), 0 1px 3px rgba(11, 61, 145, 0.05);
#         border: 1px solid #e3ecf7;
#     }

#     div.stFileUploader label {
#         font-weight: 600;
#         color: #14395e;
#         font-size: 0.9rem;
#     }
#     div.stFileUploader section {
#         border: 2px dashed #b8cceb !important;
#         border-radius: 10px !important;
#         background: #f8fbff !important;
#     }
#     div.stFileUploader {
#         transition: box-shadow 0.2s ease, transform 0.2s ease;
#     }
#     div.stFileUploader:hover {
#         box-shadow: 0 6px 16px rgba(11, 61, 145, 0.12);
#         transform: translateY(-1px);
#     }

#     /* Stat chips */
#     .stat-chip {
#         background: #ffffff;
#         border-radius: 10px;
#         padding: 0.5rem 0.7rem;
#         text-align: center;
#         box-shadow: 0 2px 8px rgba(11,61,145,0.06);
#         border: 1px solid #e3ecf7;
#         color: #14395e;
#         font-size: 0.78rem;
#     }
#     .stat-chip b { color: #0b3d91; font-size: 0.85rem; }

#     /* File result cards */
#     .file-card {
#         display: flex;
#         align-items: center;
#         gap: 0.6rem;
#         background: #f8fbff;
#         border: 1px solid #e3ecf7;
#         border-radius: 8px;
#         padding: 0.45rem 0.7rem;
#         margin-bottom: 0.4rem;
#         transition: all 0.2s ease;
#     }
#     .file-card:hover {
#         border-color: #1a56c4;
#         box-shadow: 0 3px 10px rgba(11,61,145,0.1);
#     }
#     .file-icon { font-size: 1.05rem; }
#     .file-name { font-weight: 600; color: #14395e; font-size: 0.82rem; }

#     /* Primary button */
#     button[kind="primary"] {
#         background: linear-gradient(135deg, #1a56c4 0%, #0b3d91 100%) !important;
#         border: none !important;
#         border-radius: 8px !important;
#         font-weight: 600 !important;
#         padding: 0.55rem 1.1rem !important;
#         color: #ffffff !important;
#         box-shadow: 0 3px 10px rgba(11, 61, 145, 0.3);
#         transition: transform 0.15s ease, box-shadow 0.15s ease;
#     }
#     button[kind="primary"]:hover {
#         transform: translateY(-1px);
#         box-shadow: 0 6px 16px rgba(11, 61, 145, 0.4);
#     }

#     /* Download buttons */
#     div.stDownloadButton > button {
#         background: #ffffff;
#         color: #0b3d91;
#         border: 1.5px solid #0b3d91;
#         border-radius: 8px;
#         font-weight: 600;
#         font-size: 0.85rem;
#         transition: all 0.15s ease;
#         box-shadow: 0 2px 6px rgba(11, 61, 145, 0.06);
#     }
#     div.stDownloadButton > button:hover {
#         background: #0b3d91;
#         color: #ffffff;
#         box-shadow: 0 4px 12px rgba(11, 61, 145, 0.25);
#         transform: translateY(-1px);
#     }

#     div[role="radiogroup"] {
#         background: #ffffff;
#         padding: 0.5rem 0.7rem;
#         border-radius: 10px;
#         box-shadow: 0 2px 6px rgba(11, 61, 145, 0.05);
#         border: 1px solid #e3ecf7;
#     }

#     div[data-testid="stAlert"] {
#         border-radius: 10px;
#         box-shadow: 0 2px 8px rgba(0,0,0,0.05);
#         font-size: 0.85rem;
#     }

#     hr {
#         border-top: 2px solid #dce6f5;
#         margin: 0.6rem 0;
#     }

#     div[data-testid="stProgress"] > div > div {
#         background: linear-gradient(90deg, #1a56c4, #0b3d91);
#     }

#     div[data-testid="stHorizontalBlock"] { gap: 0.6rem; }

#     .app-footer {
#         text-align: center;
#         padding: 1rem 0 0.4rem 0;
#         color: #8fa3c2;
#         font-size: 0.75rem;
#         border-top: 1px solid #dce6f5;
#         margin-top: 1.2rem;
#     }
# </style>
# """, unsafe_allow_html=True)

# DEFAULT_TEMPLATE_B64 = """
# UEsDBBQABgAIAAAAIQCVZpbtkAEAAIIFAAATAAgCW0NvbnRlbnRfVHlwZXNdLnhtbCCiBAIooAACAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACslM9O4zAQxu8r7TtEvqLEhcMKoaYc+CMh7QIS7AMM9qSx6tiWZ4D27Zm4gBDqtqzoJVZiz/f9PJmZ6ely8NUTZnIxtOqwmagKg4nWhXmr/t5f1seqIoZgwceArVohqdPZzx/T+1VCqiQ6UKt65nSiNZkeB6AmJgyy08U8AMtrnusEZgFz1EeTyS9tYmAMXPOooWbTc+zg0XN1sZTPa5KMnlR1tj44erUKUvLOAAupfgr2k0v96tBIZDlDvUt0IBhKb3QYd/5t8Bp3I6nJzmJ1C5mvYRAMvfT6OebFQ4yLZrvIBsrYdc6gjeZxkAw0lDKCpR6RB9+UtRnAhTfuLf7lMOmyHO4ZZLxfEd7BwfK/UZfn9xGKzA5D4pVH2nfai+gu5x4y2jvO0hl7B/iovYPDgDdnvZTInpPwrrvNX+r2DzJYYNC/4QH9VejiF0AGqteF3xgPRE56oPSoHzXem3RTsYvjbY6JZGZk/P8rvw2FMbpOIoSZHX7NUebNt3OM40SzaDd46zJBZy8AAAD//wMAUEsDBBQABgAIAAAAIQAxHYnNIgEAAN4CAAALAAgCX3JlbHMvLnJlbHMgogQCKKAAAgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAArJLRSgMxEEXfBf8h5L0721VEpNu+iFBQEKkfME1mt6FJJiRR2783rRZdqEXQx2Tu3Jy5k8ls46x4pZgM+1aOq1oK8oq18X0rnxd3o2spUkav0bKnVm4pydn0/GzyRBZzaUorE5IoLj61cpVzuAFIakUOU8WBfKl0HB3mcow9BFRr7Amaur6C+N1DTgeeYq5bGef6QorFNpSX/+INjjJqzAiKI41CLGQxmzKLWGDsKbdSs3os12mvqAq1hONAzQ9AzqjIibtcKXbAXWfUbsymhroZTgrKYkqmCPYJWlySHZI8HHDvd7W57/gU0fj3EX1g3bJ6ceTzkS18gh8UX/lsLLxxXC+Z16dYLv+ThTaZvCZ9emEYwoEIBr9y+g4AAP//AwBQSwMEFAAGAAgAAAAhAD4Ms3q3AwAAeQkAAA8AAAB4bC93b3JrYm9vay54bWysVm1vozgQ/n7S/QfEdxcbDAHUdBXedJXaVZVm2zup0soFp0EBnDNOk6ra/75jAknbnE657kUJxvbw+JmZZ4acf9nWlfHMZVuKZmySM2wavMlFUTZPY/PbLEO+abSKNQWrRMPH5gtvzS8Xv/92vhFy+SjE0gCAph2bC6VWoWW1+YLXrD0TK97AzlzImimYyierXUnOinbBuaory8bYs2pWNuYOIZSnYIj5vMx5IvJ1zRu1A5G8Ygrot4ty1Q5odX4KXM3kcr1CuahXAPFYVqV66UBNo87Dy6dGSPZYgdtb4hpbCV8PfgTDxR5Ogq2jo+oyl6IVc3UG0NaO9JH/BFuEvAvB9jgGpyFRS/LnUudwz0p6n2Tl7bG8AxjBv4xGQFqdVkII3ifR3D0327w4n5cVv9tJ12Cr1VdW60xVplGxVqVFqXgxNkcwFRv+bkGuV9G6rGDXwcQhpnWxl/ONNAo+Z+tKzUDIA/zYtLHtYKwtQRiTSnHZMMVj0SjQYe/Xr2quw44XAhRuTPnf61JyKCzQF/gKV5aH7LG9YWphrGU1NuPw4VsL7j+81KyqyiV7SMSmqQTU2MMbcbLjSvgP8mS59tkCp3fEdvcfAwD8ZDhI8EZJA+4vkytIwy17hqRA6ou+Zi8h6v73V4c6MfGdBI2CNEEUTwLkByRCmZf4qR+7JKWTH+CF9MJcsLVa9InWmGOTQlaPtq7ZdtghOFyXxeH8V9x/kB4/XIa9H9pT3dLuSr5pD5LQU2N7XzaF2IxNRLSQX95PN93mfVmoBTTRgNpgslv7g5dPC2BMiAvy0X1DMxubry5xUpplPoojb4Jo4HrIx06E/HSSxK7tZI7ndoysN5S65gnUutFoOsHf6oZKoEvrUUcX7mWoz5CXRadta3gsZ1UOAtdDZxgQbAfaa75VV63qRtBWCfQIpGSEA4pw6riI+oGNfOrYKKaJnbqjNEkjoDcU9P/QAjuJh8NbRbNcMKlmkuVLeBdN+TxiLShp5xDwfUs2cv0IO0CRZiRDlAQYRZFHkZtkjjsiSZy62YGsdn/+yQbkW93TnKk1FKeuy24e6mvWr+4X57uFPk/vii6cJjru/dP/ZngL3lf8ROPs7kTD+Ov17PpE26t09v0+O9V4ch0lk9PtJ9Pp5K9Z+udwhPWPAbU+JDwhNMBOOkGOE1NER9kI+Rl2kUNHNHZplBI8OiS82uTPn8u3Ta1BkfHb/wh9M9L51+Bh/wfKaLnqt+CN0TXOjrim39XXHu3iJwAAAP//AwBQSwMEFAAGAAgAAAAhAJIHlOwEAQAAPwMAABoACAF4bC9fcmVscy93b3JrYm9vay54bWwucmVscyCiBAEooAABAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAKySy2rEMAxF94X+g9G+cTJ9UIZxZtFSmG2bfoBwlDhMYgdbfeTva1I6ycCQbrIxSML3Hom72393rfgkHxpnFWRJCoKsdmVjawXvxcvNI4jAaEtsnSUFAwXY59dXu1dqkeOnYJo+iKhigwLD3G+lDNpQhyFxPdk4qZzvkGPpa9mjPmJNcpOmD9LPNSA/0xSHUoE/lLcgiqGPzv9ru6pqND07/dGR5QsWMvDQxgVEgb4mVvBbJ5ER5GX7zZr2HM9Ck/tYyvHNlhiyNRm+nD8GQ8QTx6kV5DhZhLlfE0Zjq58MNnaCObWWLnK3aigMeirf2MfMz7Mxb//ByLPY5z8AAAD//wMAUEsDBBQABgAIAAAAIQDDG6yZDw8AAG1IAAAYAAAAeGwvd29ya3NoZWV0cy9zaGVldDEueG1spFxrb9vIFf1eoP9BIBaFg40t8SXJqq2FJVGSgc1mkTi724+KTMdCJNOV5Dxa9L/3zMy94uXMOKbpoq3iM/femXvmdUgOefbLt8269SXf7lbF3XkQnnSCVn63LK5Xd5/Ogw9X0+N+0NrtF3fXi3Vxl58H3/Nd8Mvw7387+1psP+9u83zfQoS73Xlwu9/fD9rt3fI23yx2J8V9foeSm2K7Wezx5/ZTe3e/zRfX2mmzbkedTre9WazuAhNhsK0To7i5WS3zSbF82OR3exNkm68Xe7R/d7u633G0zbJOuM1i+/nh/nhZbO4R4uNqvdp/10GD1mY5uPx0V2wXH9fI+1uYLJatb1v8N8L/Yq5G405Nm9VyW+yKm/0JIrdNm930T9un7cXyEMnNv1aYMGlv8y8r1YFlqKhZk8L0ECsqg8UNg3UPwRRd28HD6vo8+G+UJtHkojM97l10kuPk9AJj7LTbO06TbBpPo36/34n+FwzPrlfoYZVVa5vfnAejcHDZ7QTt4ZkeQH+s8q878e/WfvHxfb7Ol/sclYRBS43Pj0XxWRleAuog5E4bqJC7f+ugF+Hgr+kk7CT9tNdVsduH4PLfXNFUD+bft63r/GbxsN6/K77O89Wn270K3/pPvi34TzRAj4zB9fdJvltiqMLkJE5VHctijYD4/9ZmpeYchtrim2nz6np/q2Zh1D1N4zBKg9byYbcvNn9SAbkbR/SQdsTvV1N+etKLumn3CT90hvbDL/mFaP0PKkrIAb/kEHVq1YT265rwyzX1Trpx1H8qtS454pcd41o19sgRv+TYP4nr1IhlTjcVv2VT67B5So74Zccfdl/bdL8eZ5PFfjE82xZfW5j7GAe7+4VaSaMBgqlBhfRNt9QZVhQGo8EOEyKOZzQmmBFLVfNI+egGoM4d0C/Dzln7C2bAkizGxiLS40+5TGwgs4GpDcxsYG6AWM87tP7ABAamk0JygkHkn1KchXI7DzBID1mEVhZsoSahzsIAiHxwiaouGVuwy9QGZjYwN4BZTmRaaFmTtJTbeYABfWhjbKXFFoe0DNA/9FZmgNMDMDWAWY4UEzMDqIXgUE1SrWZOJhinB5P0YNKWmYLQJpkqNwxDDItDBV0rVTLRY8Z0oUFErgYQuRpA5EpB9Gaig8wJCZ2hiJnjyaR7WJifnJg8OFUg5Cb57Vm5HUwO/UhIUnakg0wNIrLz1NS3epJMQj3Vo05UslzpRoy5Jt2o3KxuPLVSJRPRjYSUeWSE6Hmke2lqEJGqDcwN4E49JWKttTWssaIoN2tFsRdGNjn0mQHCnuY2STud6DRMQqsHMrIqh+3UICI5G5gbwE0Om0WT5JQbugk/5XppL5gHm0N6BomUmvoyTNKkZ60SGbmIzAwiMqMYeraZ+UeI3l8qQ1BNmCbJaT87O2ttH5dGh/QI4vx6ccfq8IydRIIEiQw5jEiRILf7Qs/GX2dwaj87RXtrKI3KFFV954FYMclILJmEyIyMVyQzMpAnI58GqTHd1Mqv5E+5LIwJKhsycZDMQaYcqKIFrIE6Y6MUQ/lm+P7Dm6NZmAxmSfLqrH3j0UBzdpD7U/jIDhhaGgZpxYCMmjOXEVpwKZUX+eUZLhZ489DRjGE5V+2tkY2kUgitPWbCRlhLykj24sRGckAYURPL/Su0VvUZ+VWMImv+zNlIyojIWncu2cizIDRUUaFRL7ESvTfDycVVdjTvvn7z9rer+dE8+cdic//P8NXrkPs+6YahM5coBMnF/VaF0jTeLrb5dWCuU8cYRONEma70ReeH33/P3h1dZX9dHY3C5HUwwX+CV1zPn9nkt+z95OJf1Q1yws0td7+MoV6p4hgq5zKGsJ5Dcdl3c4KSUu9cMuQKnrChdtN+2CsjLymjMB2MErVegBTIDjDxM/pcTzNFtTU3xxwNo04NU5dqQy7WtC/Dq/mHdz4GjaaLJYMESQYJkgwaCF3IS+acm1MuQ5cMeRj0a0ZE/fHli7pKU2rDYdBwpnNVXJUrjt47x+xXh6vpu0vPWDMVV5giSDJFkGSKWiyZIkgyRZCHqYYCE+KqFlPOCkl+dZh6f3H14Z2HKwohRxVBkisDxZIrspJcEVRaXXJmHq4aStiQBOpTo8q+GGG/Wlx9+M3DFFUsmSJIMmWgClNkJZkiSI4qgjxMNdTDSgnXmX/WbjlmvzpMYb/xMGUqrsw/giRTBMkxRS2WTBEkmSLIZQoqt5G41n41Vir7qo/9KkwN5QYZdewN8uqDb3ukSJIyhgRlDAnKuAmCMoYEZQx5KGso1qGb6wwuWy2N2a/O4HpMS1CMClmmObEkiyBJFjVakkWQJIsgD1kNrwMicx3w1E5oq8Yx+71ANVCICld0m1JyRZDkitosuSJIrO/cQg9X1gWDujCocxWIi6VaA8u+9mW/xqqBAlSYoosEyRRBkilqsWSKIDmqCPIw1fAKIDLK+MlRZct+9nuBauAQYi9kSHJF0l1yRW2WXBEkuSLIw1VDNa+eINXYCyNbt7NfY9XAASRT1BbJlKva2VEyRY6SKYI8TDVU7erivQ5Ttmpnv8aqgQNIpqgtkilXtbOjZIoc5Ur1qGqPGqp27fe0ahA3o831DfvVWtX9asEV7RxUUuWKdraSVLmina08g6qhaI/qifbIFu3s9yK14Op2DivJcnU7W0myXN3OVh6yGur2qJ5uj2zdzn4vUQtUtZyDBEmuXOXOdUuuXOXOVi5X6hZbo0ekyq/GHLSVu64Pfo3VAgcQTDEkmCJIXg2ylWCKIbGuM+RhqqFgj+sJ9th+HsR+L1ALHEJyRa2RXLl6nR0lV65eZysPVw31elxPr8fO43jya6wWuGLJFMWUTLlqnR0lU+QoRxVBHqaUOm3wQFE9A6gx/2JbrbNfY7XAASRT1BbJlIEq84+sJFMESaYI8jDVUK3jRn0tpmy1zn7N1QJHkFRRYyRVrlhnR0mVK9bZykNVQ7GOB0W1qLLFOvu9RC1wDEmWq9fJqjKu3LvsHEuOq0f1Oq47m83Aeno9tvW6rq/mDvjIEwkOIblyFTtZVbhy77NzLMnVo4odt3macVXvPnts32fX9b1ILbiCnWPKKegKdraSU5BiSaYI8kzBhoIdvVVrCtqCnf1eohZcvc5RJVeuXmcryZWr19nKw1VDvY5Hk7W4svU6+zVXC65a55iSKVets5VkylXrbOUyhSevzU7+1VPrsa3WdX0155/3iQQHECsVQ4IpguRKxVaCKYbE/GPIw1RDta4fLT99XZPYap39mqsFjiCpcsU6WVWocm+ucyxJ1aM313Efs9mgkmJ9eDk9wgOpo+ytOQnx0+gnnFHovHp1HvVfB8FrPLP/WZ6JSGwJr1tRc6g99nyCY0gKXRVPVhUK3XvuHEtS+KiKV5ezjU7kShWvKBwleN8iUIQ9zuep4TMCn+KQSWIrfd2ml+kMDiH5dLU+WVX4dLU+x5J8Pqr19amXBldF2g85Y00tT3sYXuOneI3xxFCN07jkVb3GcTi4TrFTOrj0o5Mk0m/CfnqN0jciM4bKU/BThsoTazg/pre3VC6BBGkSK8cbk4aSX/udB2nZxWOGygV64kIZQ9joyyPv1hiclkZ8AmZGEO6y1zwsxzH0OaRqzg2Ve2L0rXkHyNwaJqgcnBMHyRxk6iAzB5kTgud1alRUTvE3FNOJEZ2V5htINt9GMnIrbaYOMnOQOSG+5nsUbhI//XJFYqRhF/tROXCcFxHISLwnwn7lDMlcaOpCMxeaM6SnVrVTpBZFI+o+PVVLDu7HVDqFINkrBMkF1YGmHKu0mrnQnCE9TasvTUiV+Iwc1NJm5cBQ+TxnwlB5LjBzoakLzVxoTlBPHyis5iD123NyMDJH9kNqINENDpI5yNRBZg4yJ8QzOVKpqcrmN3jdQ0fCWzuV6WLfgGEjMV0YEtOFoXJDmbqOM9dxXnGsdpMUPi/L04gBeWQ7NVBP9pwDZa7VlCGs8+USY92HmZVGh9OZDHmmlLz5+JzhaDbsXmWXdF5FIiOZpwNl0AFqdgo2pgSVfjMHmRPiG6NSMjwnJ3MrrTLFDFTpKQfKUgeaEiQzMEYlMicbXwZSADwnA1cAQABpdkk5mveZ7DPvbCSWb4bE1SxBMicTXOZkEF9OUhU8Jyez4ffLs9njlKBSeU5cKHOhqQvNXGjOkKtsUikNKAck/Ow3UNNnbsaYyPTuKgZneUFWg0WlzcSLtJt8+ykf5+v1rrUsHtQr11jqh2cHmA7op91BZpSzVTJJ4kGGKx6EdUoSlOgV2I6W9AdqQ3R9xmmEejTNlg/Wb/j4SsYpWoCZ40bD6ggfb9vi3iAzJ/nsVsd9lGgR4JScokRrASfTDjL15TNJQpToUen4IFNceXt4i+GDB46+EvjgMZyvBBzgeZWvBL2A5ye+khQlPt4mMXrbHFKzWx3BBxc0nmgRfHBMxlcCrnHew1cCrnG4wVcCrvEo39dqcI1H156SED44WOyLBh+cT/WVgGvzfpaTKbg2i5ZTAq7N5u2UgGscUvO1DT54QcdXAh+8WuIrAdd4ZcJXAq7xCoOvBFzjQLyvBFzjgLpbMuoNxl4PhPLho/5g7I1zOhj72B+FnYF6bdBTcxiixDvSUZJ5S0ZYBnwzYBQPxv5VCNPMV3c6GHu5BeleHJRrxtvliqk+SHF3vVKfUVmszQtqe3wDhj+UgReSBrgn0prid46FEOY37x7WeWv//R5fR8m/4dMuO/MJkutvN+Z7G6377arY4oMq+pWfM/UlmIf1Yvj23dFPeEnqPOBXLILX9Ld+jSBQd8fIFPdxdC3qho6vdfhICL6g8MdivcKv+gAMr/tq2a8WUUOv8+Vqs1gHrcV6XXwdrRd3n82XF26Lr5d39w/7N0hj8QkpqcsPgNl2W2wlmCvgarVXH4UZXf7668Xo16w1f4t3kPDlHFUGeLVeq6/GtG6Lh+2utcEnPFof81b/JG0V29YaNZwguPkCCb1zKL6P0u+k0WTU6x73o3H/OMEXUY5xe+viOBn38a7BRSfsjrrq+yhEUqi+1XD4N6PREA8FCI1AX5UMB8D3SO6R9ZvF9tMKLK7zG2ye+MYH3mmjz5uof++Le41C630s9vgsCf91i8/65EgcXzgJWjdFsec/ME/ahw8FDf8PAAD//wMAUEsDBBQABgAIAAAAIQD2YLRBuAcAABEiAAATAAAAeGwvdGhlbWUvdGhlbWUxLnhtbOxazY8btxW/B8j/QMxd1szoe2E50Kc39u564ZVd5EhJlIZeznBAUrsrFAEK59RLgQJp0UuB3nooigZogAa55I8xYCNN/4g8ckaa4YqKvf5AkmJ3LzPU7z3+5r3HxzePc/eTq5ihCyIk5UnXC+74HiLJjM9psux6TybjSttDUuFkjhlPSNdbE+l9cu/jj+7iAxWRmCCQT+QB7nqRUulBtSpnMIzlHZ6SBH5bcBFjBbdiWZ0LfAl6Y1YNfb9ZjTFNPJTgGNQ+WizojKCJVund2ygfMbhNlNQDMybOtGpiSRjs/DzQCLmWAybQBWZdD+aZ88sJuVIeYlgq+KHr+ebPq967W8UHuRBTe2RLcmPzl8vlAvPz0MwpltPtpP4obNeDrX4DYGoXN2rr/60+A8CzGTxpxqWsM2g0/XaYY0ug7NKhu9MKaja+pL+2wznoNPth3dJvQJn++u4zjjujYcPCG1CGb+zge37Y79QsvAFl+OYOvj7qtcKRhTegiNHkfBfdbLXbzRy9hSw4O3TCO82m3xrm8AIF0bCNLj3FgidqX6zF+BkXYwBoIMOKJkitU7LAM4jiXqq4REMqU4bXHkpxwiUM+2EQQOjV/XD7byyODwguSWtewETuDGk+SM4ETVXXewBavRLk5TffvHj+9Yvn/3nxxRcvnv8LHdFlpDJVltwhTpZluR/+/sf//fV36L///tsPX/7JjZdl/Kt//v7Vt9/9lHpYaoUpXv75q1dff/XyL3/4/h9fOrT3BJ6W4RMaE4lOyCV6zGN4QGMKmz+ZiptJTCJMLQkcgW6H6pGKLODJGjMXrk9sEz4VkGVcwPurZxbXs0isFHXM/DCKLeAx56zPhdMAD/VcJQtPVsnSPblYlXGPMb5wzT3AieXg0SqF9EpdKgcRsWieMpwovCQJUUj/xs8JcTzdZ5Radj2mM8ElXyj0GUV9TJ0mmdCpFUiF0CGNwS9rF0FwtWWb46eoz5nrqYfkwkbCssDMQX5CmGXG+3ilcOxSOcExKxv8CKvIRfJsLWZl3Egq8PSSMI5GcyKlS+aRgOctOf0hhsTmdPsxW8c2Uih67tJ5hDkvI4f8fBDhOHVypklUxn4qzyFEMTrlygU/5vYK0ffgB5zsdfdTSix3vz4RPIEEV6ZUBIj+ZSUcvrxPuL0e12yBiSvL9ERsZdeeoM7o6K+WVmgfEcLwJZ4Tgp586mDQ56ll84L0gwiyyiFxBdYDbMeqvk+IhDJJ1zW7KfKISitkz8iS7+FzvL6WeNY4ibHYp/kEvG6F7lTAYnRQeMRm52XgCYXyD+LFaZRHEnSUgnu0T+tphK29S99Ld7yuheW/N1ljsC6f3XRdggy5sQwk9je2zQQza4IiYCaYoiNXugURy/2FiN5XjdjKKbewF23hBiiMrHonpsnrip8TLAS//Hlqnw9W9bgVv0u9sy+vHF6rcvbhfoW1zRCvklMC28lu4rotbW5LG+//vrTZt5ZvC5rbgua2oHG9gn2QgqaoYaC8KVo9pvET7+37LChjZ2rNyJE0rR8JrzXzMQyanpRpTG77gGkEl/p5YAILtxTYyCDB1W+ois4inEJ/KDBdzKXMVS8lSrmEtpEZNv1Uck23aT6t4mM+z9qdpr/kZyaUWBXjfgMaT9k4tKpUhm628kHNb0PdsF2aVuuGgJa9CYnSZDaJmoNEazP4GhK6c/Z+WHQcLNpa/cZVO6YAaluvwHs3grf1rteoZ4ygIwc1+lz7KXP1xrvaOe/V0/uMycoRAK3FXU93NNe9j6efLgu1N/C0RcI4JQsrm4TxlSnwZARvw3l0lvvuPxVwN/V1p3CpRU+bYrMaChqt9ofwtU4i13IDS8qZgiXoEtZ4CIvOQzOcdr0F9I3hMk4heKR+98JsCYcvMyWyFf82qSUVUg2xjDKLm6yT+SemigjEaNz19PNvw4ElJolk5DqwdH+p5EK94H5p5MDrtpfJYkFmquz30oi2dHYLKT5LFs5fjfjbg7UkX4G7z6L5JZqylXiMIcQarUB7d04lHB8EmavnFM7DtpmsiL9rO1Oe/a1DriIfY5ZGON9Sytk8g5sNZUvH3G1tULrLnxkMumvC6VLvsO+87b5+r9aWK/bHTrFpWmlFb5vubPrhdvkSq2IXtVhluft6zu1skh0EqnObePe9v0StmMyiphnv5mGdtPNRm9p7rAhKu09zj922m4TTEm+79YPc9ajVO8SmsDSBbw7Oy2fbfPoMkscQThFXLDvtZgncmdIyPRXGt1M+X+eXTGaJJvO5LkqzVP6YLBCdX3W90FU55ofHeTXAEkCbmhdW2FbQWe3Zgnqzy0WzBbsVzsrYa/WqLbyV2ByzboVNa9FFW11tTtR1rW5m1g7LntqkYWMpuNq1IrTJBYbSOTvMzXIv5JkrlVfacIVWgna93/qNXn0QNgYVv90YVeq1ul9pN3q1Sq/RqAWjRuAP++HnQE9FcdDIvnwYw2kQW+ffP5jxnW8g4s2B150Zj6vcfONQNd4330AE4f5vIMCRQCscBfWwFw4qg2HQrNTDYbPSbtV6lUHYHIY92LSb497nHrow4KA/HI7HjbDSHACu7vcalV6/Nqg026N+OA5G9aEP4Hz7uYK3GJ1zc1vApeF170cAAAD//wMAUEsDBBQABgAIAAAAIQBowXh5+wcAADNSAAANAAAAeGwvc3R5bGVzLnhtbOxcWY/bNhB+L9D/IAh99eqw5WNrO83uxkCANAiQLVCg6YNWlr1CdBgSvbET9L93qMMiLVOiTmvR7kNiyTL5zT2coTh/c3Bs4cX0A8tzF6JyI4uC6Rre2nK3C/GPx9VgKgoB0t21bnuuuRCPZiC+Wf780zxAR9v8/GyaSIAh3GAhPiO0u5WkwHg2HT248XamC99sPN/REVz6WynY+aa+DvCPHFtSZXksObrlitEIt47BM4ij+1/3u4HhOTsdWU+WbaFjOJYoOMbt+63r+fqTDVAPykg3hIMy9lXh4CeThHcz8ziW4XuBt0E3MK7kbTaWYWbhzqSZpBvpSDBytZEUTZJVivaDX3GkkeSbLxYWn7icu3tn5aBAMLy9i0Ccp1tC9M37Ndwcj0Qhksq9twY+/fXLQJHh7+/1F8FxnC/CEf5+/U2UlnMpHnA533huOu4YWICZe/vV9b65K/xVNBl+ajkPvgsvug13FDyG4dmeLyBQCpgrvOPqjhk98XaHvED4qPu+9w0/u9Edyz5G36n4RqhM8cOOBaINUUXTRP8+4aeuNmEIkqJQxggbpTChbtYBN0+cDMnoUnTtTiiF+gsKZtn2yTwmYB74xnIOngSZvruCCyH+/Hjcgb664PQilQufK3h66+tHRdX4fxB4trXGKLb3GSt5iu9Z7to8mGC4YLfYJAmscBUhKsDFmAZ8PbKwnxjIN+poNptMFPw3mc6GeKZuASgAYDYdTUbyZKSpYzX0FF1A6JzSM0LLsjoUeQDi8fw1xO3E1w9l0KPo3nJumxsEdPnW9hn/j7wdptJDCILbcr629K3n6jbWpuQX5C8h4ENsX4joGWJz4sDP1RBPQc3A9StAUgIITUKLE3ANHXIzZGYLSMryvGXEXMNHClWsT4kYS2Lun8r2gStlda9RzD3T0tZ8WZuqV9UKrirIsyDSK/fQPw2PFztkfnUh6EtpQDnp20VaOIfjdcd15iCifp1hXjflkTE04ogbFm1upC2tlX2ksFfKWwcMkQm/dqM/Xxr0ybpzli3NOLD/dSCsXLB14L/HoJz4wOmDO3YOHYSNfhLeyHqmmbhankFVVz+cMzWe6jXkCLoju307PveajVRUrqWP7a/DGtLc1mMVh4rSpc389V9dLWlwvVIGSlxbhlK1Ydr2Z1xT/nND9SYPG6IvCatl3KTBLUr8Efot8ceoNB1dACkSOVo0NjHsdAgV6/LjCofNaQLWr9UUFXxMUSnQDY1/Lei7nX3EDVHc6oyvgJL06i6s2qfXb21r6zpm0q3Vk0vh2fOt7zAQbqRiVxz2fg4bNnEseHC/z/CG/YBHqBwgSoULvfIz7nUkTgIQYCBtgAEITIPS3LjPH1kU0JRYFDWadk7ex73zZPqrcNMGocNNa7QBGm9Gzfw8na4tlVymEKMDH1IW4+0NlEEzmMKvCZEn+OR7yDRQtNMG3FSOseMtOcgysPHDWgDb/u70Y8H2jK+4IRy2zKUM/yjBs2icvG4aLweL3lkuwMxxJbmxoLTKlNIRDv5lrADvjugikLGMPqOyVwBECRS2xtF+oh8Cxdu74uwp361VzlPK6Gbq6fk9GKGdLGKyylA1RHVIDCumt5FklI+wVGLQBqSwLlyYxxKpSoE/7y7NZoXUfJbVThtYKUKebLnSPlBErgwnh8NcC5eaSpjJTlke/zzNuxzeKGnNeBP7CmNn2Evj5mIdn7HAIvNiXp813ytl8CDCBCAkFMSiOcujlhASqTSnOl4Rc7jSbxBx+7WJmKcNYm6/YNE85itqs/DN13eP5iEsOeH6WN6CmqXa7fiLdBGbcLwMViL1o9LuAl4zvWsGTRGziHQtH0DNKblCtpIplPWrVMOqMCkVVSsn7BJTUXJRsvkCFVPOl2fn9ZhKCxQmmOz6+YpgCpZLnXKmJBbKuDpcrDEFm1n20+voVnhZ2hdWqgYrmaT4lfiY5qoBuVXZooDBVJlXrvJqQcjt1H3Am0G5Da1uwRQkq82BwXE06Y+y9CwrJx77bRBj9GppfstJfaWJjNr+mpqrz8sUfp+yH7VP2U+R5Lr1GAVq1FgPoWbhnalmnaWSXEXXTDAoV3StUNzLenxGXyxnbFbZrmLpowIV2W0XnFRwSYU3j+zRnNHKgrmSL1O4IOoGVAMum7jU76iWr7BTkLiVuYA95G6lmpC4NbM7SNy9gu4gNdFiqCko7u5Md1yBE0ea365V18Sy9YKrm73SD7uv1pAroU58TTSWv66zA+Ya23doz17HaV0dfZ29PlcHX2fvRA92o6hthDuupS1rf0y3O7+6yHDb9GGUV+3DLj4KEH+IrrI24461FQbnD5q4TozfUoD3EoiXH6hXH04vMQj4rLCF+BHvuLaJ7OVpb9lwOtOF1x5gzPWBfpECri8cZZUe4XThxRl83tJ4MtM0bazJmhydt3TxcCkpHB3/CzMjfLZe+C7HiRyw2rW50fc2ejx9uRDTz7+ba2vvQMYRP/XJevFQOMRCTD9/wGcVKWO8PQx62R8COFwI/hf2vrUQf7y7m8we3q3UwVS+mw5GQ1MbzLS7h4E2ur97eFjNZFW+/4c44a/G+X7hgYSwlV4Z3QY2nALox8TG4D+n9xYicRHBDze3AWwS+0wdy281RR6shrIyGI316WA6HmqDlaaoD+PR3TttpRHYtYrnAMqSokQnCmLw2i2yHNO23ERWiYTIuyAkuMwhQkokIaWnPS7/BQAA//8DAFBLAwQUAAYACAAAACEAjqCrVJIBAACqAwAAFAAAAHhsL3NoYXJlZFN0cmluZ3MueG1sfJNRT+QgFIXfN9n/cNN3p6MPZmPaGmxxprEtDUWNjziDMyQFukDN+u+3xs0mwugj59zvcHIJ2fUfNcKrsE4anSfnq3UCQu/MXupDntyz27NfCTjP9Z6PRos8eRMuuS5+/sic87Cw2uXJ0fvpKk3d7igUdyszCb04L8Yq7pejPaRusoLv3VEIr8b0Yr2+TBWXOoGdmbXPk4vlllnL37Mo/wtF5mSR+YLVLR62GDM4g4FRxPCmLmEg97Ssu02W+iJL3yc/pntEWYdpKJeo3+C27moYFgj1hGLoGYaGVV8kAG77FXSoxeFASzq2DUU0WTlGt9bV1bfxlDRR/BNGUf3KvK8riiLQzepZ2BPG4Ln1UHEv4gYEsN6f9gby+C8zwm7qpkE3DYbtsvoBIn9rZuvCJtXyXLH2FG2vZPVDzSL9kdA7uKWkDYHPZUKX4hbRu1BlhKEGvidbo/0R+M7LV+mlcOBmpbh9C7M6ruLFlnw6CCW1BCrcZLSTz6MIyUEeNPezjXG8QafBdPlrxV8AAAD//wMAUEsDBBQABgAIAAAAIQDfLqcEcwEAANIFAAAQAAAAeGwvY2FsY0NoYWluLnhtbHSU627CIBiG/y/ZPRD+T+Qwd4jVBJJ5A9sFkIq2SUtNaZbt7scW+Zgc/phIH5/yva+w3X+NA/o0s+sn22C6WmNkbDsde3tu8Mf728MzRm7R9qiHyZoGfxuH97v7u22rh1Z1urfIG6xrcLcsl1dCXNuZUbvVdDHWPzlN86gX/3U+E3eZjT66zphlHAhbrzdk9AK827ZobvBBPGLU+01gNPx+kuu6pOK6HleA9C+O5IGyhFS05txkZP6Wp+J+FM1/WyYl9en9TRR3/lJxZiTzVRTSUDQYwMl8YiUSDJEM+dwmrMAQSMXKpGQ8nYiF3FJnSqoKKVnoKO4zJJw6c7LSEUs7kiwknDqDIc5eJiXLkueVjgpkpSMwwOy80hEPhkiGhJOJwAAT8TIpefaf5+XzogpkpSMwxLeXScnT5CWvdFQgK+cIDJCSqHTEszZFpSMwwEQVUors/oGV5E4T2TkSlXMU12/uOvXfQOA23v0AAAD//wMAUEsDBBQABgAIAAAAIQAwIYPQ9gAAAEUBAAAZAAAAZG9jTWV0YWRhdGEvTGFiZWxJbmZvLnhtbFTQ3WqEMBAF4FcJuY+JrjGuqAu9K2yhrxCTcQ3kZzFTaSl998a79u4wMB9zZrx9Bk8O2LNLcaJ1JSiBaJJ18THRD1xZT0lGHa32KcJEvyDT2zwav/jB6wX83WUkBYl5OIcT3RCfA+fZbBB0roIze8ppxcqkwNO6OgO8EY3gwT3vp/AGqK1GTf+yxNmJfoPqhGmvLeukFKxtl5b1jeyZUqZW0irVL+rnvFgvHspCTUkA3FKJ77s7nIcH2FLAIbyenup0o0Fqdl1F8dZuYVcJlkl7uRTNtLari2dSRIj44jBPtHxkh5CO0y+ZzyP/337+BQAA//8DAFBLAwQUAAYACAAAACEA0B2j6k4BAAB/AgAAEQAIAWRvY1Byb3BzL2NvcmUueG1sIKIEASigAAEAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAjJJfS8MwFMXfBb9DybNt2k2nlLYDlT05EJz45+2S3G1xTRqSzG7f3rTdamUKPibn3F/OuSSb7mQZfKKxolI5SaKYBKhYxYVa5eR5MQtvSGAdKA5lpTAne7RkWpyfZUynrDL4aCqNxgm0gScpmzKdk7VzOqXUsjVKsJF3KC8uKyPB+aNZUQ1sAyukozieUIkOODigDTDUPZEckJz1SL01ZQvgjGKJEpWzNIkS+u11aKT9daBVBk4p3F77Toe4QzZnndi7d1b0xrquo3rcxvD5E/o6f3hqq4ZCNbtiSIqMs5QZBFeZYg5lKTZgPrYKLoI3kBpK2GwzOvA0+yzBurlf/VIgv93/PXZq9a+15bonkQc+btqVOyov47v7xYwUo3g0CeObMIkXyXUaX6WX4/cmyY/5Jn53IQ95/k+ceOiAeAQUGT35MsUXAAAA//8DAFBLAwQUAAYACAAAACEAb3gbEpEBAAAaAwAAEAAIAWRvY1Byb3BzL2FwcC54bWwgogQBKKAAAQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACckkFv2zAMhe8D9h8M3Rs53VAMgayiSDf0sGEBkrZnVqZjobIkiKyR7NdPtpHU2XbajXx8ePpESd0eOlf0mMgGX4nlohQFehNq6/eVeNx9u/oiCmLwNbjgsRJHJHGrP35QmxQiJrZIRY7wVImWOa6kJNNiB7TIY58nTUgdcG7TXoamsQbvg3nr0LO8LssbiQdGX2N9Fc+BYkpc9fy/oXUwAx897Y4xA2t1F6OzBjjfUv+wJgUKDRdfDwadkvOhynRbNG/J8lGXSs5btTXgcJ2DdQOOUMl3QT0gDEvbgE2kVc+rHg2HVJD9ldd2LYoXIBxwKtFDsuA5Yw22qRlrF4mTfg7plVpEJiWzYRLHcu6d1/azXo6GXFwah4AJJA8uEXeWHdLPZgOJ/0G8nBOPDBPvhLMd+KYz53zjlfNJf2SvQxfBH/Ua4h47662SJ0l9t/6VHuMu3APjaa+Xotq2kLDOT3He+1lQD3mlyQ0h6xb8HuuT5+/B8Auepq+ulzeL8lOZH3imKfn+qfVvAAAA//8DAFBLAQItABQABgAIAAAAIQCVZpbtkAEAAIIFAAATAAAAAAAAAAAAAAAAAAAAAABbQ29udGVudF9UeXBlc10ueG1sUEsBAi0AFAAGAAgAAAAhADEdic0iAQAA3gIAAAsAAAAAAAAAAAAAAAAAyQMAAF9yZWxzLy5yZWxzUEsBAi0AFAAGAAgAAAAhAD4Ms3q3AwAAeQkAAA8AAAAAAAAAAAAAAAAAHAcAAHhsL3dvcmtib29rLnhtbFBLAQItABQABgAIAAAAIQCSB5TsBAEAAD8DAAAaAAAAAAAAAAAAAAAAAAALAAB4bC9fcmVscy93b3JrYm9vay54bWwucmVsc1BLAQItABQABgAIAAAAIQDDG6yZDw8AAG1IAAAYAAAAAAAAAAAAAAAAAEQNAAB4bC93b3Jrc2hlZXRzL3NoZWV0MS54bWxQSwECLQAUAAYACAAAACEA9mC0QbgHAAARIgAAEwAAAAAAAAAAAAAAAACJHAAAeGwvdGhlbWUvdGhlbWUxLnhtbFBLAQItABQABgAIAAAAIQBowXh5+wcAADNSAAANAAAAAAAAAAAAAAAAAHIkAAB4bC9zdHlsZXMueG1sUEsBAi0AFAAGAAgAAAAhAI6gq1SSAQAAqgMAABQAAAAAAAAAAAAAAAAAmCwAAHhsL3NoYXJlZFN0cmluZ3MueG1sUEsBAi0AFAAGAAgAAAAhAN8upwRzAQAA0gUAABAAAAAAAAAAAAAAAAAAXC4AAHhsL2NhbGNDaGFpbi54bWxQSwECLQAUAAYACAAAACEAMCGD0PYAAABFAQAAGQAAAAAAAAAAAAAAAAD9LwAAZG9jTWV0YWRhdGEvTGFiZWxJbmZvLnhtbFBLAQItABQABgAIAAAAIQDQHaPqTgEAAH8CAAARAAAAAAAAAAAAAAAAACoxAABkb2NQcm9wcy9jb3JlLnhtbFBLAQItABQABgAIAAAAIQBveBsSkQEAABoDAAAQAAAAAAAAAAAAAAAAAK8zAABkb2NQcm9wcy9hcHAueG1sUEsFBgAAAAAMAAwABQMAAHY2AAAAAA=="""

# def get_default_template_bytes():
#     """Decode the hardcoded default template back into raw .xlsx bytes."""
#     return base64.b64decode(DEFAULT_TEMPLATE_B64.strip())


# # =====================================================================
# # 2. HELPER FUNCTIONS
# # =====================================================================

# def norm_name(name):
#     """Normalize a name to a comparable token set regardless of 'First Last'
#     vs 'Last, First' formatting differences between source files."""
#     parts = re.split(r"[,\s]+", str(name).strip().lower())
#     parts = [p for p in parts if p]
#     return " ".join(sorted(parts))


# def name_tokens(name):
#     parts = re.split(r"[,\s]+", str(name).strip().lower())
#     return {p for p in parts if p}


# def find_matching_rep_key(sow_name, rep_name_keys):
#     """Match a SOW resource name to a Replicon name key even when one side
#     has an extra middle name (e.g. SOW 'Kalyani Ghaytadkar' vs Replicon
#     'Ghaytadkar, Kalyani Popatrao'). A match is any rep key whose token set
#     fully contains the SOW name's tokens, or vice versa."""
#     sow_tok = name_tokens(sow_name)
#     best_match = None
#     for rep_key in rep_name_keys:
#         rep_tok = set(rep_key.split())
#         if sow_tok.issubset(rep_tok) or rep_tok.issubset(sow_tok):
#             # Prefer the closest-length match if multiple candidates exist
#             if best_match is None or abs(len(rep_tok) - len(sow_tok)) < abs(len(set(best_match.split())) - len(sow_tok)):
#                 best_match = rep_key
#     return best_match

# def get_field(row, candidates, default=""):
#     """
#     Flexible lookup of a value from a SOW row, regardless of small
#     header-naming differences (spacing, trailing colon, casing) between
#     what the code expects and what the uploaded SOW file actually has.

#     `candidates` is a list of possible header names to try, in priority
#     order. Returns `default` if none of them are found or all are blank.
#     """
#     def normalize(s):
#         return re.sub(r"\s+", " ", str(s).strip().lower().rstrip(":"))

#     normalized_map = {normalize(col): col for col in row.index}

#     for cand in candidates:
#         key = normalize(cand)
#         if key in normalized_map:
#             val = row[normalized_map[key]]
#             if pd.notna(val) and str(val).strip() != "":
#                 return val
#     return default


# # def load_sow(file):
# #     """Load the SOW & Resource Details file into a DataFrame."""
# #     df = pd.read_excel(file)
# #     df["name_key"] = df["Resource name"].apply(norm_name)
# #     return df
# def load_sow(file):
#     df = pd.read_excel(file)
#     name_col = None
#     for candidate in ["Resource name", "Employee Name", "Name"]:
#         for col in df.columns:
#             if re.sub(r"\s+", " ", str(col).strip().lower()) == candidate.lower():
#                 name_col = col
#                 break
#         if name_col:
#             break
#     if name_col is None:
#         raise ValueError("Could not find a resource/employee name column in the SOW file.")
#     df["name_key"] = df[name_col].apply(norm_name)
#     return df

# # def load_replicon(file):
# #     """Load the Replicon dump (Expenditure Details Report) into a DataFrame.
# #     Row 1 is an instructions banner, row 2 holds the real headers."""
# #     df = pd.read_excel(file, sheet_name="Expenditure Details Report", header=1)
# #     df = df.dropna(subset=["Employee Name/Supplier Name"])
# #     df["name_key"] = df["Employee Name/Supplier Name"].apply(norm_name)
# #     df["Item Date"] = pd.to_datetime(df["Item Date"], format="%d-%b-%Y")
# #     df["Quantity"] = pd.to_numeric(df["Quantity"], errors="coerce").fillna(0)
# #     return df

# # def load_replicon(file):
# #     """Load the Replicon dump into a DataFrame. Sheet is 'Sheet1' with
# #     headers on row 1. Employee name, date, and hours columns are renamed
# #     to the internal standard names used throughout the rest of the app."""
# #     df = pd.read_excel(file, sheet_name="Sheet1")
# #     df = df.dropna(subset=["Employee Name"])
# #     df["name_key"] = df["Employee Name"].apply(norm_name)
# #     df["Item Date"] = pd.to_datetime(df["Expenditure Item Date"])
# #     df["Quantity"] = pd.to_numeric(df["QTT"], errors="coerce").fillna(0)
# #     return df
# def load_replicon(file):
#     """Load the Replicon dump into a DataFrame. Sheet is 'Sheet1' with
#     headers on row 1. Employee name, date, and hours columns are renamed
#     to the internal standard names used throughout the rest of the app."""
#     # df = pd.read_excel(file, sheet_name="Sheet1")
#     xls = pd.ExcelFile(file)
#     if "Sheet1" in xls.sheet_names:
#         sheet_to_use = "Sheet1"
#     else:
#         sheet_to_use = xls.sheet_names[0]
#         st.warning(f"Replicon file has no sheet named 'Sheet1' — using the first sheet found: '{sheet_to_use}'.")
#     df = pd.read_excel(file, sheet_name=sheet_to_use)
    
#     df = df.dropna(subset=["Employee Name"])

#     # Drop exact duplicate line items: same Exp. Item Id AND same
#     # Expd Line Num means the same row got exported twice (not a real
#     # correction). Real corrections have the SAME Exp. Item Id but a
#     # DIFFERENT Expd Line Num (e.g. line 2 = reversal, line 3 = rebooking),
#     # so those are left alone.
#     before = len(df)
#     df = df.drop_duplicates(subset=["Exp. Item Id", "Expd Line Num"], keep="first")
#     dropped = before - len(df)
#     if dropped:
#         st.info(f"Removed {dropped} duplicate Replicon line(s) (same Exp. Item Id and Line Num).")

#     df["name_key"] = df["Employee Name"].apply(norm_name)
#     df["Item Date"] = pd.to_datetime(df["Expenditure Item Date"])
#     df["Quantity"] = pd.to_numeric(df["QTT"], errors="coerce").fillna(0)
#     return df

# # def build_daily_hours(rep_df):
# #     """Sum hours per employee per date (nets out negative correction rows),
# #     and track which (employee, date) pairs had a correction so we can flag
# #     them in the Remark column."""
# #     daily = rep_df.groupby(["name_key", "Item Date"])["Quantity"].sum().reset_index()
# #     neg = rep_df[rep_df["Quantity"] < 0][["name_key", "Item Date"]]
# #     neg_set = set(zip(neg["name_key"], neg["Item Date"]))
# #     return daily, neg_set

# def build_daily_hours(rep_df):
#     """Sum hours per employee per PROJECT per date (nets out negative
#     correction rows), and track which (employee, project, date) pairs
#     had a correction so we can flag them in the Remark column."""
#     daily = rep_df.groupby(["name_key", "Project Number", "Item Date"])["Quantity"].sum().reset_index()
#     neg = rep_df[rep_df["Quantity"] < 0][["name_key", "Project Number", "Item Date"]]
#     neg_set = set(zip(neg["name_key"], neg["Project Number"], neg["Item Date"]))
#     return daily, neg_set


# def determine_work_from(role):
#     """Determine Work From value from the Partner Emp. Role string.

#     Only fills "WFO" when the role string explicitly contains "onsite".
#     Everything else (offshore, unrecognized, blank) is left empty —
#     we don't guess WFH.
#     """
#     role_str = str(role).lower()

#     if "onsite" in role_str:
#         return "WFO"

#     return ""  # not onsite -> leave blank, don't assume WFH

# # def determine_work_from(role):
# #     """Determine WFH vs WFO from the Partner Emp. Role string.

# #     Rule (per business input): the role code contains a segment like
# #     'T&M_13' — the number right after 'T&M_'. Only the LAST digit of that
# #     number matters (the leading digit, e.g. the '1' in '13', is ignored):
# #         - last digit 3  -> WFO (work from office)
# #         - last digit 2  -> WFH (work from home)
# #     Falls back to a keyword check ("onsite" -> WFO) and then defaults to
# #     WFH if the pattern can't be parsed, so nothing breaks on unexpected
# #     role formats.
# #     """
# #     role_str = str(role)

# #     # Explicit keyword override, if present
# #     if "onsite" in role_str.lower():
# #         return "WFO"

# #     match = re.search(r"T&M[_\s]*(\d+)", role_str)
# #     if match:
# #         last_digit = match.group(1)[-1]
# #         if last_digit == "3":
# #             return "WFO"
# #         elif last_digit == "2":
# #             return "WFH"

# #     # Unrecognized pattern - default to WFH, but this should be reviewed
# #     return "WFH"

# # def fill_timesheet_for_employee(template_bytes, res_row, daily_hours, neg_set, month, year, matched_rep_key):
# def fill_timesheet_for_employee(template_bytes, res_row, daily_hours, neg_set, month, year, matched_rep_key, hours_used_so_far=0):
#     """Fill one copy of the timesheet template for a single employee and
#     return it as an in-memory .xlsx (BytesIO)."""
#     wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
#     # # ws = wb["Sheet1"]
#     # if "Sheet1" in wb.sheetnames:
#     #     ws = wb["Sheet1"]
#     # else:
#     #     ws = wb.active  # falls back to the workbook's active/default sheet
#     # if len(wb.sheetnames) > 1:
#     #     st.warning(f"Custom template has multiple sheets ({', '.join(wb.sheetnames)}). Using '{wb.sheetnames[0]}'.")
#     # ws = wb[wb.sheetnames[0]]
#     if "Actual template" in wb.sheetnames:
#         ws = wb["Actual template"]
#     elif len(wb.sheetnames) > 1:
#         st.warning(f"Custom template has multiple sheets ({', '.join(wb.sheetnames)}) and no sheet named 'Actual template' — using '{wb.sheetnames[0]}'.")
#         ws = wb[wb.sheetnames[0]]
#     else:
#         ws = wb[wb.sheetnames[0]]
#     ws.column_dimensions["H"].width = 32

#     name_key = matched_rep_key  # the Replicon-side key matched to this SOW resource

#     # ---- Pull all resource/PO/SOW fields from the SOW & Resource Details
#     # ---- file, tolerant of header-name variations in that file.
#     emp_name = get_field(res_row, ["Resource name", "Employee Name", "Name"])
#     cid = get_field(res_row, ["CID", "C.I.D", "C.I.D.", "Cid"])
#     role = get_field(res_row, ["Partner Emp. role", "Partner Emp Role", "PARTNER EMP. ROLE", "Emp Role", "Role"])
#     domain = get_field(res_row, ["Domain"])
#     po_number = get_field(res_row, ["PO Number", "PO Number:", "PO No", "PO No."])
#     po_start = get_field(res_row, ["PO Start Date", "PO Start Date:", "PO Start"])
#     po_end = get_field(res_row, ["PO End Date", "PO End Date:", "PO End"])
#     sow_number = get_field(res_row, ["SOW number", "SOW Number", "SOW Number:", "SOW No"])

#     work_from_value = determine_work_from(role)
#     planned_hours_raw = get_field(res_row, ["SOW planned duration hours", "SOW Planned Duration Hours", "Planned Hours"], None)
#     planned_hours = float(planned_hours_raw) if pd.notna(planned_hours_raw) and str(planned_hours_raw).strip() != "" else None

#     # ---- Header fields ----
#     ws["D4"] = emp_name
#     ws["D5"] = cid
#     ws["D6"] = role
#     ws["D7"] = domain
#     if po_number != "":
#         ws["D8"] = po_number
#     if po_start != "":
#         ws["D9"] = po_start
#     if po_end != "":
#         ws["D10"] = po_end
#     ws["D11"] = sow_number
#     ws["H4"] = calendar.month_name[month]
#     ws["H6"] = year

#     # ---- Daily rows (row 14 = day 1) ----
#     # emp_hours = daily_hours[daily_hours["name_key"] == name_key].set_index("Item Date")["Quantity"]
#     # sow_project_code = str(get_field(res_row, ["Project Code", "Project Number"], "")).strip()
#     sow_project_code_raw = get_field(res_row, ["Project Code", "Project Number"], "")
#     sow_project_code = re.sub(r"\.0$", "", str(sow_project_code_raw).strip())
#     emp_hours = daily_hours[
#         (daily_hours["name_key"] == name_key) &
#         # (daily_hours["Project Number"].astype(str).str.strip() == sow_project_code)
#         (daily_hours["Project Number"].astype(str).str.strip().str.replace(r"\.0$", "", regex=True) == sow_project_code)
#     ].set_index("Item Date")["Quantity"]
#     days_in_month = calendar.monthrange(year, month)[1]
#     date_number_format = ws["B14"].number_format  # capture template's date format before overwriting

#     proj_start_raw = get_field(res_row, ["Start date", "Start Date", "Project Start Date"], None)
#     proj_end_raw = get_field(res_row, ["End date", "End Date", "Project End Date"], None)

#     month_start = datetime(year, month, 1)
#     month_end = datetime(year, month, days_in_month)

#     proj_start = pd.to_datetime(proj_start_raw) if pd.notna(proj_start_raw) else month_start
#     proj_end = pd.to_datetime(proj_end_raw) if pd.notna(proj_end_raw) else month_end

#     effective_start = max(proj_start, month_start)
#     effective_end = min(proj_end, month_end)

#     for day in range(1, days_in_month + 1):
#         row = 13 + day
#         this_date = datetime(year, month, day)

#         if this_date < effective_start or this_date > effective_end:
#             for col in range(2, 9):
#                 ws.cell(row=row, column=col, value=None)
#             continue

#         hours = emp_hours.get(this_date, 0)
#         # remark = "Correction adjusted" if (name_key, this_date) in neg_set else ""
#         remark = "Correction adjusted" if (name_key, sow_project_code, this_date) in neg_set else ""
#         is_weekend = this_date.weekday() >= 5  # 5 = Saturday, 6 = Sunday

#         # Don't let total billed hours cross the SOW's planned duration hours
#         if planned_hours is not None and hours > 0:
#             remaining = planned_hours - hours_used_so_far
#             if remaining <= 0:
#                 hours = 0
#                 remark = (remark + "; " if remark else "") + "Exceeds SOW planned hours - not billed"
#             elif hours > remaining:
#                 hours = remaining
#                 remark = (remark + "; " if remark else "") + "Capped to SOW planned hours"
#             hours_used_so_far += hours

#         date_cell = ws.cell(row=row, column=2, value=this_date)
#         date_cell.number_format = date_number_format
#         ws.cell(row=row, column=3, value=this_date.strftime("%A").upper())

#         if is_weekend and hours == 0:
#             ws.cell(row=row, column=7, value=None)   # blank instead of 0
#         else:
#             ws.cell(row=row, column=7, value=float(hours))

#         ws.cell(row=row, column=8, value=remark)
#         remark_cell = ws.cell(row=row, column=8, value=remark)
#         remark_cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left") 

#         # Activity (D) and Work From (F) driven by whether hours were billed
#         if hours == 0:
#             if not is_weekend:
#                 ws.cell(row=row, column=4, value="On Leave")
#                 ws.cell(row=row, column=6, value="On Leave")
#             # weekend + 0 hours -> leave D and F blank, same as G
#         else:
#             ws.cell(row=row, column=4, value="Project work")
#             ws.cell(row=row, column=6, value=work_from_value)

#     # ---- Dropdown for Work From column (F) — WFO / WFH ----
#     # from openpyxl.worksheet.datavalidation import DataValidation
    
#     # for day in range(1, days_in_month + 1):
#     #     row = 13 + day
#     #     this_date = datetime(year, month, day)

#     #     if this_date < effective_start or this_date > effective_end:
#     #         for col in range(2, 9):
#     #             ws.cell(row=row, column=col, value=None)
#     #         continue


        
#     #     # hours = emp_hours.get(this_date, 0)
#     #     # remark = "Correction adjusted" if (name_key, this_date) in neg_set else ""
#     #     # is_weekend = this_date.weekday() >= 5  # 5 = Saturday, 6 = Sunday

#     #     # date_cell = ws.cell(row=row, column=2, value=this_date)
#     #     hours = emp_hours.get(this_date, 0)
#     #     remark = "Correction adjusted" if (name_key, this_date) in neg_set else ""
#     #     is_weekend = this_date.weekday() >= 5  # 5 = Saturday, 6 = Sunday

#     #     # Don't let total billed hours cross the SOW's planned duration hours
#     #     if planned_hours is not None and hours > 0:
#     #         remaining = planned_hours - hours_used_so_far
#     #         if remaining <= 0:
#     #             hours = 0
#     #             remark = (remark + "; " if remark else "") + "Exceeds SOW planned hours - not billed"
#     #         elif hours > remaining:
#     #             hours = remaining
#     #             remark = (remark + "; " if remark else "") + "Capped to SOW planned hours"
#     #         hours_used_so_far += hours

#     #     date_cell = ws.cell(row=row, column=2, value=this_date)
#     #     date_cell.number_format = date_number_format
#     #     ws.cell(row=row, column=3, value=this_date.strftime("%A").upper())

#     #     if is_weekend and hours == 0:
#     #         ws.cell(row=row, column=7, value=None)   # blank instead of 0
#     #     else:
#     #         ws.cell(row=row, column=7, value=float(hours))

#     #     ws.cell(row=row, column=8, value=remark)
#     #     date_cell = ws.cell(row=row, column=2, value=this_date)
#     #     date_cell.number_format = date_number_format
#     #     ws.cell(row=row, column=3, value=this_date.strftime("%A").upper())

#     #     if is_weekend and hours == 0:
#     #         ws.cell(row=row, column=7, value=None)   # blank instead of 0
#     #     else:
#     #         ws.cell(row=row, column=7, value=float(hours))

#     #     ws.cell(row=row, column=8, value=remark)

#     #     # Activity (D) and Work From (F) driven by whether hours were billed
#     #     if hours == 0:
#     #         if not is_weekend:
#     #             ws.cell(row=row, column=4, value="On Leave")
#     #             ws.cell(row=row, column=6, value="On Leave")
#     #         # weekend + 0 hours -> leave D and F blank, same as G
#     #     else:
#     #         ws.cell(row=row, column=4, value="Project work")
#     #         ws.cell(row=row, column=6, value=work_from_value)
#     #     # date_cell = ws.cell(row=row, column=2, value=this_date)
#     #     # date_cell.number_format = date_number_format
#     #     # ws.cell(row=row, column=3, value=this_date.strftime("%A").upper())

#     #     # if is_weekend and hours == 0:
#     #     #     ws.cell(row=row, column=7, value=None)   # blank instead of 0
#     #     # else:
#     #     #     ws.cell(row=row, column=7, value=float(hours))
#     #     # if hours == 0:
#     #     #     if not is_weekend:
#     #     #         ws.cell(row=row, column=4, value="On Leave")
#     #     #         ws.cell(row=row, column=6, value="On Leave")
#     #     #     # weekend + 0 hours -> leave D and F blank, same as G
            
#     #     # else:
#     #     #     ws.cell(row=row, column=4, value="Project work")
#     #     #     ws.cell(row=row, column=6, value=work_from_value)

#     #     # ws.cell(row=row, column=8, value=remark)

#     #     # # NEW: Activity (D) and Work From (F) driven by whether hours were billed
#     #     # if hours == 0:
#     #     #     ws.cell(row=row, column=4, value="On Leave")
#     #     #     ws.cell(row=row, column=6, value="On Leave")
#     #     # else:
#     #     #     ws.cell(row=row, column=4, value="Project work")
#     #     #     ws.cell(row=row, column=6, value=work_from_value)

#     # # ---- Dropdown for Work From column (F) — WFO / WFH ----
#     from openpyxl.worksheet.datavalidation import DataValidation
#     # dv = DataValidation(type="list", formula1='"WFO,WFH"', allow_blank=True)
#     dv = DataValidation(type="list", formula1='"WFO,WFH,On Leave"', allow_blank=True)
#     dv.error = "Please select WFO or WFH"
#     dv.errorTitle = "Invalid entry"
#     dv.prompt = "Select WFO or WFH"
#     dv.promptTitle = "Work From"
#     ws.add_data_validation(dv)
#     dv.add(f"F14:F{13 + days_in_month}")
#     # ---- Dropdown for Activity column (D) — with custom entry allowed ----
#     dv_activity = DataValidation(
#         type="list",
#         formula1='"On Leave,Sick Leave,Public Holiday,Project work"',
#         allow_blank=True,
#         showErrorMessage=False,  # allows typing a custom value beyond the 3 options
#     )
#     dv_activity.prompt = "Select an activity or type your own"
#     dv_activity.promptTitle = "Activity"
#     ws.add_data_validation(dv_activity)
#     dv_activity.add(f"D14:D{13 + days_in_month}")
#     # ---- Dropdown for Month (H4) ----
#     dv_month = DataValidation(
#         type="list",
#         formula1='"January,February,March,April,May,June,July,August,September,October,November,December"',
#         allow_blank=True,
#     )
#     dv_month.error = "Please select a valid month"
#     dv_month.errorTitle = "Invalid entry"
#     dv_month.prompt = "Select a month"
#     dv_month.promptTitle = "Month"
#     ws.add_data_validation(dv_month)
#     dv_month.add("H4")

#     # ---- Dropdown for Year (H6) ----
#     dv_year = DataValidation(
#         type="list",
#         formula1='"2024,2025,2026,2027,2028,2029,2030,2031"',
#         allow_blank=True,
#     )
#     dv_year.error = "Please select a valid year"
#     dv_year.errorTitle = "Invalid entry"
#     dv_year.prompt = "Select a year"
#     dv_year.promptTitle = "Year"
#     ws.add_data_validation(dv_year)
#     dv_year.add("H6")
#     # Blank out any leftover template rows beyond this month's day count
#     for row in range(14 + days_in_month, 45):
#         for col in range(2, 9):
#             ws.cell(row=row, column=col, value=None)
    
#     # ---- Sign-off block ----
#     ws["C52"] = emp_name
#     ws["F52"] = get_field(res_row, ["Capgemini Resposible", "Capgemini Responsible"])
#     ws["F56"] = get_field(res_row, ["EGA Resposible", "EGA Responsible"])
#     # ---- Lock only these specific cells; everything else stays editable ----
#     # locked_cells = ["D6", "D7", "D8", "D9", "D10", "D11"]

#     # # Unlock all cells first
#     # for row in ws.iter_rows():
#     #     for cell in row:
#     #         cell.protection = Protection(locked=False)

#     # # Lock only the required ones
#     # for coord in locked_cells:
#     #     ws[coord].protection = Protection(locked=True)

#     # # Enable sheet protection so locking actually takes effect
#     # ws.protection.sheet = True
#     # ws.protection.password = "yourpassword"  # optional, remove this line if no password needed
#     # ---- Freeze specific fields so employees can't edit them ----
#     # from openpyxl.styles import Protection

#     # locked_cells = ["D6", "D7", "D8", "D9", "D10", "D11"]

#     # for row in ws.iter_rows():
#     #     for cell in row:
#     #         cell.protection = Protection(locked=False)

#     # for coord in locked_cells:
#     #     ws[coord].protection = Protection(locked=True)

#     # ws.protection.sheet = True
#     # ws.protection.enable()

#     # out = io.BytesIO()
#     # # wb.save(out)

#     # # out = io.BytesIO()
#     # ---- Lock everything EXCEPT Activity (D) and Work From (F) daily cells ----
#     from openpyxl.styles import Protection

#     # editable_cols = {4, 6, 7}  # column D = Activity, column F = Work From
#     # editable_rows = set(range(14, 14 + days_in_month))  # daily table rows only

#     # for row in ws.iter_rows():
#     #     for cell in row:
#     #         is_editable = (cell.row in editable_rows) and (cell.column in editable_cols)
#     #         cell.protection = Protection(locked=not is_editable)
#     # ws["D5"].protection = Protection(locked=False)   # ← new line: unlocks CID

#     # ws.protection.sheet = True
#     # ws.protection.enable()
#     # editable_cols = {4, 6}  # column D = Activity, column F = Work From
#     editable_cols = {4, 6, 7}  # column D = Activity, column F = Work From, column G = Billable Hours
#     editable_rows = set(range(14, 14 + days_in_month))  # daily table rows only

#     for row in ws.iter_rows():
#         for cell in row:
#             is_editable = (cell.row in editable_rows) and (cell.column in editable_cols)
#             cell.protection = Protection(locked=not is_editable)

#     # Explicitly unlock header cells that people should be able to fill in/edit
#     # extra_unlocked_cells = ["D4", "D5", "H4" , "C52"]  # Name, CID, Month
#     # for coord in extra_unlocked_cells:
#     #     ws[coord].protection = Protection(locked=False)
#     # extra_unlocked_cells = ["D4", "D5", "H4", "C52"]  # Name, CID, Month, Signature
#     extra_unlocked_cells = ["D4", "D5", "C52", "F52"]  # Name, CID, Month, Signature, Capgemini Responsible
#     for coord in extra_unlocked_cells:
#         ws[coord].protection = Protection(locked=False)

#     # Month activities summary is a merged block (B48:H50) — unlock every
#     # cell in the merged range, since protection is per-cell even when merged
#     # for row in ws["B48:H50"]:
#     #     for cell in row:
#     #         cell.protection = Protection(locked=False)
#     # from openpyxl.styles import Alignment

#     # Month activities summary — unlock AND align text/cursor to top-left
#     for row in ws["B48:H50"]:
#         for cell in row:
#             cell.protection = Protection(locked=False)
#             cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)

#     ws.protection.sheet = True
#     ws.protection.enable()

#     # out = io.BytesIO()
#     # # wb.save(out)
#     # wb.save(out)
#     # out.seek(0)
#     # return out.getvalue()
#     out = io.BytesIO()
#     # wb.save(out)
#     wb.save(out)
#     out.seek(0)
#     return out.getvalue(), hours_used_so_far


# # =====================================================================
# # 3. STREAMLIT UI
# # =====================================================================
# # =====================================================================
# # 3. STREAMLIT UI
# # =====================================================================

# # st.title("📋 Timesheet Auto-Filler")
# # st.write(
# #     "Upload **SOW & Resource Details** and **Replicon Dump** files below. "
# #     "The app will automatically generate a filled timesheet for **every employee** "
# #     "found in the SOW file, and save each one with that employee's name."
# # )
# st.markdown("""
# <div class="hero-banner">
#     <div class="hero-icon">🗓️</div>
#     <div class="hero-text">
#         <h1>Timesheet Auto-Filler</h1>
#         <p>Upload <b>SOW & Resource Details</b> and <b>Replicon Dump</b> files below.
#         The app will automatically generate a filled timesheet for <b>every employee</b>
#         found in the SOW file, and save each one with that employee's name.</p>
#     </div>
# </div>
# """, unsafe_allow_html=True)
# st.markdown("""
# <div class="step-tracker">
#     <div class="step active"><div class="step-num">1</div><div class="step-label">Upload SOW</div></div>
#     <div class="step-line"></div>
#     <div class="step active"><div class="step-num">2</div><div class="step-label">Upload Replicon</div></div>
#     <div class="step-line"></div>
#     <div class="step"><div class="step-num">3</div><div class="step-label">Generate</div></div>
#     <div class="step-line"></div>
#     <div class="step"><div class="step-num">4</div><div class="step-label">Download</div></div>
# </div>
# """, unsafe_allow_html=True)

# col1, col2 = st.columns(2)
# with col1:
#     sow_file = st.file_uploader("1️⃣ Upload SOW & Resource Details (.xlsx)", type=["xlsx"])
# with col2:
#     replicon_file = st.file_uploader("2️⃣ Upload Replicon Dump (.xlsx)", type=["xlsx"])
# stat_col1, stat_col2, stat_col3 = st.columns(3)
# with stat_col1:
#     st.markdown(f"""<div class="stat-chip">📄 SOW File<br><b>{'✅ Uploaded' if sow_file else '⏳ Pending'}</b></div>""", unsafe_allow_html=True)
# with stat_col2:
#     st.markdown(f"""<div class="stat-chip">🕒 Replicon Dump<br><b>{'✅ Uploaded' if replicon_file else '⏳ Pending'}</b></div>""", unsafe_allow_html=True)
# with stat_col3:
#     st.markdown(f"""<div class="stat-chip">📋 Template<br><b>Default</b></div>""", unsafe_allow_html=True)

# # st.subheader("3️⃣ Timesheet Template")
# # template_option = st.radio(
# #     "Choose which timesheet template to fill:",
# #     ["Use built-in default template (hardcoded in app)", "Upload a custom template"],
# #     index=0,
# # )

# # custom_template_file = None
# # if template_option == "Upload a custom template":
# #     custom_template_file = st.file_uploader(
# #         "Upload Timesheet Template (.xlsx)", type=["xlsx"], key="template_upload"
# #     )
# # else:
# #     st.caption("✅ Using the built-in default timesheet template embedded in this app.")
# template_option = "Use built-in default template (hardcoded in app)"
# custom_template_file = None
# st.divider()
# generate_btn = st.button("🚀 Generate Timesheets", type="primary", use_container_width=True)

# if generate_btn:
#     if not sow_file or not replicon_file:
#         st.error("Please upload both the SOW & Resource Details and Replicon Dump files.")
#     elif template_option == "Upload a custom template" and not custom_template_file:
#         st.error("Please upload a custom timesheet template, or switch to the built-in default.")
#     else:
#         try:
#             with st.spinner("Reading files..."):
#                 sow_df = load_sow(sow_file)
#                 rep_df = load_replicon(replicon_file)
#                 daily_hours, neg_set = build_daily_hours(rep_df)

#                 # Auto-detect month/year from the Replicon dump's dates
#                 # month = int(rep_df["Item Date"].dt.month.mode()[0])
#                 # year = int(rep_df["Item Date"].dt.year.mode()[0])
#                 # Detect ALL distinct months present in the Replicon dump
#                 # (not just the single most-common one) so no billable
#                 # hours from any month get silently dropped.
#                 month_year_pairs = sorted(
#                     set(zip(rep_df["Item Date"].dt.year, rep_df["Item Date"].dt.month))
#                 )

#                 template_bytes = (
#                     custom_template_file.read() if custom_template_file else get_default_template_bytes()
#                 )

#             # st.success(
#             #     f"Detected period: **{calendar.month_name[month]} {year}**. "
#             #     f"Found **{len(sow_df)}** resource(s) in the SOW file."
#             # )

#             # generated_files = {}
#             # unmatched_employees = []
#             # rep_name_keys = daily_hours["name_key"].unique().tolist()
#             # progress = st.progress(0.0)
#             # for i, (_, res_row) in enumerate(sow_df.iterrows()):
#             #     # matched_key = find_matching_rep_key(res_row["Resource name"], rep_name_keys)
#             #     matched_key = find_matching_rep_key(get_field(res_row, ["Resource name", "Employee Name", "Name"]), rep_name_keys)
#             #     # if matched_key is None:
#             #     #     unmatched_employees.append(res_row["Resource name"])
#             #     if matched_key is None:
#             #         unmatched_employees.append(get_field(res_row, ["Resource name", "Employee Name", "Name"]))
#             #     data = fill_timesheet_for_employee(
#             #         template_bytes, res_row, daily_hours, neg_set, month, year, matched_key
#             #     )
#             #     # emp_name_clean = str(res_row.get("Resource name", f"Employee_{i+1}")).strip().replace(" ", "_")
#             #     emp_name_clean = str(get_field(res_row, ["Resource name", "Employee Name", "Name"], f"Employee_{i+1}")).strip().replace(" ", "_")
#             #     fname = f"Timesheet_{calendar.month_name[month]}_{year}_{emp_name_clean}.xlsx"
#             #     generated_files[fname] = data
#             #     progress.progress((i + 1) / len(sow_df))
#             months_label = ", ".join(f"{calendar.month_name[m]} {y}" for y, m in month_year_pairs)
#             st.success(
#                 f"Detected **{len(month_year_pairs)}** month(s) in the Replicon dump: **{months_label}**. "
#                 f"Found **{len(sow_df)}** resource(s) in the SOW file."
#             )

#             generated_files = {}
#             unmatched_employees = set()
#             rep_name_keys = daily_hours["name_key"].unique().tolist()

#             total_steps = len(sow_df) * len(month_year_pairs)
#             progress = st.progress(0.0)
#             step = 0

#             for _, res_row in sow_df.iterrows():
#                 emp_display_name = get_field(res_row, ["Resource name", "Employee Name", "Name"], "Employee")
#                 # matched_key = find_matching_rep_key(emp_display_name, rep_name_keys)
#                 # if matched_key is None:
#                 #     unmatched_employees.add(emp_display_name)
#                 #     hours_used_so_far = 0
#                 matched_key = find_matching_rep_key(emp_display_name, rep_name_keys)
#                 hours_used_so_far = 0
#                 if matched_key is None:
#                     unmatched_employees.add(emp_display_name)

#                 for year, month in month_year_pairs:
#                     step += 1
#                     progress.progress(step / total_steps)

#                     # Skip generating a file for months with zero overlap
#                     # between this employee's SOW range and the month
#                     proj_start_raw = get_field(res_row, ["Start date", "Start Date", "Project Start Date"], None)
#                     proj_end_raw = get_field(res_row, ["End date", "End Date", "Project End Date"], None)
#                     days_in_month = calendar.monthrange(year, month)[1]
#                     month_start = datetime(year, month, 1)
#                     month_end = datetime(year, month, days_in_month)
#                     proj_start = pd.to_datetime(proj_start_raw) if pd.notna(proj_start_raw) else month_start
#                     proj_end = pd.to_datetime(proj_end_raw) if pd.notna(proj_end_raw) else month_end
#                     effective_start = max(proj_start, month_start)
#                     effective_end = min(proj_end, month_end)
#                     if effective_start > effective_end:
#                         continue  # no overlap with SOW range -> don't generate a blank file

#                     # data = fill_timesheet_for_employee(
#                     #     template_bytes, res_row, daily_hours, neg_set, month, year, matched_key
#                     # )
#                     data, hours_used_so_far = fill_timesheet_for_employee(
#                         template_bytes, res_row, daily_hours, neg_set, month, year, matched_key, hours_used_so_far
#                     )
#                     # emp_name_clean = str(emp_display_name).strip().replace(" ", "_")
#                     emp_name_clean = re.sub(r"[^\w\-]", "_", str(emp_display_name).strip())
#                     fname = f"Timesheet_{calendar.month_name[month]}_{year}_{emp_name_clean}.xlsx"
#                     generated_files[fname] = data

#             unmatched_employees = list(unmatched_employees)
#             if unmatched_employees:
#                 st.warning(
#                     "⚠️ No matching hours found in the Replicon dump for: "
#                     + ", ".join(unmatched_employees)
#                     + ". Their timesheet was generated with all days marked as Leave — "
#                     "double check the name spelling in both files."
#                 )


#             # st.subheader("✅ Generated Timesheets")
#             # for fname, data in generated_files.items():
#             #     st.download_button(
#             #         label=f"⬇️ Download {fname}",
#             #         data=data,
#             #         file_name=fname,
#             #         mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#             #         key=fname,
#             #     )

#             # # Zip download of all files together
#             # zip_buffer = io.BytesIO()
#             # with zipfile.ZipFile(zip_buffer, "w") as zf:
#             #     for fname, data in generated_files.items():
#             #         zf.writestr(fname, data)
#             # zip_buffer.seek(0)

#             # st.download_button(
#             #     "📦 Download All Timesheets as ZIP",
#             #     data=zip_buffer.getvalue(),
#             #     # file_name=f"All_Timesheets_{calendar.month_name[month]}_{year}.zip",
#             #     file_name="All_Timesheets.zip",
#             #     mime="application/zip",
#             #     use_container_width=True,
#             # )
#             zip_buffer = io.BytesIO()
#             with zipfile.ZipFile(zip_buffer, "w") as zf:
#                 for fname, data in generated_files.items():
#                     zf.writestr(fname, data)
#             zip_buffer.seek(0)

#             st.session_state["generated_files"] = generated_files
#             st.session_state["zip_bytes"] = zip_buffer.getvalue()

#         except Exception as e:
#             st.error(f"Something went wrong: {e}")
#             st.exception(e)
# if "generated_files" in st.session_state:
#     st.divider()
#     st.subheader("✅ Generated Timesheets")

#     st.download_button(
#         "📦 Download All Timesheets as ZIP",
#         data=st.session_state["zip_bytes"],
#         file_name="All_Timesheets.zip",
#         mime="application/zip",
#         use_container_width=True,
#         type="primary",
#     )

#     # # with st.expander(f"Or download individual files ({len(st.session_state['generated_files'])})"):
#     # with st.expander(f"Or download individual files ({len(st.session_state['generated_files'])})", expanded=True):
#     #     for fname, data in st.session_state["generated_files"].items():
#     #         st.download_button(
#     #             label=f"⬇️ {fname}",
#     #             data=data,
#     #             file_name=fname,
#     #             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#     #             key=fname,
#     #         )
#     with st.expander(f"Or download individual files ({len(st.session_state['generated_files'])})", expanded=True):
#         for fname, data in st.session_state["generated_files"].items():
#             display_name = fname.replace("Timesheet_", "").replace(".xlsx", "").replace("_", " ")
#             card_col1, card_col2 = st.columns([4, 1])
#             with card_col1:
#                 st.markdown(f"""
#                     <div class="file-card">
#                         <div class="file-icon">📊</div>
#                         <div class="file-name">{display_name}</div>
#                     </div>
#                 """, unsafe_allow_html=True)
#             with card_col2:
#                 st.download_button("⬇️", data=data, file_name=fname,
#                     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=fname)
#     st.markdown("""
# <div class="app-footer">
#     <span>⚡ Powered by Capgemini Automation</span>
# </div>
# """, unsafe_allow_html=True)

# # st.divider()
# # with st.expander("ℹ️ How field mapping works"):
# #     st.markdown(
# #         """
# #         - **Partner Emp. Name / CID / Role / Domain / PO Number / SOW Number** ← from SOW & Resource Details file
# #         - **Daily Billable Hours** ← summed per employee per date from the Replicon dump
# #           (negative correction rows are netted automatically)
# #         - **Activity** ← SOW name, filled only on days with hours > 0
# #         - **Work From** ← "WFO" if role contains "onsite", otherwise "WFH"
# #         - **Days with 0 hours** ← marked as "Leave"
# #         - **Capgemini/EGA Responsible, sign-off Name** ← from SOW & Resource Details file 
# #         """
# #     )
# st.divider()
# with st.expander("ℹ️ How field mapping works"):
#     st.markdown(
#         """
#         - **Partner Emp. Name / CID / Role / Domain / PO Number / SOW Number** ← from SOW & Resource Details file
#         - **Daily Billable Hours** ← summed per employee per date from the Replicon dump
#           (negative correction rows are netted automatically)
#         - **Activity** ← "Project work" on days with billable hours, "On Leave" on 0-hour weekdays
#         - **Work From** ← "WFO" only if role contains "onsite"; left blank otherwise (not auto-filled as WFH)
#         - **Days with 0 hours (weekdays)** ← marked as "On Leave"
#         - **Weekends with 0 hours** ← left blank
#         - **Capgemini/EGA Responsible, sign-off Name** ← from SOW & Resource Details file 
#         """
#     )
# """
# Timesheet Auto-Filler - Streamlit App
# =======================================
# Upload SOW & Resource Details + Replicon Dump files, and this app will
# automatically generate one filled Timesheet per employee.

# Features:
# - Upload SOW & Resource Details (.xlsx)
# - Upload Replicon Dump (.xlsx)
# - Timesheet template: use the BUILT-IN default template (hardcoded/embedded
#   in this file as base64) OR upload your own custom template (.xlsx)
# - Generates one filled timesheet per employee found in the SOW file
# - Download each file individually, or all together as a ZIP

# Run with:
#     streamlit run timesheet_app.py
# """

# import streamlit as st
# import pandas as pd
# import openpyxl
# import re
# import io
# import base64
# import calendar
# import zipfile
# from datetime import datetime
# from openpyxl.worksheet.protection import SheetProtection
# from openpyxl.styles import Protection

# st.set_page_config(page_title="Timesheet Auto-Filler", layout="wide", page_icon="📋")

# # =====================================================================
# # 1. HARDCODED DEFAULT TIMESHEET TEMPLATE (base64-embedded .xlsx)
# #    This lets the app work out-of-the-box with no template upload needed.
# #    A second option below lets the user upload a custom template instead.
# # =====================================================================
# DEFAULT_TEMPLATE_B64 = """
# UEsDBBQABgAIAAAAIQCVZpbtkAEAAIIFAAATAAgCW0NvbnRlbnRfVHlwZXNdLnhtbCCiBAIooAACAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACslM9O4zAQxu8r7TtEvqLEhcMKoaYc+CMh7QIS7AMM9qSx6tiWZ4D27Zm4gBDqtqzoJVZiz/f9PJmZ6ely8NUTZnIxtOqwmagKg4nWhXmr/t5f1seqIoZgwceArVohqdPZzx/T+1VCqiQ6UKt65nSiNZkeB6AmJgyy08U8AMtrnusEZgFz1EeTyS9tYmAMXPOooWbTc+zg0XN1sZTPa5KMnlR1tj44erUKUvLOAAupfgr2k0v96tBIZDlDvUt0IBhKb3QYd/5t8Bp3I6nJzmJ1C5mvYRAMvfT6OebFQ4yLZrvIBsrYdc6gjeZxkAw0lDKCpR6RB9+UtRnAhTfuLf7lMOmyHO4ZZLxfEd7BwfK/UZfn9xGKzA5D4pVH2nfai+gu5x4y2jvO0hl7B/iovYPDgDdnvZTInpPwrrvNX+r2DzJYYNC/4QH9VejiF0AGqteF3xgPRE56oPSoHzXem3RTsYvjbY6JZGZk/P8rvw2FMbpOIoSZHX7NUebNt3OM40SzaDd46zJBZy8AAAD//wMAUEsDBBQABgAIAAAAIQAxHYnNIgEAAN4CAAALAAgCX3JlbHMvLnJlbHMgogQCKKAAAgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAArJLRSgMxEEXfBf8h5L0721VEpNu+iFBQEKkfME1mt6FJJiRR2783rRZdqEXQx2Tu3Jy5k8ls46x4pZgM+1aOq1oK8oq18X0rnxd3o2spUkav0bKnVm4pydn0/GzyRBZzaUorE5IoLj61cpVzuAFIakUOU8WBfKl0HB3mcow9BFRr7Amaur6C+N1DTgeeYq5bGef6QorFNpSX/+INjjJqzAiKI41CLGQxmzKLWGDsKbdSs3os12mvqAq1hONAzQ9AzqjIibtcKXbAXWfUbsymhroZTgrKYkqmCPYJWlySHZI8HHDvd7W57/gU0fj3EX1g3bJ6ceTzkS18gh8UX/lsLLxxXC+Z16dYLv+ThTaZvCZ9emEYwoEIBr9y+g4AAP//AwBQSwMEFAAGAAgAAAAhAD4Ms3q3AwAAeQkAAA8AAAB4bC93b3JrYm9vay54bWysVm1vozgQ/n7S/QfEdxcbDAHUdBXedJXaVZVm2zup0soFp0EBnDNOk6ra/75jAknbnE657kUJxvbw+JmZZ4acf9nWlfHMZVuKZmySM2wavMlFUTZPY/PbLEO+abSKNQWrRMPH5gtvzS8Xv/92vhFy+SjE0gCAph2bC6VWoWW1+YLXrD0TK97AzlzImimYyierXUnOinbBuaory8bYs2pWNuYOIZSnYIj5vMx5IvJ1zRu1A5G8Ygrot4ty1Q5odX4KXM3kcr1CuahXAPFYVqV66UBNo87Dy6dGSPZYgdtb4hpbCV8PfgTDxR5Ogq2jo+oyl6IVc3UG0NaO9JH/BFuEvAvB9jgGpyFRS/LnUudwz0p6n2Tl7bG8AxjBv4xGQFqdVkII3ifR3D0327w4n5cVv9tJ12Cr1VdW60xVplGxVqVFqXgxNkcwFRv+bkGuV9G6rGDXwcQhpnWxl/ONNAo+Z+tKzUDIA/zYtLHtYKwtQRiTSnHZMMVj0SjQYe/Xr2quw44XAhRuTPnf61JyKCzQF/gKV5aH7LG9YWphrGU1NuPw4VsL7j+81KyqyiV7SMSmqQTU2MMbcbLjSvgP8mS59tkCp3fEdvcfAwD8ZDhI8EZJA+4vkytIwy17hqRA6ou+Zi8h6v73V4c6MfGdBI2CNEEUTwLkByRCmZf4qR+7JKWTH+CF9MJcsLVa9InWmGOTQlaPtq7ZdtghOFyXxeH8V9x/kB4/XIa9H9pT3dLuSr5pD5LQU2N7XzaF2IxNRLSQX95PN93mfVmoBTTRgNpgslv7g5dPC2BMiAvy0X1DMxubry5xUpplPoojb4Jo4HrIx06E/HSSxK7tZI7ndoysN5S65gnUutFoOsHf6oZKoEvrUUcX7mWoz5CXRadta3gsZ1UOAtdDZxgQbAfaa75VV63qRtBWCfQIpGSEA4pw6riI+oGNfOrYKKaJnbqjNEkjoDcU9P/QAjuJh8NbRbNcMKlmkuVLeBdN+TxiLShp5xDwfUs2cv0IO0CRZiRDlAQYRZFHkZtkjjsiSZy62YGsdn/+yQbkW93TnKk1FKeuy24e6mvWr+4X57uFPk/vii6cJjru/dP/ZngL3lf8ROPs7kTD+Ov17PpE26t09v0+O9V4ch0lk9PtJ9Pp5K9Z+udwhPWPAbU+JDwhNMBOOkGOE1NER9kI+Rl2kUNHNHZplBI8OiS82uTPn8u3Ta1BkfHb/wh9M9L51+Bh/wfKaLnqt+CN0TXOjrim39XXHu3iJwAAAP//AwBQSwMEFAAGAAgAAAAhAJIHlOwEAQAAPwMAABoACAF4bC9fcmVscy93b3JrYm9vay54bWwucmVscyCiBAEooAABAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAKySy2rEMAxF94X+g9G+cTJ9UIZxZtFSmG2bfoBwlDhMYgdbfeTva1I6ycCQbrIxSML3Hom72393rfgkHxpnFWRJCoKsdmVjawXvxcvNI4jAaEtsnSUFAwXY59dXu1dqkeOnYJo+iKhigwLD3G+lDNpQhyFxPdk4qZzvkGPpa9mjPmJNcpOmD9LPNSA/0xSHUoE/lLcgiqGPzv9ru6pqND07/dGR5QsWMvDQxgVEgb4mVvBbJ5ER5GX7zZr2HM9Ck/tYyvHNlhiyNRm+nD8GQ8QTx6kV5DhZhLlfE0Zjq58MNnaCObWWLnK3aigMeirf2MfMz7Mxb//ByLPY5z8AAAD//wMAUEsDBBQABgAIAAAAIQDDG6yZDw8AAG1IAAAYAAAAeGwvd29ya3NoZWV0cy9zaGVldDEueG1spFxrb9vIFf1eoP9BIBaFg40t8SXJqq2FJVGSgc1mkTi724+KTMdCJNOV5Dxa9L/3zMy94uXMOKbpoq3iM/femXvmdUgOefbLt8269SXf7lbF3XkQnnSCVn63LK5Xd5/Ogw9X0+N+0NrtF3fXi3Vxl58H3/Nd8Mvw7387+1psP+9u83zfQoS73Xlwu9/fD9rt3fI23yx2J8V9foeSm2K7Wezx5/ZTe3e/zRfX2mmzbkedTre9WazuAhNhsK0To7i5WS3zSbF82OR3exNkm68Xe7R/d7u633G0zbJOuM1i+/nh/nhZbO4R4uNqvdp/10GD1mY5uPx0V2wXH9fI+1uYLJatb1v8N8L/Yq5G405Nm9VyW+yKm/0JIrdNm930T9un7cXyEMnNv1aYMGlv8y8r1YFlqKhZk8L0ECsqg8UNg3UPwRRd28HD6vo8+G+UJtHkojM97l10kuPk9AJj7LTbO06TbBpPo36/34n+FwzPrlfoYZVVa5vfnAejcHDZ7QTt4ZkeQH+s8q878e/WfvHxfb7Ol/sclYRBS43Pj0XxWRleAuog5E4bqJC7f+ugF+Hgr+kk7CT9tNdVsduH4PLfXNFUD+bft63r/GbxsN6/K77O89Wn270K3/pPvi34TzRAj4zB9fdJvltiqMLkJE5VHctijYD4/9ZmpeYchtrim2nz6np/q2Zh1D1N4zBKg9byYbcvNn9SAbkbR/SQdsTvV1N+etKLumn3CT90hvbDL/mFaP0PKkrIAb/kEHVq1YT265rwyzX1Trpx1H8qtS454pcd41o19sgRv+TYP4nr1IhlTjcVv2VT67B5So74Zccfdl/bdL8eZ5PFfjE82xZfW5j7GAe7+4VaSaMBgqlBhfRNt9QZVhQGo8EOEyKOZzQmmBFLVfNI+egGoM4d0C/Dzln7C2bAkizGxiLS40+5TGwgs4GpDcxsYG6AWM87tP7ABAamk0JygkHkn1KchXI7DzBID1mEVhZsoSahzsIAiHxwiaouGVuwy9QGZjYwN4BZTmRaaFmTtJTbeYABfWhjbKXFFoe0DNA/9FZmgNMDMDWAWY4UEzMDqIXgUE1SrWZOJhinB5P0YNKWmYLQJpkqNwxDDItDBV0rVTLRY8Z0oUFErgYQuRpA5EpB9Gaig8wJCZ2hiJnjyaR7WJifnJg8OFUg5Cb57Vm5HUwO/UhIUnakg0wNIrLz1NS3epJMQj3Vo05UslzpRoy5Jt2o3KxuPLVSJRPRjYSUeWSE6Hmke2lqEJGqDcwN4E49JWKttTWssaIoN2tFsRdGNjn0mQHCnuY2STud6DRMQqsHMrIqh+3UICI5G5gbwE0Om0WT5JQbugk/5XppL5gHm0N6BomUmvoyTNKkZ60SGbmIzAwiMqMYeraZ+UeI3l8qQ1BNmCbJaT87O2ttH5dGh/QI4vx6ccfq8IydRIIEiQw5jEiRILf7Qs/GX2dwaj87RXtrKI3KFFV954FYMclILJmEyIyMVyQzMpAnI58GqTHd1Mqv5E+5LIwJKhsycZDMQaYcqKIFrIE6Y6MUQ/lm+P7Dm6NZmAxmSfLqrH3j0UBzdpD7U/jIDhhaGgZpxYCMmjOXEVpwKZUX+eUZLhZ489DRjGE5V+2tkY2kUgitPWbCRlhLykj24sRGckAYURPL/Su0VvUZ+VWMImv+zNlIyojIWncu2cizIDRUUaFRL7ESvTfDycVVdjTvvn7z9rer+dE8+cdic//P8NXrkPs+6YahM5coBMnF/VaF0jTeLrb5dWCuU8cYRONEma70ReeH33/P3h1dZX9dHY3C5HUwwX+CV1zPn9nkt+z95OJf1Q1yws0td7+MoV6p4hgq5zKGsJ5Dcdl3c4KSUu9cMuQKnrChdtN+2CsjLymjMB2MErVegBTIDjDxM/pcTzNFtTU3xxwNo04NU5dqQy7WtC/Dq/mHdz4GjaaLJYMESQYJkgwaCF3IS+acm1MuQ5cMeRj0a0ZE/fHli7pKU2rDYdBwpnNVXJUrjt47x+xXh6vpu0vPWDMVV5giSDJFkGSKWiyZIkgyRZCHqYYCE+KqFlPOCkl+dZh6f3H14Z2HKwohRxVBkisDxZIrspJcEVRaXXJmHq4aStiQBOpTo8q+GGG/Wlx9+M3DFFUsmSJIMmWgClNkJZkiSI4qgjxMNdTDSgnXmX/WbjlmvzpMYb/xMGUqrsw/giRTBMkxRS2WTBEkmSLIZQoqt5G41n41Vir7qo/9KkwN5QYZdewN8uqDb3ukSJIyhgRlDAnKuAmCMoYEZQx5KGso1qGb6wwuWy2N2a/O4HpMS1CMClmmObEkiyBJFjVakkWQJIsgD1kNrwMicx3w1E5oq8Yx+71ANVCICld0m1JyRZDkitosuSJIrO/cQg9X1gWDujCocxWIi6VaA8u+9mW/xqqBAlSYoosEyRRBkilqsWSKIDmqCPIw1fAKIDLK+MlRZct+9nuBauAQYi9kSHJF0l1yRW2WXBEkuSLIw1VDNa+eINXYCyNbt7NfY9XAASRT1BbJlKva2VEyRY6SKYI8TDVU7erivQ5Ttmpnv8aqgQNIpqgtkilXtbOjZIoc5Ur1qGqPGqp27fe0ahA3o831DfvVWtX9asEV7RxUUuWKdraSVLmina08g6qhaI/qifbIFu3s9yK14Op2DivJcnU7W0myXN3OVh6yGur2qJ5uj2zdzn4vUQtUtZyDBEmuXOXOdUuuXOXOVi5X6hZbo0ekyq/GHLSVu64Pfo3VAgcQTDEkmCJIXg2ylWCKIbGuM+RhqqFgj+sJ9th+HsR+L1ALHEJyRa2RXLl6nR0lV65eZysPVw31elxPr8fO43jya6wWuGLJFMWUTLlqnR0lU+QoRxVBHqaUOm3wQFE9A6gx/2JbrbNfY7XAASRT1BbJlIEq84+sJFMESaYI8jDVUK3jRn0tpmy1zn7N1QJHkFRRYyRVrlhnR0mVK9bZykNVQ7GOB0W1qLLFOvu9RC1wDEmWq9fJqjKu3LvsHEuOq0f1Oq47m83Aeno9tvW6rq/mDvjIEwkOIblyFTtZVbhy77NzLMnVo4odt3macVXvPnts32fX9b1ILbiCnWPKKegKdraSU5BiSaYI8kzBhoIdvVVrCtqCnf1eohZcvc5RJVeuXmcryZWr19nKw1VDvY5Hk7W4svU6+zVXC65a55iSKVets5VkylXrbOUyhSevzU7+1VPrsa3WdX0155/3iQQHECsVQ4IpguRKxVaCKYbE/GPIw1RDta4fLT99XZPYap39mqsFjiCpcsU6WVWocm+ucyxJ1aM313Efs9mgkmJ9eDk9wgOpo+ytOQnx0+gnnFHovHp1HvVfB8FrPLP/WZ6JSGwJr1tRc6g99nyCY0gKXRVPVhUK3XvuHEtS+KiKV5ezjU7kShWvKBwleN8iUIQ9zuep4TMCn+KQSWIrfd2ml+kMDiH5dLU+WVX4dLU+x5J8Pqr19amXBldF2g85Y00tT3sYXuOneI3xxFCN07jkVb3GcTi4TrFTOrj0o5Mk0m/CfnqN0jciM4bKU/BThsoTazg/pre3VC6BBGkSK8cbk4aSX/udB2nZxWOGygV64kIZQ9joyyPv1hiclkZ8AmZGEO6y1zwsxzH0OaRqzg2Ve2L0rXkHyNwaJqgcnBMHyRxk6iAzB5kTgud1alRUTvE3FNOJEZ2V5htINt9GMnIrbaYOMnOQOSG+5nsUbhI//XJFYqRhF/tROXCcFxHISLwnwn7lDMlcaOpCMxeaM6SnVrVTpBZFI+o+PVVLDu7HVDqFINkrBMkF1YGmHKu0mrnQnCE9TasvTUiV+Iwc1NJm5cBQ+TxnwlB5LjBzoakLzVxoTlBPHyis5iD123NyMDJH9kNqINENDpI5yNRBZg4yJ8QzOVKpqcrmN3jdQ0fCWzuV6WLfgGEjMV0YEtOFoXJDmbqOM9dxXnGsdpMUPi/L04gBeWQ7NVBP9pwDZa7VlCGs8+USY92HmZVGh9OZDHmmlLz5+JzhaDbsXmWXdF5FIiOZpwNl0AFqdgo2pgSVfjMHmRPiG6NSMjwnJ3MrrTLFDFTpKQfKUgeaEiQzMEYlMicbXwZSADwnA1cAQABpdkk5mveZ7DPvbCSWb4bE1SxBMicTXOZkEF9OUhU8Jyez4ffLs9njlKBSeU5cKHOhqQvNXGjOkKtsUikNKAck/Ow3UNNnbsaYyPTuKgZneUFWg0WlzcSLtJt8+ykf5+v1rrUsHtQr11jqh2cHmA7op91BZpSzVTJJ4kGGKx6EdUoSlOgV2I6W9AdqQ3R9xmmEejTNlg/Wb/j4SsYpWoCZ40bD6ggfb9vi3iAzJ/nsVsd9lGgR4JScokRrASfTDjL15TNJQpToUen4IFNceXt4i+GDB46+EvjgMZyvBBzgeZWvBL2A5ye+khQlPt4mMXrbHFKzWx3BBxc0nmgRfHBMxlcCrnHew1cCrnG4wVcCrvEo39dqcI1H156SED44WOyLBh+cT/WVgGvzfpaTKbg2i5ZTAq7N5u2UgGscUvO1DT54QcdXAh+8WuIrAdd4ZcJXAq7xCoOvBFzjQLyvBFzjgLpbMuoNxl4PhPLho/5g7I1zOhj72B+FnYF6bdBTcxiixDvSUZJ5S0ZYBnwzYBQPxv5VCNPMV3c6GHu5BeleHJRrxtvliqk+SHF3vVKfUVmszQtqe3wDhj+UgReSBrgn0prid46FEOY37x7WeWv//R5fR8m/4dMuO/MJkutvN+Z7G6377arY4oMq+pWfM/UlmIf1Yvj23dFPeEnqPOBXLILX9Ld+jSBQd8fIFPdxdC3qho6vdfhICL6g8MdivcKv+gAMr/tq2a8WUUOv8+Vqs1gHrcV6XXwdrRd3n82XF26Lr5d39w/7N0hj8QkpqcsPgNl2W2wlmCvgarVXH4UZXf7668Xo16w1f4t3kPDlHFUGeLVeq6/GtG6Lh+2utcEnPFof81b/JG0V29YaNZwguPkCCb1zKL6P0u+k0WTU6x73o3H/OMEXUY5xe+viOBn38a7BRSfsjrrq+yhEUqi+1XD4N6PREA8FCI1AX5UMB8D3SO6R9ZvF9tMKLK7zG2ye+MYH3mmjz5uof++Le41C630s9vgsCf91i8/65EgcXzgJWjdFsec/ME/ahw8FDf8PAAD//wMAUEsDBBQABgAIAAAAIQD2YLRBuAcAABEiAAATAAAAeGwvdGhlbWUvdGhlbWUxLnhtbOxazY8btxW/B8j/QMxd1szoe2E50Kc39u564ZVd5EhJlIZeznBAUrsrFAEK59RLgQJp0UuB3nooigZogAa55I8xYCNN/4g8ckaa4YqKvf5AkmJ3LzPU7z3+5r3HxzePc/eTq5ihCyIk5UnXC+74HiLJjM9psux6TybjSttDUuFkjhlPSNdbE+l9cu/jj+7iAxWRmCCQT+QB7nqRUulBtSpnMIzlHZ6SBH5bcBFjBbdiWZ0LfAl6Y1YNfb9ZjTFNPJTgGNQ+WizojKCJVund2ygfMbhNlNQDMybOtGpiSRjs/DzQCLmWAybQBWZdD+aZ88sJuVIeYlgq+KHr+ebPq967W8UHuRBTe2RLcmPzl8vlAvPz0MwpltPtpP4obNeDrX4DYGoXN2rr/60+A8CzGTxpxqWsM2g0/XaYY0ug7NKhu9MKaja+pL+2wznoNPth3dJvQJn++u4zjjujYcPCG1CGb+zge37Y79QsvAFl+OYOvj7qtcKRhTegiNHkfBfdbLXbzRy9hSw4O3TCO82m3xrm8AIF0bCNLj3FgidqX6zF+BkXYwBoIMOKJkitU7LAM4jiXqq4REMqU4bXHkpxwiUM+2EQQOjV/XD7byyODwguSWtewETuDGk+SM4ETVXXewBavRLk5TffvHj+9Yvn/3nxxRcvnv8LHdFlpDJVltwhTpZluR/+/sf//fV36L///tsPX/7JjZdl/Kt//v7Vt9/9lHpYaoUpXv75q1dff/XyL3/4/h9fOrT3BJ6W4RMaE4lOyCV6zGN4QGMKmz+ZiptJTCJMLQkcgW6H6pGKLODJGjMXrk9sEz4VkGVcwPurZxbXs0isFHXM/DCKLeAx56zPhdMAD/VcJQtPVsnSPblYlXGPMb5wzT3AieXg0SqF9EpdKgcRsWieMpwovCQJUUj/xs8JcTzdZ5Radj2mM8ElXyj0GUV9TJ0mmdCpFUiF0CGNwS9rF0FwtWWb46eoz5nrqYfkwkbCssDMQX5CmGXG+3ilcOxSOcExKxv8CKvIRfJsLWZl3Egq8PSSMI5GcyKlS+aRgOctOf0hhsTmdPsxW8c2Uih67tJ5hDkvI4f8fBDhOHVypklUxn4qzyFEMTrlygU/5vYK0ffgB5zsdfdTSix3vz4RPIEEV6ZUBIj+ZSUcvrxPuL0e12yBiSvL9ERsZdeeoM7o6K+WVmgfEcLwJZ4Tgp586mDQ56ll84L0gwiyyiFxBdYDbMeqvk+IhDJJ1zW7KfKISitkz8iS7+FzvL6WeNY4ibHYp/kEvG6F7lTAYnRQeMRm52XgCYXyD+LFaZRHEnSUgnu0T+tphK29S99Ld7yuheW/N1ljsC6f3XRdggy5sQwk9je2zQQza4IiYCaYoiNXugURy/2FiN5XjdjKKbewF23hBiiMrHonpsnrip8TLAS//Hlqnw9W9bgVv0u9sy+vHF6rcvbhfoW1zRCvklMC28lu4rotbW5LG+//vrTZt5ZvC5rbgua2oHG9gn2QgqaoYaC8KVo9pvET7+37LChjZ2rNyJE0rR8JrzXzMQyanpRpTG77gGkEl/p5YAILtxTYyCDB1W+ois4inEJ/KDBdzKXMVS8lSrmEtpEZNv1Uck23aT6t4mM+z9qdpr/kZyaUWBXjfgMaT9k4tKpUhm628kHNb0PdsF2aVuuGgJa9CYnSZDaJmoNEazP4GhK6c/Z+WHQcLNpa/cZVO6YAaluvwHs3grf1rteoZ4ygIwc1+lz7KXP1xrvaOe/V0/uMycoRAK3FXU93NNe9j6efLgu1N/C0RcI4JQsrm4TxlSnwZARvw3l0lvvuPxVwN/V1p3CpRU+bYrMaChqt9ofwtU4i13IDS8qZgiXoEtZ4CIvOQzOcdr0F9I3hMk4heKR+98JsCYcvMyWyFf82qSUVUg2xjDKLm6yT+SemigjEaNz19PNvw4ElJolk5DqwdH+p5EK94H5p5MDrtpfJYkFmquz30oi2dHYLKT5LFs5fjfjbg7UkX4G7z6L5JZqylXiMIcQarUB7d04lHB8EmavnFM7DtpmsiL9rO1Oe/a1DriIfY5ZGON9Sytk8g5sNZUvH3G1tULrLnxkMumvC6VLvsO+87b5+r9aWK/bHTrFpWmlFb5vubPrhdvkSq2IXtVhluft6zu1skh0EqnObePe9v0StmMyiphnv5mGdtPNRm9p7rAhKu09zj922m4TTEm+79YPc9ajVO8SmsDSBbw7Oy2fbfPoMkscQThFXLDvtZgncmdIyPRXGt1M+X+eXTGaJJvO5LkqzVP6YLBCdX3W90FU55ofHeTXAEkCbmhdW2FbQWe3Zgnqzy0WzBbsVzsrYa/WqLbyV2ByzboVNa9FFW11tTtR1rW5m1g7LntqkYWMpuNq1IrTJBYbSOTvMzXIv5JkrlVfacIVWgna93/qNXn0QNgYVv90YVeq1ul9pN3q1Sq/RqAWjRuAP++HnQE9FcdDIvnwYw2kQW+ffP5jxnW8g4s2B150Zj6vcfONQNd4330AE4f5vIMCRQCscBfWwFw4qg2HQrNTDYbPSbtV6lUHYHIY92LSb497nHrow4KA/HI7HjbDSHACu7vcalV6/Nqg026N+OA5G9aEP4Hz7uYK3GJ1zc1vApeF170cAAAD//wMAUEsDBBQABgAIAAAAIQBowXh5+wcAADNSAAANAAAAeGwvc3R5bGVzLnhtbOxcWY/bNhB+L9D/IAh99eqw5WNrO83uxkCANAiQLVCg6YNWlr1CdBgSvbET9L93qMMiLVOiTmvR7kNiyTL5zT2coTh/c3Bs4cX0A8tzF6JyI4uC6Rre2nK3C/GPx9VgKgoB0t21bnuuuRCPZiC+Wf780zxAR9v8/GyaSIAh3GAhPiO0u5WkwHg2HT248XamC99sPN/REVz6WynY+aa+DvCPHFtSZXksObrlitEIt47BM4ij+1/3u4HhOTsdWU+WbaFjOJYoOMbt+63r+fqTDVAPykg3hIMy9lXh4CeThHcz8ziW4XuBt0E3MK7kbTaWYWbhzqSZpBvpSDBytZEUTZJVivaDX3GkkeSbLxYWn7icu3tn5aBAMLy9i0Ccp1tC9M37Ndwcj0Qhksq9twY+/fXLQJHh7+/1F8FxnC/CEf5+/U2UlnMpHnA533huOu4YWICZe/vV9b65K/xVNBl+ajkPvgsvug13FDyG4dmeLyBQCpgrvOPqjhk98XaHvED4qPu+9w0/u9Edyz5G36n4RqhM8cOOBaINUUXTRP8+4aeuNmEIkqJQxggbpTChbtYBN0+cDMnoUnTtTiiF+gsKZtn2yTwmYB74xnIOngSZvruCCyH+/Hjcgb664PQilQufK3h66+tHRdX4fxB4trXGKLb3GSt5iu9Z7to8mGC4YLfYJAmscBUhKsDFmAZ8PbKwnxjIN+poNptMFPw3mc6GeKZuASgAYDYdTUbyZKSpYzX0FF1A6JzSM0LLsjoUeQDi8fw1xO3E1w9l0KPo3nJumxsEdPnW9hn/j7wdptJDCILbcr629K3n6jbWpuQX5C8h4ENsX4joGWJz4sDP1RBPQc3A9StAUgIITUKLE3ANHXIzZGYLSMryvGXEXMNHClWsT4kYS2Lun8r2gStlda9RzD3T0tZ8WZuqV9UKrirIsyDSK/fQPw2PFztkfnUh6EtpQDnp20VaOIfjdcd15iCifp1hXjflkTE04ogbFm1upC2tlX2ksFfKWwcMkQm/dqM/Xxr0ybpzli3NOLD/dSCsXLB14L/HoJz4wOmDO3YOHYSNfhLeyHqmmbhankFVVz+cMzWe6jXkCLoju307PveajVRUrqWP7a/DGtLc1mMVh4rSpc389V9dLWlwvVIGSlxbhlK1Ydr2Z1xT/nND9SYPG6IvCatl3KTBLUr8Efot8ceoNB1dACkSOVo0NjHsdAgV6/LjCofNaQLWr9UUFXxMUSnQDY1/Lei7nX3EDVHc6oyvgJL06i6s2qfXb21r6zpm0q3Vk0vh2fOt7zAQbqRiVxz2fg4bNnEseHC/z/CG/YBHqBwgSoULvfIz7nUkTgIQYCBtgAEITIPS3LjPH1kU0JRYFDWadk7ex73zZPqrcNMGocNNa7QBGm9Gzfw8na4tlVymEKMDH1IW4+0NlEEzmMKvCZEn+OR7yDRQtNMG3FSOseMtOcgysPHDWgDb/u70Y8H2jK+4IRy2zKUM/yjBs2icvG4aLweL3lkuwMxxJbmxoLTKlNIRDv5lrADvjugikLGMPqOyVwBECRS2xtF+oh8Cxdu74uwp361VzlPK6Gbq6fk9GKGdLGKyylA1RHVIDCumt5FklI+wVGLQBqSwLlyYxxKpSoE/7y7NZoXUfJbVThtYKUKebLnSPlBErgwnh8NcC5eaSpjJTlke/zzNuxzeKGnNeBP7CmNn2Evj5mIdn7HAIvNiXp813ytl8CDCBCAkFMSiOcujlhASqTSnOl4Rc7jSbxBx+7WJmKcNYm6/YNE85itqs/DN13eP5iEsOeH6WN6CmqXa7fiLdBGbcLwMViL1o9LuAl4zvWsGTRGziHQtH0DNKblCtpIplPWrVMOqMCkVVSsn7BJTUXJRsvkCFVPOl2fn9ZhKCxQmmOz6+YpgCpZLnXKmJBbKuDpcrDEFm1n20+voVnhZ2hdWqgYrmaT4lfiY5qoBuVXZooDBVJlXrvJqQcjt1H3Am0G5Da1uwRQkq82BwXE06Y+y9CwrJx77bRBj9GppfstJfaWJjNr+mpqrz8sUfp+yH7VP2U+R5Lr1GAVq1FgPoWbhnalmnaWSXEXXTDAoV3StUNzLenxGXyxnbFbZrmLpowIV2W0XnFRwSYU3j+zRnNHKgrmSL1O4IOoGVAMum7jU76iWr7BTkLiVuYA95G6lmpC4NbM7SNy9gu4gNdFiqCko7u5Md1yBE0ea365V18Sy9YKrm73SD7uv1pAroU58TTSWv66zA+Ya23doz17HaV0dfZ29PlcHX2fvRA92o6hthDuupS1rf0y3O7+6yHDb9GGUV+3DLj4KEH+IrrI24461FQbnD5q4TozfUoD3EoiXH6hXH04vMQj4rLCF+BHvuLaJ7OVpb9lwOtOF1x5gzPWBfpECri8cZZUe4XThxRl83tJ4MtM0bazJmhydt3TxcCkpHB3/CzMjfLZe+C7HiRyw2rW50fc2ejx9uRDTz7+ba2vvQMYRP/XJevFQOMRCTD9/wGcVKWO8PQx62R8COFwI/hf2vrUQf7y7m8we3q3UwVS+mw5GQ1MbzLS7h4E2ur97eFjNZFW+/4c44a/G+X7hgYSwlV4Z3QY2nALox8TG4D+n9xYicRHBDze3AWwS+0wdy281RR6shrIyGI316WA6HmqDlaaoD+PR3TttpRHYtYrnAMqSokQnCmLw2i2yHNO23ERWiYTIuyAkuMwhQkokIaWnPS7/BQAA//8DAFBLAwQUAAYACAAAACEAjqCrVJIBAACqAwAAFAAAAHhsL3NoYXJlZFN0cmluZ3MueG1sfJNRT+QgFIXfN9n/cNN3p6MPZmPaGmxxprEtDUWNjziDMyQFukDN+u+3xs0mwugj59zvcHIJ2fUfNcKrsE4anSfnq3UCQu/MXupDntyz27NfCTjP9Z6PRos8eRMuuS5+/sic87Cw2uXJ0fvpKk3d7igUdyszCb04L8Yq7pejPaRusoLv3VEIr8b0Yr2+TBWXOoGdmbXPk4vlllnL37Mo/wtF5mSR+YLVLR62GDM4g4FRxPCmLmEg97Ssu02W+iJL3yc/pntEWYdpKJeo3+C27moYFgj1hGLoGYaGVV8kAG77FXSoxeFASzq2DUU0WTlGt9bV1bfxlDRR/BNGUf3KvK8riiLQzepZ2BPG4Ln1UHEv4gYEsN6f9gby+C8zwm7qpkE3DYbtsvoBIn9rZuvCJtXyXLH2FG2vZPVDzSL9kdA7uKWkDYHPZUKX4hbRu1BlhKEGvidbo/0R+M7LV+mlcOBmpbh9C7M6ruLFlnw6CCW1BCrcZLSTz6MIyUEeNPezjXG8QafBdPlrxV8AAAD//wMAUEsDBBQABgAIAAAAIQDfLqcEcwEAANIFAAAQAAAAeGwvY2FsY0NoYWluLnhtbHSU627CIBiG/y/ZPRD+T+Qwd4jVBJJ5A9sFkIq2SUtNaZbt7scW+Zgc/phIH5/yva+w3X+NA/o0s+sn22C6WmNkbDsde3tu8Mf728MzRm7R9qiHyZoGfxuH97v7u22rh1Z1urfIG6xrcLcsl1dCXNuZUbvVdDHWPzlN86gX/3U+E3eZjT66zphlHAhbrzdk9AK827ZobvBBPGLU+01gNPx+kuu6pOK6HleA9C+O5IGyhFS05txkZP6Wp+J+FM1/WyYl9en9TRR3/lJxZiTzVRTSUDQYwMl8YiUSDJEM+dwmrMAQSMXKpGQ8nYiF3FJnSqoKKVnoKO4zJJw6c7LSEUs7kiwknDqDIc5eJiXLkueVjgpkpSMwwOy80hEPhkiGhJOJwAAT8TIpefaf5+XzogpkpSMwxLeXScnT5CWvdFQgK+cIDJCSqHTEszZFpSMwwEQVUors/oGV5E4T2TkSlXMU12/uOvXfQOA23v0AAAD//wMAUEsDBBQABgAIAAAAIQAwIYPQ9gAAAEUBAAAZAAAAZG9jTWV0YWRhdGEvTGFiZWxJbmZvLnhtbFTQ3WqEMBAF4FcJuY+JrjGuqAu9K2yhrxCTcQ3kZzFTaSl998a79u4wMB9zZrx9Bk8O2LNLcaJ1JSiBaJJ18THRD1xZT0lGHa32KcJEvyDT2zwav/jB6wX83WUkBYl5OIcT3RCfA+fZbBB0roIze8ppxcqkwNO6OgO8EY3gwT3vp/AGqK1GTf+yxNmJfoPqhGmvLeukFKxtl5b1jeyZUqZW0irVL+rnvFgvHspCTUkA3FKJ77s7nIcH2FLAIbyenup0o0Fqdl1F8dZuYVcJlkl7uRTNtLari2dSRIj44jBPtHxkh5CO0y+ZzyP/337+BQAA//8DAFBLAwQUAAYACAAAACEA0B2j6k4BAAB/AgAAEQAIAWRvY1Byb3BzL2NvcmUueG1sIKIEASigAAEAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAjJJfS8MwFMXfBb9DybNt2k2nlLYDlT05EJz45+2S3G1xTRqSzG7f3rTdamUKPibn3F/OuSSb7mQZfKKxolI5SaKYBKhYxYVa5eR5MQtvSGAdKA5lpTAne7RkWpyfZUynrDL4aCqNxgm0gScpmzKdk7VzOqXUsjVKsJF3KC8uKyPB+aNZUQ1sAyukozieUIkOODigDTDUPZEckJz1SL01ZQvgjGKJEpWzNIkS+u11aKT9daBVBk4p3F77Toe4QzZnndi7d1b0xrquo3rcxvD5E/o6f3hqq4ZCNbtiSIqMs5QZBFeZYg5lKTZgPrYKLoI3kBpK2GwzOvA0+yzBurlf/VIgv93/PXZq9a+15bonkQc+btqVOyov47v7xYwUo3g0CeObMIkXyXUaX6WX4/cmyY/5Jn53IQ95/k+ceOiAeAQUGT35MsUXAAAA//8DAFBLAwQUAAYACAAAACEAb3gbEpEBAAAaAwAAEAAIAWRvY1Byb3BzL2FwcC54bWwgogQBKKAAAQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACckkFv2zAMhe8D9h8M3Rs53VAMgayiSDf0sGEBkrZnVqZjobIkiKyR7NdPtpHU2XbajXx8ePpESd0eOlf0mMgGX4nlohQFehNq6/eVeNx9u/oiCmLwNbjgsRJHJHGrP35QmxQiJrZIRY7wVImWOa6kJNNiB7TIY58nTUgdcG7TXoamsQbvg3nr0LO8LssbiQdGX2N9Fc+BYkpc9fy/oXUwAx897Y4xA2t1F6OzBjjfUv+wJgUKDRdfDwadkvOhynRbNG/J8lGXSs5btTXgcJ2DdQOOUMl3QT0gDEvbgE2kVc+rHg2HVJD9ldd2LYoXIBxwKtFDsuA5Yw22qRlrF4mTfg7plVpEJiWzYRLHcu6d1/azXo6GXFwah4AJJA8uEXeWHdLPZgOJ/0G8nBOPDBPvhLMd+KYz53zjlfNJf2SvQxfBH/Ua4h47662SJ0l9t/6VHuMu3APjaa+Xotq2kLDOT3He+1lQD3mlyQ0h6xb8HuuT5+/B8Auepq+ulzeL8lOZH3imKfn+qfVvAAAA//8DAFBLAQItABQABgAIAAAAIQCVZpbtkAEAAIIFAAATAAAAAAAAAAAAAAAAAAAAAABbQ29udGVudF9UeXBlc10ueG1sUEsBAi0AFAAGAAgAAAAhADEdic0iAQAA3gIAAAsAAAAAAAAAAAAAAAAAyQMAAF9yZWxzLy5yZWxzUEsBAi0AFAAGAAgAAAAhAD4Ms3q3AwAAeQkAAA8AAAAAAAAAAAAAAAAAHAcAAHhsL3dvcmtib29rLnhtbFBLAQItABQABgAIAAAAIQCSB5TsBAEAAD8DAAAaAAAAAAAAAAAAAAAAAAALAAB4bC9fcmVscy93b3JrYm9vay54bWwucmVsc1BLAQItABQABgAIAAAAIQDDG6yZDw8AAG1IAAAYAAAAAAAAAAAAAAAAAEQNAAB4bC93b3Jrc2hlZXRzL3NoZWV0MS54bWxQSwECLQAUAAYACAAAACEA9mC0QbgHAAARIgAAEwAAAAAAAAAAAAAAAACJHAAAeGwvdGhlbWUvdGhlbWUxLnhtbFBLAQItABQABgAIAAAAIQBowXh5+wcAADNSAAANAAAAAAAAAAAAAAAAAHIkAAB4bC9zdHlsZXMueG1sUEsBAi0AFAAGAAgAAAAhAI6gq1SSAQAAqgMAABQAAAAAAAAAAAAAAAAAmCwAAHhsL3NoYXJlZFN0cmluZ3MueG1sUEsBAi0AFAAGAAgAAAAhAN8upwRzAQAA0gUAABAAAAAAAAAAAAAAAAAAXC4AAHhsL2NhbGNDaGFpbi54bWxQSwECLQAUAAYACAAAACEAMCGD0PYAAABFAQAAGQAAAAAAAAAAAAAAAAD9LwAAZG9jTWV0YWRhdGEvTGFiZWxJbmZvLnhtbFBLAQItABQABgAIAAAAIQDQHaPqTgEAAH8CAAARAAAAAAAAAAAAAAAAACoxAABkb2NQcm9wcy9jb3JlLnhtbFBLAQItABQABgAIAAAAIQBveBsSkQEAABoDAAAQAAAAAAAAAAAAAAAAAK8zAABkb2NQcm9wcy9hcHAueG1sUEsFBgAAAAAMAAwABQMAAHY2AAAAAA=="""

# def get_default_template_bytes():
#     """Decode the hardcoded default template back into raw .xlsx bytes."""
#     return base64.b64decode(DEFAULT_TEMPLATE_B64.strip())


# # =====================================================================
# # 2. HELPER FUNCTIONS
# # =====================================================================

# def norm_name(name):
#     """Normalize a name to a comparable token set regardless of 'First Last'
#     vs 'Last, First' formatting differences between source files."""
#     parts = re.split(r"[,\s]+", str(name).strip().lower())
#     parts = [p for p in parts if p]
#     return " ".join(sorted(parts))


# def name_tokens(name):
#     parts = re.split(r"[,\s]+", str(name).strip().lower())
#     return {p for p in parts if p}


# def find_matching_rep_key(sow_name, rep_name_keys):
#     """Match a SOW resource name to a Replicon name key even when one side
#     has an extra middle name (e.g. SOW 'Kalyani Ghaytadkar' vs Replicon
#     'Ghaytadkar, Kalyani Popatrao'). A match is any rep key whose token set
#     fully contains the SOW name's tokens, or vice versa."""
#     sow_tok = name_tokens(sow_name)
#     best_match = None
#     for rep_key in rep_name_keys:
#         rep_tok = set(rep_key.split())
#         if sow_tok.issubset(rep_tok) or rep_tok.issubset(sow_tok):
#             # Prefer the closest-length match if multiple candidates exist
#             if best_match is None or abs(len(rep_tok) - len(sow_tok)) < abs(len(set(best_match.split())) - len(sow_tok)):
#                 best_match = rep_key
#     return best_match

# def get_field(row, candidates, default=""):
#     """
#     Flexible lookup of a value from a SOW row, regardless of small
#     header-naming differences (spacing, trailing colon, casing) between
#     what the code expects and what the uploaded SOW file actually has.

#     `candidates` is a list of possible header names to try, in priority
#     order. Returns `default` if none of them are found or all are blank.
#     """
#     def normalize(s):
#         return re.sub(r"\s+", " ", str(s).strip().lower().rstrip(":"))

#     normalized_map = {normalize(col): col for col in row.index}

#     for cand in candidates:
#         key = normalize(cand)
#         if key in normalized_map:
#             val = row[normalized_map[key]]
#             if pd.notna(val) and str(val).strip() != "":
#                 return val
#     return default


# # def load_sow(file):
# #     """Load the SOW & Resource Details file into a DataFrame."""
# #     df = pd.read_excel(file)
# #     df["name_key"] = df["Resource name"].apply(norm_name)
# #     return df
# def load_sow(file):
#     df = pd.read_excel(file)
#     name_col = None
#     for candidate in ["Resource name", "Employee Name", "Name"]:
#         for col in df.columns:
#             if re.sub(r"\s+", " ", str(col).strip().lower()) == candidate.lower():
#                 name_col = col
#                 break
#         if name_col:
#             break
#     if name_col is None:
#         raise ValueError("Could not find a resource/employee name column in the SOW file.")
#     df["name_key"] = df[name_col].apply(norm_name)
#     return df

# # def load_replicon(file):
# #     """Load the Replicon dump (Expenditure Details Report) into a DataFrame.
# #     Row 1 is an instructions banner, row 2 holds the real headers."""
# #     df = pd.read_excel(file, sheet_name="Expenditure Details Report", header=1)
# #     df = df.dropna(subset=["Employee Name/Supplier Name"])
# #     df["name_key"] = df["Employee Name/Supplier Name"].apply(norm_name)
# #     df["Item Date"] = pd.to_datetime(df["Item Date"], format="%d-%b-%Y")
# #     df["Quantity"] = pd.to_numeric(df["Quantity"], errors="coerce").fillna(0)
# #     return df

# def load_replicon(file):
#     """Load the Replicon dump into a DataFrame. Sheet is 'Sheet1' with
#     headers on row 1. Employee name, date, and hours columns are renamed
#     to the internal standard names used throughout the rest of the app."""
#     df = pd.read_excel(file, sheet_name="Sheet1")
#     df = df.dropna(subset=["Employee Name"])
#     df["name_key"] = df["Employee Name"].apply(norm_name)
#     df["Item Date"] = pd.to_datetime(df["Expenditure Item Date"])
#     df["Quantity"] = pd.to_numeric(df["QTT"], errors="coerce").fillna(0)
#     return df


# def build_daily_hours(rep_df):
#     """Sum hours per employee per date (nets out negative correction rows),
#     and track which (employee, date) pairs had a correction so we can flag
#     them in the Remark column."""
#     daily = rep_df.groupby(["name_key", "Item Date"])["Quantity"].sum().reset_index()
#     neg = rep_df[rep_df["Quantity"] < 0][["name_key", "Item Date"]]
#     neg_set = set(zip(neg["name_key"], neg["Item Date"]))
#     return daily, neg_set


# def determine_work_from(role):
#     """Determine WFH vs WFO from the Partner Emp. Role string.

#     Rule (per business input): the role code contains a segment like
#     'T&M_13' — the number right after 'T&M_'. Only the LAST digit of that
#     number matters (the leading digit, e.g. the '1' in '13', is ignored):
#         - last digit 3  -> WFO (work from office)
#         - last digit 2  -> WFH (work from home)
#     Falls back to a keyword check ("onsite" -> WFO) and then defaults to
#     WFH if the pattern can't be parsed, so nothing breaks on unexpected
#     role formats.
#     """
#     role_str = str(role)

#     # Explicit keyword override, if present
#     if "onsite" in role_str.lower():
#         return "WFO"

#     match = re.search(r"T&M[_\s]*(\d+)", role_str)
#     if match:
#         last_digit = match.group(1)[-1]
#         if last_digit == "3":
#             return "WFO"
#         elif last_digit == "2":
#             return "WFH"

#     # Unrecognized pattern - default to WFH, but this should be reviewed
#     return "WFH"

# def fill_timesheet_for_employee(template_bytes, res_row, daily_hours, neg_set, month, year, matched_rep_key):
#     """Fill one copy of the timesheet template for a single employee and
#     return it as an in-memory .xlsx (BytesIO)."""
#     wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
#     ws = wb["Sheet1"]

#     name_key = matched_rep_key  # the Replicon-side key matched to this SOW resource

#     # ---- Pull all resource/PO/SOW fields from the SOW & Resource Details
#     # ---- file, tolerant of header-name variations in that file.
#     emp_name = get_field(res_row, ["Resource name", "Employee Name", "Name"])
#     cid = get_field(res_row, ["CID", "C.I.D", "C.I.D.", "Cid"])
#     role = get_field(res_row, ["Partner Emp. role", "Partner Emp Role", "PARTNER EMP. ROLE", "Emp Role", "Role"])
#     domain = get_field(res_row, ["Domain"])
#     po_number = get_field(res_row, ["PO Number", "PO Number:", "PO No", "PO No."])
#     po_start = get_field(res_row, ["PO Start Date", "PO Start Date:", "PO Start"])
#     po_end = get_field(res_row, ["PO End Date", "PO End Date:", "PO End"])
#     sow_number = get_field(res_row, ["SOW number", "SOW Number", "SOW Number:", "SOW No"])

#     work_from_value = determine_work_from(role)

#     # ---- Header fields ----
#     ws["D4"] = emp_name
#     ws["D5"] = cid
#     ws["D6"] = role
#     ws["D7"] = domain
#     if po_number != "":
#         ws["D8"] = po_number
#     if po_start != "":
#         ws["D9"] = po_start
#     if po_end != "":
#         ws["D10"] = po_end
#     ws["D11"] = sow_number
#     ws["H4"] = calendar.month_name[month]
#     ws["H6"] = year

#     # ---- Daily rows (row 14 = day 1) ----
#     emp_hours = daily_hours[daily_hours["name_key"] == name_key].set_index("Item Date")["Quantity"]
#     days_in_month = calendar.monthrange(year, month)[1]
#     date_number_format = ws["B14"].number_format  # capture template's date format before overwriting

#     proj_start_raw = get_field(res_row, ["Start date", "Start Date", "Project Start Date"], None)
#     proj_end_raw = get_field(res_row, ["End date", "End Date", "Project End Date"], None)

#     month_start = datetime(year, month, 1)
#     month_end = datetime(year, month, days_in_month)

#     proj_start = pd.to_datetime(proj_start_raw) if pd.notna(proj_start_raw) else month_start
#     proj_end = pd.to_datetime(proj_end_raw) if pd.notna(proj_end_raw) else month_end

#     effective_start = max(proj_start, month_start)
#     effective_end = min(proj_end, month_end)


    
#     for day in range(1, days_in_month + 1):
#         row = 13 + day
#         this_date = datetime(year, month, day)

#         if this_date < effective_start or this_date > effective_end:
#             for col in range(2, 9):
#                 ws.cell(row=row, column=col, value=None)
#             continue


        
#         hours = emp_hours.get(this_date, 0)
#         remark = "Correction adjusted" if (name_key, this_date) in neg_set else ""
#         is_weekend = this_date.weekday() >= 5  # 5 = Saturday, 6 = Sunday

#         date_cell = ws.cell(row=row, column=2, value=this_date)
#         date_cell.number_format = date_number_format
#         ws.cell(row=row, column=3, value=this_date.strftime("%A").upper())

#         if is_weekend and hours == 0:
#             ws.cell(row=row, column=7, value=None)   # blank instead of 0
#         else:
#             ws.cell(row=row, column=7, value=float(hours))

#         ws.cell(row=row, column=8, value=remark)

#     # ---- Dropdown for Work From column (F) — WFO / WFH ----
#     from openpyxl.worksheet.datavalidation import DataValidation
#     # dv = DataValidation(type="list", formula1='"WFO,WFH"', allow_blank=True)
#     dv = DataValidation(type="list", formula1='"WFO,WFH,On Leave"', allow_blank=True)
#     dv.error = "Please select WFO or WFH"
#     dv.errorTitle = "Invalid entry"
#     dv.prompt = "Select WFO or WFH"
#     dv.promptTitle = "Work From"
#     ws.add_data_validation(dv)
#     dv.add(f"F14:F{13 + days_in_month}")
#     # ---- Dropdown for Activity column (D) — with custom entry allowed ----
#     dv_activity = DataValidation(
#         type="list",
#         formula1='"On Leave,Sick Leave,Public Holiday,Project work"',
#         allow_blank=True,
#         showErrorMessage=False,  # allows typing a custom value beyond the 3 options
#     )
#     dv_activity.prompt = "Select an activity or type your own"
#     dv_activity.promptTitle = "Activity"
#     ws.add_data_validation(dv_activity)
#     dv_activity.add(f"D14:D{13 + days_in_month}")
#     # ---- Dropdown for Month (H4) ----
#     dv_month = DataValidation(
#         type="list",
#         formula1='"January,February,March,April,May,June,July,August,September,October,November,December"',
#         allow_blank=True,
#     )
#     dv_month.error = "Please select a valid month"
#     dv_month.errorTitle = "Invalid entry"
#     dv_month.prompt = "Select a month"
#     dv_month.promptTitle = "Month"
#     ws.add_data_validation(dv_month)
#     dv_month.add("H4")

#     # ---- Dropdown for Year (H6) ----
#     dv_year = DataValidation(
#         type="list",
#         formula1='"2024,2025,2026,2027,2028,2029,2030,2031"',
#         allow_blank=True,
#     )
#     dv_year.error = "Please select a valid year"
#     dv_year.errorTitle = "Invalid entry"
#     dv_year.prompt = "Select a year"
#     dv_year.promptTitle = "Year"
#     ws.add_data_validation(dv_year)
#     dv_year.add("H6")
#     # Blank out any leftover template rows beyond this month's day count
#     for row in range(14 + days_in_month, 45):
#         for col in range(2, 9):
#             ws.cell(row=row, column=col, value=None)
    
#     # ---- Sign-off block ----
#     ws["C52"] = emp_name
#     ws["F52"] = get_field(res_row, ["Capgemini Resposible", "Capgemini Responsible"])
#     ws["F56"] = get_field(res_row, ["EGA Resposible", "EGA Responsible"])
#     # ---- Lock only these specific cells; everything else stays editable ----
#     # locked_cells = ["D6", "D7", "D8", "D9", "D10", "D11"]

#     # # Unlock all cells first
#     # for row in ws.iter_rows():
#     #     for cell in row:
#     #         cell.protection = Protection(locked=False)

#     # # Lock only the required ones
#     # for coord in locked_cells:
#     #     ws[coord].protection = Protection(locked=True)

#     # # Enable sheet protection so locking actually takes effect
#     # ws.protection.sheet = True
#     # ws.protection.password = "yourpassword"  # optional, remove this line if no password needed
#     # ---- Freeze specific fields so employees can't edit them ----
#     # from openpyxl.styles import Protection

#     # locked_cells = ["D6", "D7", "D8", "D9", "D10", "D11"]

#     # for row in ws.iter_rows():
#     #     for cell in row:
#     #         cell.protection = Protection(locked=False)

#     # for coord in locked_cells:
#     #     ws[coord].protection = Protection(locked=True)

#     # ws.protection.sheet = True
#     # ws.protection.enable()

#     # out = io.BytesIO()
#     # # wb.save(out)

#     # # out = io.BytesIO()
#     # ---- Lock everything EXCEPT Activity (D) and Work From (F) daily cells ----
#     from openpyxl.styles import Protection

#     editable_cols = {4, 6, 7}  # column D = Activity, column F = Work From
#     editable_rows = set(range(14, 14 + days_in_month))  # daily table rows only

#     for row in ws.iter_rows():
#         for cell in row:
#             is_editable = (cell.row in editable_rows) and (cell.column in editable_cols)
#             cell.protection = Protection(locked=not is_editable)
#     ws["D5"].protection = Protection(locked=False)   # ← new line: unlocks CID

#     ws.protection.sheet = True
#     ws.protection.enable()

#     out = io.BytesIO()
#     # wb.save(out)
#     wb.save(out)
#     out.seek(0)
#     return out.getvalue()


# # =====================================================================
# # 3. STREAMLIT UI
# # =====================================================================
# # =====================================================================
# # 3. STREAMLIT UI
# # =====================================================================

# st.title("📋 Timesheet Auto-Filler")
# st.write(
#     "Upload **SOW & Resource Details** and **Replicon Dump** files below. "
#     "The app will automatically generate a filled timesheet for **every employee** "
#     "found in the SOW file, and save each one with that employee's name."
# )

# col1, col2 = st.columns(2)
# with col1:
#     sow_file = st.file_uploader("1️⃣ Upload SOW & Resource Details (.xlsx)", type=["xlsx"])
# with col2:
#     replicon_file = st.file_uploader("2️⃣ Upload Replicon Dump (.xlsx)", type=["xlsx"])

# st.subheader("3️⃣ Timesheet Template")
# template_option = st.radio(
#     "Choose which timesheet template to fill:",
#     ["Use built-in default template (hardcoded in app)", "Upload a custom template"],
#     index=0,
# )

# custom_template_file = None
# if template_option == "Upload a custom template":
#     custom_template_file = st.file_uploader(
#         "Upload Timesheet Template (.xlsx)", type=["xlsx"], key="template_upload"
#     )
# else:
#     st.caption("✅ Using the built-in default timesheet template embedded in this app.")

# st.divider()
# generate_btn = st.button("🚀 Generate Timesheets", type="primary", use_container_width=True)

# if generate_btn:
#     if not sow_file or not replicon_file:
#         st.error("Please upload both the SOW & Resource Details and Replicon Dump files.")
#     elif template_option == "Upload a custom template" and not custom_template_file:
#         st.error("Please upload a custom timesheet template, or switch to the built-in default.")
#     else:
#         try:
#             with st.spinner("Reading files..."):
#                 sow_df = load_sow(sow_file)
#                 rep_df = load_replicon(replicon_file)
#                 daily_hours, neg_set = build_daily_hours(rep_df)

#                 # Auto-detect month/year from the Replicon dump's dates
#                 month = int(rep_df["Item Date"].dt.month.mode()[0])
#                 year = int(rep_df["Item Date"].dt.year.mode()[0])

#                 template_bytes = (
#                     custom_template_file.read() if custom_template_file else get_default_template_bytes()
#                 )

#             st.success(
#                 f"Detected period: **{calendar.month_name[month]} {year}**. "
#                 f"Found **{len(sow_df)}** resource(s) in the SOW file."
#             )

#             generated_files = {}
#             unmatched_employees = []
#             rep_name_keys = daily_hours["name_key"].unique().tolist()
#             progress = st.progress(0.0)
#             for i, (_, res_row) in enumerate(sow_df.iterrows()):
#                 # matched_key = find_matching_rep_key(res_row["Resource name"], rep_name_keys)
#                 matched_key = find_matching_rep_key(get_field(res_row, ["Resource name", "Employee Name", "Name"]), rep_name_keys)
#                 # if matched_key is None:
#                 #     unmatched_employees.append(res_row["Resource name"])
#                 if matched_key is None:
#                     unmatched_employees.append(get_field(res_row, ["Resource name", "Employee Name", "Name"]))
#                 data = fill_timesheet_for_employee(
#                     template_bytes, res_row, daily_hours, neg_set, month, year, matched_key
#                 )
#                 # emp_name_clean = str(res_row.get("Resource name", f"Employee_{i+1}")).strip().replace(" ", "_")
#                 emp_name_clean = str(get_field(res_row, ["Resource name", "Employee Name", "Name"], f"Employee_{i+1}")).strip().replace(" ", "_")
#                 fname = f"Timesheet_{calendar.month_name[month]}_{year}_{emp_name_clean}.xlsx"
#                 generated_files[fname] = data
#                 progress.progress((i + 1) / len(sow_df))

#             if unmatched_employees:
#                 st.warning(
#                     "⚠️ No matching hours found in the Replicon dump for: "
#                     + ", ".join(unmatched_employees)
#                     + ". Their timesheet was generated with all days marked as Leave — "
#                     "double check the name spelling in both files."
#                 )


#             st.subheader("✅ Generated Timesheets")
#             for fname, data in generated_files.items():
#                 st.download_button(
#                     label=f"⬇️ Download {fname}",
#                     data=data,
#                     file_name=fname,
#                     mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#                     key=fname,
#                 )

#             # Zip download of all files together
#             zip_buffer = io.BytesIO()
#             with zipfile.ZipFile(zip_buffer, "w") as zf:
#                 for fname, data in generated_files.items():
#                     zf.writestr(fname, data)
#             zip_buffer.seek(0)

#             st.download_button(
#                 "📦 Download All Timesheets as ZIP",
#                 data=zip_buffer.getvalue(),
#                 file_name=f"All_Timesheets_{calendar.month_name[month]}_{year}.zip",
#                 mime="application/zip",
#                 use_container_width=True,
#             )

#         except Exception as e:
#             st.error(f"Something went wrong: {e}")
#             st.exception(e)

# st.divider()
# with st.expander("ℹ️ How field mapping works"):
#     st.markdown(
#         """
#         - **Partner Emp. Name / CID / Role / Domain / PO Number / SOW Number** ← from SOW & Resource Details file
#         - **Daily Billable Hours** ← summed per employee per date from the Replicon dump
#           (negative correction rows are netted automatically)
#         - **Activity** ← SOW name, filled only on days with hours > 0
#         - **Work From** ← "WFO" if role contains "onsite", otherwise "WFH"
#         - **Days with 0 hours** ← marked as "Leave"
#         - **Capgemini/EGA Responsible, sign-off Name** ← from SOW & Resource Details file 
#         """
#     )





"""
Timesheet Auto-Filler - Streamlit App
=======================================
Upload SOW & Resource Details + Replicon Dump files, and this app will
automatically generate one filled Timesheet per employee.

Features:
- Upload SOW & Resource Details (.xlsx)
- Upload Replicon Dump (.xlsx)
- Timesheet template: use the BUILT-IN default template (hardcoded/embedded
  in this file as base64) OR upload your own custom template (.xlsx)
- Generates one filled timesheet per employee found in the SOW file
- Download each file individually, or all together as a ZIP

Run with:
    streamlit run timesheet_app.py
"""

import streamlit as st
import pandas as pd
import openpyxl
import re
import io
import base64
import calendar
import zipfile
from datetime import datetime
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.styles import Protection
from openpyxl.styles import Protection, Alignment

st.set_page_config(page_title="Timesheet Auto-Filler", layout="wide", page_icon="📋")
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

    html, body, [class*="css"] {
        font-family: 'Inter', -apple-system, sans-serif !important;
    }

    /* Overall background */
    .stApp {
        background: linear-gradient(-45deg, #f4f7fb, #eaf0f8, #e6eefc, #eef4fc);
        background-size: 400% 400%;
        animation: gradientShift 18s ease infinite;
    }
    @keyframes gradientShift {
        0% { background-position: 0% 50%; }
        50% { background-position: 100% 50%; }
        100% { background-position: 0% 50%; }
    }

    /* Top header bar */
    header[data-testid="stHeader"] {
        background: transparent;
    }

    div.block-container {
        padding-top: 0.5rem;
        padding-bottom: 1rem;
        padding-left: 2rem;
        padding-right: 2rem;
        max-width: 100%;
    }

    div[data-testid="stVerticalBlock"] > div {
        gap: 0.5rem;
    }

    h1 {
        color: #0b3d91;
        font-weight: 700;
        letter-spacing: -0.5px;
    }
    h2, h3 {
        color: #14395e;
        display: flex;
        align-items: center;
        gap: 0.5rem;
        margin-top: 0.2rem;
        margin-bottom: 0.2rem;
        font-size: 1.1rem;
    }

    /* Hero banner */
    .hero-banner {
        display: flex;
        align-items: center;
        gap: 1rem;
        background: linear-gradient(135deg, #0b3d91 0%, #1a56c4 55%, #3b7ce0 100%);
        border-radius: 14px;
        padding: 1rem 1.4rem;
        margin-bottom: 0.8rem;
        box-shadow: 0 6px 18px rgba(11, 61, 145, 0.3);
        position: relative;
        overflow: hidden;
    }
    .hero-banner::after {
        content: "";
        position: absolute;
        top: -40px;
        right: -40px;
        width: 180px;
        height: 180px;
        background: rgba(255,255,255,0.08);
        border-radius: 50%;
    }
    .hero-banner::before {
        content: "";
        position: absolute;
        bottom: -60px;
        right: 80px;
        width: 140px;
        height: 140px;
        background: rgba(255,255,255,0.06);
        border-radius: 50%;
    }
    .hero-icon {
        font-size: 1.8rem;
        background: rgba(255,255,255,0.15);
        border-radius: 10px;
        padding: 0.5rem 0.65rem;
        box-shadow: inset 0 0 0 1px rgba(255,255,255,0.25);
        z-index: 1;
    }
    .hero-text { z-index: 1; }
    .hero-text h1 {
        color: #ffffff !important;
        margin: 0 0 0.2rem 0;
        font-size: 1.3rem;
    }
    .hero-text p {
        color: #dce8fb;
        font-size: 0.82rem;
        margin: 0;
        line-height: 1.4;
    }
    .hero-text p b { color: #ffffff; }

    /* Step tracker */
    .step-tracker {
        display: flex;
        align-items: center;
        justify-content: center;
        margin: 0.8rem 0 1rem 0;
    }
    .step {
        display: flex;
        flex-direction: column;
        align-items: center;
        gap: 0.3rem;
    }
    .step-num {
        width: 24px;
        height: 24px;
        border-radius: 50%;
        display: flex;
        align-items: center;
        justify-content: center;
        background: #d7e3f5;
        color: #6f89ad;
        font-weight: 700;
        font-size: 0.75rem;
        transition: all 0.3s ease;
    }
    .step.active .step-num {
        background: linear-gradient(135deg, #1a56c4, #0b3d91);
        color: #ffffff;
        box-shadow: 0 4px 12px rgba(11,61,145,0.35);
    }
    .step-label {
        font-size: 0.68rem;
        color: #5c7699;
        font-weight: 600;
    }
    .step.active .step-label { color: #0b3d91; }
    .step-line {
        width: 45px;
        height: 2px;
        background: #d7e3f5;
        margin: 0 0.6rem;
        margin-bottom: 1rem;
    }

    /* Card-style containers */
    [data-testid="stVerticalBlockBorderWrapper"],
    div.stFileUploader, div[data-testid="stExpander"] {
        background: #ffffff;
        border-radius: 10px;
        padding: 0.6rem;
        box-shadow: 0 3px 10px rgba(11, 61, 145, 0.07), 0 1px 3px rgba(11, 61, 145, 0.05);
        border: 1px solid #e3ecf7;
    }

    div.stFileUploader label {
        font-weight: 600;
        color: #14395e;
        font-size: 0.9rem;
    }
    div.stFileUploader section {
        border: 2px dashed #b8cceb !important;
        border-radius: 10px !important;
        background: #f8fbff !important;
    }
    div.stFileUploader {
        transition: box-shadow 0.2s ease, transform 0.2s ease;
    }
    div.stFileUploader:hover {
        box-shadow: 0 6px 16px rgba(11, 61, 145, 0.12);
        transform: translateY(-1px);
    }

    /* Stat chips */
    .stat-chip {
        background: #ffffff;
        border-radius: 10px;
        padding: 0.5rem 0.7rem;
        text-align: center;
        box-shadow: 0 2px 8px rgba(11,61,145,0.06);
        border: 1px solid #e3ecf7;
        color: #14395e;
        font-size: 0.78rem;
    }
    .stat-chip b { color: #0b3d91; font-size: 0.85rem; }

    /* File result cards */
    .file-card {
        display: flex;
        align-items: center;
        gap: 0.6rem;
        background: #f8fbff;
        border: 1px solid #e3ecf7;
        border-radius: 8px;
        padding: 0.45rem 0.7rem;
        margin-bottom: 0.4rem;
        transition: all 0.2s ease;
    }
    .file-card:hover {
        border-color: #1a56c4;
        box-shadow: 0 3px 10px rgba(11,61,145,0.1);
    }
    .file-icon { font-size: 1.05rem; }
    .file-name { font-weight: 600; color: #14395e; font-size: 0.82rem; }

    /* Primary button */
    button[kind="primary"] {
        background: linear-gradient(135deg, #1a56c4 0%, #0b3d91 100%) !important;
        border: none !important;
        border-radius: 8px !important;
        font-weight: 600 !important;
        padding: 0.55rem 1.1rem !important;
        color: #ffffff !important;
        box-shadow: 0 3px 10px rgba(11, 61, 145, 0.3);
        transition: transform 0.15s ease, box-shadow 0.15s ease;
    }
    button[kind="primary"]:hover {
        transform: translateY(-1px);
        box-shadow: 0 6px 16px rgba(11, 61, 145, 0.4);
    }

    /* Download buttons */
    div.stDownloadButton > button {
        background: #ffffff;
        color: #0b3d91;
        border: 1.5px solid #0b3d91;
        border-radius: 8px;
        font-weight: 600;
        font-size: 0.85rem;
        transition: all 0.15s ease;
        box-shadow: 0 2px 6px rgba(11, 61, 145, 0.06);
    }
    div.stDownloadButton > button:hover {
        background: #0b3d91;
        color: #ffffff;
        box-shadow: 0 4px 12px rgba(11, 61, 145, 0.25);
        transform: translateY(-1px);
    }

    div[role="radiogroup"] {
        background: #ffffff;
        padding: 0.5rem 0.7rem;
        border-radius: 10px;
        box-shadow: 0 2px 6px rgba(11, 61, 145, 0.05);
        border: 1px solid #e3ecf7;
    }

    div[data-testid="stAlert"] {
        border-radius: 10px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05);
        font-size: 0.85rem;
    }

    hr {
        border-top: 2px solid #dce6f5;
        margin: 0.6rem 0;
    }

    div[data-testid="stProgress"] > div > div {
        background: linear-gradient(90deg, #1a56c4, #0b3d91);
    }

    div[data-testid="stHorizontalBlock"] { gap: 0.6rem; }

    .app-footer {
        text-align: center;
        padding: 1rem 0 0.4rem 0;
        color: #8fa3c2;
        font-size: 0.75rem;
        border-top: 1px solid #dce6f5;
        margin-top: 1.2rem;
    }
</style>
""", unsafe_allow_html=True)

DEFAULT_TEMPLATE_B64 = """
UEsDBBQABgAIAAAAIQCVZpbtkAEAAIIFAAATAAgCW0NvbnRlbnRfVHlwZXNdLnhtbCCiBAIooAACAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACslM9O4zAQxu8r7TtEvqLEhcMKoaYc+CMh7QIS7AMM9qSx6tiWZ4D27Zm4gBDqtqzoJVZiz/f9PJmZ6ely8NUTZnIxtOqwmagKg4nWhXmr/t5f1seqIoZgwceArVohqdPZzx/T+1VCqiQ6UKt65nSiNZkeB6AmJgyy08U8AMtrnusEZgFz1EeTyS9tYmAMXPOooWbTc+zg0XN1sZTPa5KMnlR1tj44erUKUvLOAAupfgr2k0v96tBIZDlDvUt0IBhKb3QYd/5t8Bp3I6nJzmJ1C5mvYRAMvfT6OebFQ4yLZrvIBsrYdc6gjeZxkAw0lDKCpR6RB9+UtRnAhTfuLf7lMOmyHO4ZZLxfEd7BwfK/UZfn9xGKzA5D4pVH2nfai+gu5x4y2jvO0hl7B/iovYPDgDdnvZTInpPwrrvNX+r2DzJYYNC/4QH9VejiF0AGqteF3xgPRE56oPSoHzXem3RTsYvjbY6JZGZk/P8rvw2FMbpOIoSZHX7NUebNt3OM40SzaDd46zJBZy8AAAD//wMAUEsDBBQABgAIAAAAIQAxHYnNIgEAAN4CAAALAAgCX3JlbHMvLnJlbHMgogQCKKAAAgAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAArJLRSgMxEEXfBf8h5L0721VEpNu+iFBQEKkfME1mt6FJJiRR2783rRZdqEXQx2Tu3Jy5k8ls46x4pZgM+1aOq1oK8oq18X0rnxd3o2spUkav0bKnVm4pydn0/GzyRBZzaUorE5IoLj61cpVzuAFIakUOU8WBfKl0HB3mcow9BFRr7Amaur6C+N1DTgeeYq5bGef6QorFNpSX/+INjjJqzAiKI41CLGQxmzKLWGDsKbdSs3os12mvqAq1hONAzQ9AzqjIibtcKXbAXWfUbsymhroZTgrKYkqmCPYJWlySHZI8HHDvd7W57/gU0fj3EX1g3bJ6ceTzkS18gh8UX/lsLLxxXC+Z16dYLv+ThTaZvCZ9emEYwoEIBr9y+g4AAP//AwBQSwMEFAAGAAgAAAAhAD4Ms3q3AwAAeQkAAA8AAAB4bC93b3JrYm9vay54bWysVm1vozgQ/n7S/QfEdxcbDAHUdBXedJXaVZVm2zup0soFp0EBnDNOk6ra/75jAknbnE657kUJxvbw+JmZZ4acf9nWlfHMZVuKZmySM2wavMlFUTZPY/PbLEO+abSKNQWrRMPH5gtvzS8Xv/92vhFy+SjE0gCAph2bC6VWoWW1+YLXrD0TK97AzlzImimYyierXUnOinbBuaory8bYs2pWNuYOIZSnYIj5vMx5IvJ1zRu1A5G8Ygrot4ty1Q5odX4KXM3kcr1CuahXAPFYVqV66UBNo87Dy6dGSPZYgdtb4hpbCV8PfgTDxR5Ogq2jo+oyl6IVc3UG0NaO9JH/BFuEvAvB9jgGpyFRS/LnUudwz0p6n2Tl7bG8AxjBv4xGQFqdVkII3ifR3D0327w4n5cVv9tJ12Cr1VdW60xVplGxVqVFqXgxNkcwFRv+bkGuV9G6rGDXwcQhpnWxl/ONNAo+Z+tKzUDIA/zYtLHtYKwtQRiTSnHZMMVj0SjQYe/Xr2quw44XAhRuTPnf61JyKCzQF/gKV5aH7LG9YWphrGU1NuPw4VsL7j+81KyqyiV7SMSmqQTU2MMbcbLjSvgP8mS59tkCp3fEdvcfAwD8ZDhI8EZJA+4vkytIwy17hqRA6ou+Zi8h6v73V4c6MfGdBI2CNEEUTwLkByRCmZf4qR+7JKWTH+CF9MJcsLVa9InWmGOTQlaPtq7ZdtghOFyXxeH8V9x/kB4/XIa9H9pT3dLuSr5pD5LQU2N7XzaF2IxNRLSQX95PN93mfVmoBTTRgNpgslv7g5dPC2BMiAvy0X1DMxubry5xUpplPoojb4Jo4HrIx06E/HSSxK7tZI7ndoysN5S65gnUutFoOsHf6oZKoEvrUUcX7mWoz5CXRadta3gsZ1UOAtdDZxgQbAfaa75VV63qRtBWCfQIpGSEA4pw6riI+oGNfOrYKKaJnbqjNEkjoDcU9P/QAjuJh8NbRbNcMKlmkuVLeBdN+TxiLShp5xDwfUs2cv0IO0CRZiRDlAQYRZFHkZtkjjsiSZy62YGsdn/+yQbkW93TnKk1FKeuy24e6mvWr+4X57uFPk/vii6cJjru/dP/ZngL3lf8ROPs7kTD+Ov17PpE26t09v0+O9V4ch0lk9PtJ9Pp5K9Z+udwhPWPAbU+JDwhNMBOOkGOE1NER9kI+Rl2kUNHNHZplBI8OiS82uTPn8u3Ta1BkfHb/wh9M9L51+Bh/wfKaLnqt+CN0TXOjrim39XXHu3iJwAAAP//AwBQSwMEFAAGAAgAAAAhAJIHlOwEAQAAPwMAABoACAF4bC9fcmVscy93b3JrYm9vay54bWwucmVscyCiBAEooAABAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAKySy2rEMAxF94X+g9G+cTJ9UIZxZtFSmG2bfoBwlDhMYgdbfeTva1I6ycCQbrIxSML3Hom72393rfgkHxpnFWRJCoKsdmVjawXvxcvNI4jAaEtsnSUFAwXY59dXu1dqkeOnYJo+iKhigwLD3G+lDNpQhyFxPdk4qZzvkGPpa9mjPmJNcpOmD9LPNSA/0xSHUoE/lLcgiqGPzv9ru6pqND07/dGR5QsWMvDQxgVEgb4mVvBbJ5ER5GX7zZr2HM9Ck/tYyvHNlhiyNRm+nD8GQ8QTx6kV5DhZhLlfE0Zjq58MNnaCObWWLnK3aigMeirf2MfMz7Mxb//ByLPY5z8AAAD//wMAUEsDBBQABgAIAAAAIQDDG6yZDw8AAG1IAAAYAAAAeGwvd29ya3NoZWV0cy9zaGVldDEueG1spFxrb9vIFf1eoP9BIBaFg40t8SXJqq2FJVGSgc1mkTi724+KTMdCJNOV5Dxa9L/3zMy94uXMOKbpoq3iM/femXvmdUgOefbLt8269SXf7lbF3XkQnnSCVn63LK5Xd5/Ogw9X0+N+0NrtF3fXi3Vxl58H3/Nd8Mvw7387+1psP+9u83zfQoS73Xlwu9/fD9rt3fI23yx2J8V9foeSm2K7Wezx5/ZTe3e/zRfX2mmzbkedTre9WazuAhNhsK0To7i5WS3zSbF82OR3exNkm68Xe7R/d7u633G0zbJOuM1i+/nh/nhZbO4R4uNqvdp/10GD1mY5uPx0V2wXH9fI+1uYLJatb1v8N8L/Yq5G405Nm9VyW+yKm/0JIrdNm930T9un7cXyEMnNv1aYMGlv8y8r1YFlqKhZk8L0ECsqg8UNg3UPwRRd28HD6vo8+G+UJtHkojM97l10kuPk9AJj7LTbO06TbBpPo36/34n+FwzPrlfoYZVVa5vfnAejcHDZ7QTt4ZkeQH+s8q878e/WfvHxfb7Ol/sclYRBS43Pj0XxWRleAuog5E4bqJC7f+ugF+Hgr+kk7CT9tNdVsduH4PLfXNFUD+bft63r/GbxsN6/K77O89Wn270K3/pPvi34TzRAj4zB9fdJvltiqMLkJE5VHctijYD4/9ZmpeYchtrim2nz6np/q2Zh1D1N4zBKg9byYbcvNn9SAbkbR/SQdsTvV1N+etKLumn3CT90hvbDL/mFaP0PKkrIAb/kEHVq1YT265rwyzX1Trpx1H8qtS454pcd41o19sgRv+TYP4nr1IhlTjcVv2VT67B5So74Zccfdl/bdL8eZ5PFfjE82xZfW5j7GAe7+4VaSaMBgqlBhfRNt9QZVhQGo8EOEyKOZzQmmBFLVfNI+egGoM4d0C/Dzln7C2bAkizGxiLS40+5TGwgs4GpDcxsYG6AWM87tP7ABAamk0JygkHkn1KchXI7DzBID1mEVhZsoSahzsIAiHxwiaouGVuwy9QGZjYwN4BZTmRaaFmTtJTbeYABfWhjbKXFFoe0DNA/9FZmgNMDMDWAWY4UEzMDqIXgUE1SrWZOJhinB5P0YNKWmYLQJpkqNwxDDItDBV0rVTLRY8Z0oUFErgYQuRpA5EpB9Gaig8wJCZ2hiJnjyaR7WJifnJg8OFUg5Cb57Vm5HUwO/UhIUnakg0wNIrLz1NS3epJMQj3Vo05UslzpRoy5Jt2o3KxuPLVSJRPRjYSUeWSE6Hmke2lqEJGqDcwN4E49JWKttTWssaIoN2tFsRdGNjn0mQHCnuY2STud6DRMQqsHMrIqh+3UICI5G5gbwE0Om0WT5JQbugk/5XppL5gHm0N6BomUmvoyTNKkZ60SGbmIzAwiMqMYeraZ+UeI3l8qQ1BNmCbJaT87O2ttH5dGh/QI4vx6ccfq8IydRIIEiQw5jEiRILf7Qs/GX2dwaj87RXtrKI3KFFV954FYMclILJmEyIyMVyQzMpAnI58GqTHd1Mqv5E+5LIwJKhsycZDMQaYcqKIFrIE6Y6MUQ/lm+P7Dm6NZmAxmSfLqrH3j0UBzdpD7U/jIDhhaGgZpxYCMmjOXEVpwKZUX+eUZLhZ489DRjGE5V+2tkY2kUgitPWbCRlhLykj24sRGckAYURPL/Su0VvUZ+VWMImv+zNlIyojIWncu2cizIDRUUaFRL7ESvTfDycVVdjTvvn7z9rer+dE8+cdic//P8NXrkPs+6YahM5coBMnF/VaF0jTeLrb5dWCuU8cYRONEma70ReeH33/P3h1dZX9dHY3C5HUwwX+CV1zPn9nkt+z95OJf1Q1yws0td7+MoV6p4hgq5zKGsJ5Dcdl3c4KSUu9cMuQKnrChdtN+2CsjLymjMB2MErVegBTIDjDxM/pcTzNFtTU3xxwNo04NU5dqQy7WtC/Dq/mHdz4GjaaLJYMESQYJkgwaCF3IS+acm1MuQ5cMeRj0a0ZE/fHli7pKU2rDYdBwpnNVXJUrjt47x+xXh6vpu0vPWDMVV5giSDJFkGSKWiyZIkgyRZCHqYYCE+KqFlPOCkl+dZh6f3H14Z2HKwohRxVBkisDxZIrspJcEVRaXXJmHq4aStiQBOpTo8q+GGG/Wlx9+M3DFFUsmSJIMmWgClNkJZkiSI4qgjxMNdTDSgnXmX/WbjlmvzpMYb/xMGUqrsw/giRTBMkxRS2WTBEkmSLIZQoqt5G41n41Vir7qo/9KkwN5QYZdewN8uqDb3ukSJIyhgRlDAnKuAmCMoYEZQx5KGso1qGb6wwuWy2N2a/O4HpMS1CMClmmObEkiyBJFjVakkWQJIsgD1kNrwMicx3w1E5oq8Yx+71ANVCICld0m1JyRZDkitosuSJIrO/cQg9X1gWDujCocxWIi6VaA8u+9mW/xqqBAlSYoosEyRRBkilqsWSKIDmqCPIw1fAKIDLK+MlRZct+9nuBauAQYi9kSHJF0l1yRW2WXBEkuSLIw1VDNa+eINXYCyNbt7NfY9XAASRT1BbJlKva2VEyRY6SKYI8TDVU7erivQ5Ttmpnv8aqgQNIpqgtkilXtbOjZIoc5Ur1qGqPGqp27fe0ahA3o831DfvVWtX9asEV7RxUUuWKdraSVLmina08g6qhaI/qifbIFu3s9yK14Op2DivJcnU7W0myXN3OVh6yGur2qJ5uj2zdzn4vUQtUtZyDBEmuXOXOdUuuXOXOVi5X6hZbo0ekyq/GHLSVu64Pfo3VAgcQTDEkmCJIXg2ylWCKIbGuM+RhqqFgj+sJ9th+HsR+L1ALHEJyRa2RXLl6nR0lV65eZysPVw31elxPr8fO43jya6wWuGLJFMWUTLlqnR0lU+QoRxVBHqaUOm3wQFE9A6gx/2JbrbNfY7XAASRT1BbJlIEq84+sJFMESaYI8jDVUK3jRn0tpmy1zn7N1QJHkFRRYyRVrlhnR0mVK9bZykNVQ7GOB0W1qLLFOvu9RC1wDEmWq9fJqjKu3LvsHEuOq0f1Oq47m83Aeno9tvW6rq/mDvjIEwkOIblyFTtZVbhy77NzLMnVo4odt3macVXvPnts32fX9b1ILbiCnWPKKegKdraSU5BiSaYI8kzBhoIdvVVrCtqCnf1eohZcvc5RJVeuXmcryZWr19nKw1VDvY5Hk7W4svU6+zVXC65a55iSKVets5VkylXrbOUyhSevzU7+1VPrsa3WdX0155/3iQQHECsVQ4IpguRKxVaCKYbE/GPIw1RDta4fLT99XZPYap39mqsFjiCpcsU6WVWocm+ucyxJ1aM313Efs9mgkmJ9eDk9wgOpo+ytOQnx0+gnnFHovHp1HvVfB8FrPLP/WZ6JSGwJr1tRc6g99nyCY0gKXRVPVhUK3XvuHEtS+KiKV5ezjU7kShWvKBwleN8iUIQ9zuep4TMCn+KQSWIrfd2ml+kMDiH5dLU+WVX4dLU+x5J8Pqr19amXBldF2g85Y00tT3sYXuOneI3xxFCN07jkVb3GcTi4TrFTOrj0o5Mk0m/CfnqN0jciM4bKU/BThsoTazg/pre3VC6BBGkSK8cbk4aSX/udB2nZxWOGygV64kIZQ9joyyPv1hiclkZ8AmZGEO6y1zwsxzH0OaRqzg2Ve2L0rXkHyNwaJqgcnBMHyRxk6iAzB5kTgud1alRUTvE3FNOJEZ2V5htINt9GMnIrbaYOMnOQOSG+5nsUbhI//XJFYqRhF/tROXCcFxHISLwnwn7lDMlcaOpCMxeaM6SnVrVTpBZFI+o+PVVLDu7HVDqFINkrBMkF1YGmHKu0mrnQnCE9TasvTUiV+Iwc1NJm5cBQ+TxnwlB5LjBzoakLzVxoTlBPHyis5iD123NyMDJH9kNqINENDpI5yNRBZg4yJ8QzOVKpqcrmN3jdQ0fCWzuV6WLfgGEjMV0YEtOFoXJDmbqOM9dxXnGsdpMUPi/L04gBeWQ7NVBP9pwDZa7VlCGs8+USY92HmZVGh9OZDHmmlLz5+JzhaDbsXmWXdF5FIiOZpwNl0AFqdgo2pgSVfjMHmRPiG6NSMjwnJ3MrrTLFDFTpKQfKUgeaEiQzMEYlMicbXwZSADwnA1cAQABpdkk5mveZ7DPvbCSWb4bE1SxBMicTXOZkEF9OUhU8Jyez4ffLs9njlKBSeU5cKHOhqQvNXGjOkKtsUikNKAck/Ow3UNNnbsaYyPTuKgZneUFWg0WlzcSLtJt8+ykf5+v1rrUsHtQr11jqh2cHmA7op91BZpSzVTJJ4kGGKx6EdUoSlOgV2I6W9AdqQ3R9xmmEejTNlg/Wb/j4SsYpWoCZ40bD6ggfb9vi3iAzJ/nsVsd9lGgR4JScokRrASfTDjL15TNJQpToUen4IFNceXt4i+GDB46+EvjgMZyvBBzgeZWvBL2A5ye+khQlPt4mMXrbHFKzWx3BBxc0nmgRfHBMxlcCrnHew1cCrnG4wVcCrvEo39dqcI1H156SED44WOyLBh+cT/WVgGvzfpaTKbg2i5ZTAq7N5u2UgGscUvO1DT54QcdXAh+8WuIrAdd4ZcJXAq7xCoOvBFzjQLyvBFzjgLpbMuoNxl4PhPLho/5g7I1zOhj72B+FnYF6bdBTcxiixDvSUZJ5S0ZYBnwzYBQPxv5VCNPMV3c6GHu5BeleHJRrxtvliqk+SHF3vVKfUVmszQtqe3wDhj+UgReSBrgn0prid46FEOY37x7WeWv//R5fR8m/4dMuO/MJkutvN+Z7G6377arY4oMq+pWfM/UlmIf1Yvj23dFPeEnqPOBXLILX9Ld+jSBQd8fIFPdxdC3qho6vdfhICL6g8MdivcKv+gAMr/tq2a8WUUOv8+Vqs1gHrcV6XXwdrRd3n82XF26Lr5d39w/7N0hj8QkpqcsPgNl2W2wlmCvgarVXH4UZXf7668Xo16w1f4t3kPDlHFUGeLVeq6/GtG6Lh+2utcEnPFof81b/JG0V29YaNZwguPkCCb1zKL6P0u+k0WTU6x73o3H/OMEXUY5xe+viOBn38a7BRSfsjrrq+yhEUqi+1XD4N6PREA8FCI1AX5UMB8D3SO6R9ZvF9tMKLK7zG2ye+MYH3mmjz5uof++Le41C630s9vgsCf91i8/65EgcXzgJWjdFsec/ME/ahw8FDf8PAAD//wMAUEsDBBQABgAIAAAAIQD2YLRBuAcAABEiAAATAAAAeGwvdGhlbWUvdGhlbWUxLnhtbOxazY8btxW/B8j/QMxd1szoe2E50Kc39u564ZVd5EhJlIZeznBAUrsrFAEK59RLgQJp0UuB3nooigZogAa55I8xYCNN/4g8ckaa4YqKvf5AkmJ3LzPU7z3+5r3HxzePc/eTq5ihCyIk5UnXC+74HiLJjM9psux6TybjSttDUuFkjhlPSNdbE+l9cu/jj+7iAxWRmCCQT+QB7nqRUulBtSpnMIzlHZ6SBH5bcBFjBbdiWZ0LfAl6Y1YNfb9ZjTFNPJTgGNQ+WizojKCJVund2ygfMbhNlNQDMybOtGpiSRjs/DzQCLmWAybQBWZdD+aZ88sJuVIeYlgq+KHr+ebPq967W8UHuRBTe2RLcmPzl8vlAvPz0MwpltPtpP4obNeDrX4DYGoXN2rr/60+A8CzGTxpxqWsM2g0/XaYY0ug7NKhu9MKaja+pL+2wznoNPth3dJvQJn++u4zjjujYcPCG1CGb+zge37Y79QsvAFl+OYOvj7qtcKRhTegiNHkfBfdbLXbzRy9hSw4O3TCO82m3xrm8AIF0bCNLj3FgidqX6zF+BkXYwBoIMOKJkitU7LAM4jiXqq4REMqU4bXHkpxwiUM+2EQQOjV/XD7byyODwguSWtewETuDGk+SM4ETVXXewBavRLk5TffvHj+9Yvn/3nxxRcvnv8LHdFlpDJVltwhTpZluR/+/sf//fV36L///tsPX/7JjZdl/Kt//v7Vt9/9lHpYaoUpXv75q1dff/XyL3/4/h9fOrT3BJ6W4RMaE4lOyCV6zGN4QGMKmz+ZiptJTCJMLQkcgW6H6pGKLODJGjMXrk9sEz4VkGVcwPurZxbXs0isFHXM/DCKLeAx56zPhdMAD/VcJQtPVsnSPblYlXGPMb5wzT3AieXg0SqF9EpdKgcRsWieMpwovCQJUUj/xs8JcTzdZ5Radj2mM8ElXyj0GUV9TJ0mmdCpFUiF0CGNwS9rF0FwtWWb46eoz5nrqYfkwkbCssDMQX5CmGXG+3ilcOxSOcExKxv8CKvIRfJsLWZl3Egq8PSSMI5GcyKlS+aRgOctOf0hhsTmdPsxW8c2Uih67tJ5hDkvI4f8fBDhOHVypklUxn4qzyFEMTrlygU/5vYK0ffgB5zsdfdTSix3vz4RPIEEV6ZUBIj+ZSUcvrxPuL0e12yBiSvL9ERsZdeeoM7o6K+WVmgfEcLwJZ4Tgp586mDQ56ll84L0gwiyyiFxBdYDbMeqvk+IhDJJ1zW7KfKISitkz8iS7+FzvL6WeNY4ibHYp/kEvG6F7lTAYnRQeMRm52XgCYXyD+LFaZRHEnSUgnu0T+tphK29S99Ld7yuheW/N1ljsC6f3XRdggy5sQwk9je2zQQza4IiYCaYoiNXugURy/2FiN5XjdjKKbewF23hBiiMrHonpsnrip8TLAS//Hlqnw9W9bgVv0u9sy+vHF6rcvbhfoW1zRCvklMC28lu4rotbW5LG+//vrTZt5ZvC5rbgua2oHG9gn2QgqaoYaC8KVo9pvET7+37LChjZ2rNyJE0rR8JrzXzMQyanpRpTG77gGkEl/p5YAILtxTYyCDB1W+ois4inEJ/KDBdzKXMVS8lSrmEtpEZNv1Uck23aT6t4mM+z9qdpr/kZyaUWBXjfgMaT9k4tKpUhm628kHNb0PdsF2aVuuGgJa9CYnSZDaJmoNEazP4GhK6c/Z+WHQcLNpa/cZVO6YAaluvwHs3grf1rteoZ4ygIwc1+lz7KXP1xrvaOe/V0/uMycoRAK3FXU93NNe9j6efLgu1N/C0RcI4JQsrm4TxlSnwZARvw3l0lvvuPxVwN/V1p3CpRU+bYrMaChqt9ofwtU4i13IDS8qZgiXoEtZ4CIvOQzOcdr0F9I3hMk4heKR+98JsCYcvMyWyFf82qSUVUg2xjDKLm6yT+SemigjEaNz19PNvw4ElJolk5DqwdH+p5EK94H5p5MDrtpfJYkFmquz30oi2dHYLKT5LFs5fjfjbg7UkX4G7z6L5JZqylXiMIcQarUB7d04lHB8EmavnFM7DtpmsiL9rO1Oe/a1DriIfY5ZGON9Sytk8g5sNZUvH3G1tULrLnxkMumvC6VLvsO+87b5+r9aWK/bHTrFpWmlFb5vubPrhdvkSq2IXtVhluft6zu1skh0EqnObePe9v0StmMyiphnv5mGdtPNRm9p7rAhKu09zj922m4TTEm+79YPc9ajVO8SmsDSBbw7Oy2fbfPoMkscQThFXLDvtZgncmdIyPRXGt1M+X+eXTGaJJvO5LkqzVP6YLBCdX3W90FU55ofHeTXAEkCbmhdW2FbQWe3Zgnqzy0WzBbsVzsrYa/WqLbyV2ByzboVNa9FFW11tTtR1rW5m1g7LntqkYWMpuNq1IrTJBYbSOTvMzXIv5JkrlVfacIVWgna93/qNXn0QNgYVv90YVeq1ul9pN3q1Sq/RqAWjRuAP++HnQE9FcdDIvnwYw2kQW+ffP5jxnW8g4s2B150Zj6vcfONQNd4330AE4f5vIMCRQCscBfWwFw4qg2HQrNTDYbPSbtV6lUHYHIY92LSb497nHrow4KA/HI7HjbDSHACu7vcalV6/Nqg026N+OA5G9aEP4Hz7uYK3GJ1zc1vApeF170cAAAD//wMAUEsDBBQABgAIAAAAIQBowXh5+wcAADNSAAANAAAAeGwvc3R5bGVzLnhtbOxcWY/bNhB+L9D/IAh99eqw5WNrO83uxkCANAiQLVCg6YNWlr1CdBgSvbET9L93qMMiLVOiTmvR7kNiyTL5zT2coTh/c3Bs4cX0A8tzF6JyI4uC6Rre2nK3C/GPx9VgKgoB0t21bnuuuRCPZiC+Wf780zxAR9v8/GyaSIAh3GAhPiO0u5WkwHg2HT248XamC99sPN/REVz6WynY+aa+DvCPHFtSZXksObrlitEIt47BM4ij+1/3u4HhOTsdWU+WbaFjOJYoOMbt+63r+fqTDVAPykg3hIMy9lXh4CeThHcz8ziW4XuBt0E3MK7kbTaWYWbhzqSZpBvpSDBytZEUTZJVivaDX3GkkeSbLxYWn7icu3tn5aBAMLy9i0Ccp1tC9M37Ndwcj0Qhksq9twY+/fXLQJHh7+/1F8FxnC/CEf5+/U2UlnMpHnA533huOu4YWICZe/vV9b65K/xVNBl+ajkPvgsvug13FDyG4dmeLyBQCpgrvOPqjhk98XaHvED4qPu+9w0/u9Edyz5G36n4RqhM8cOOBaINUUXTRP8+4aeuNmEIkqJQxggbpTChbtYBN0+cDMnoUnTtTiiF+gsKZtn2yTwmYB74xnIOngSZvruCCyH+/Hjcgb664PQilQufK3h66+tHRdX4fxB4trXGKLb3GSt5iu9Z7to8mGC4YLfYJAmscBUhKsDFmAZ8PbKwnxjIN+poNptMFPw3mc6GeKZuASgAYDYdTUbyZKSpYzX0FF1A6JzSM0LLsjoUeQDi8fw1xO3E1w9l0KPo3nJumxsEdPnW9hn/j7wdptJDCILbcr629K3n6jbWpuQX5C8h4ENsX4joGWJz4sDP1RBPQc3A9StAUgIITUKLE3ANHXIzZGYLSMryvGXEXMNHClWsT4kYS2Lun8r2gStlda9RzD3T0tZ8WZuqV9UKrirIsyDSK/fQPw2PFztkfnUh6EtpQDnp20VaOIfjdcd15iCifp1hXjflkTE04ogbFm1upC2tlX2ksFfKWwcMkQm/dqM/Xxr0ybpzli3NOLD/dSCsXLB14L/HoJz4wOmDO3YOHYSNfhLeyHqmmbhankFVVz+cMzWe6jXkCLoju307PveajVRUrqWP7a/DGtLc1mMVh4rSpc389V9dLWlwvVIGSlxbhlK1Ydr2Z1xT/nND9SYPG6IvCatl3KTBLUr8Efot8ceoNB1dACkSOVo0NjHsdAgV6/LjCofNaQLWr9UUFXxMUSnQDY1/Lei7nX3EDVHc6oyvgJL06i6s2qfXb21r6zpm0q3Vk0vh2fOt7zAQbqRiVxz2fg4bNnEseHC/z/CG/YBHqBwgSoULvfIz7nUkTgIQYCBtgAEITIPS3LjPH1kU0JRYFDWadk7ex73zZPqrcNMGocNNa7QBGm9Gzfw8na4tlVymEKMDH1IW4+0NlEEzmMKvCZEn+OR7yDRQtNMG3FSOseMtOcgysPHDWgDb/u70Y8H2jK+4IRy2zKUM/yjBs2icvG4aLweL3lkuwMxxJbmxoLTKlNIRDv5lrADvjugikLGMPqOyVwBECRS2xtF+oh8Cxdu74uwp361VzlPK6Gbq6fk9GKGdLGKyylA1RHVIDCumt5FklI+wVGLQBqSwLlyYxxKpSoE/7y7NZoXUfJbVThtYKUKebLnSPlBErgwnh8NcC5eaSpjJTlke/zzNuxzeKGnNeBP7CmNn2Evj5mIdn7HAIvNiXp813ytl8CDCBCAkFMSiOcujlhASqTSnOl4Rc7jSbxBx+7WJmKcNYm6/YNE85itqs/DN13eP5iEsOeH6WN6CmqXa7fiLdBGbcLwMViL1o9LuAl4zvWsGTRGziHQtH0DNKblCtpIplPWrVMOqMCkVVSsn7BJTUXJRsvkCFVPOl2fn9ZhKCxQmmOz6+YpgCpZLnXKmJBbKuDpcrDEFm1n20+voVnhZ2hdWqgYrmaT4lfiY5qoBuVXZooDBVJlXrvJqQcjt1H3Am0G5Da1uwRQkq82BwXE06Y+y9CwrJx77bRBj9GppfstJfaWJjNr+mpqrz8sUfp+yH7VP2U+R5Lr1GAVq1FgPoWbhnalmnaWSXEXXTDAoV3StUNzLenxGXyxnbFbZrmLpowIV2W0XnFRwSYU3j+zRnNHKgrmSL1O4IOoGVAMum7jU76iWr7BTkLiVuYA95G6lmpC4NbM7SNy9gu4gNdFiqCko7u5Md1yBE0ea365V18Sy9YKrm73SD7uv1pAroU58TTSWv66zA+Ya23doz17HaV0dfZ29PlcHX2fvRA92o6hthDuupS1rf0y3O7+6yHDb9GGUV+3DLj4KEH+IrrI24461FQbnD5q4TozfUoD3EoiXH6hXH04vMQj4rLCF+BHvuLaJ7OVpb9lwOtOF1x5gzPWBfpECri8cZZUe4XThxRl83tJ4MtM0bazJmhydt3TxcCkpHB3/CzMjfLZe+C7HiRyw2rW50fc2ejx9uRDTz7+ba2vvQMYRP/XJevFQOMRCTD9/wGcVKWO8PQx62R8COFwI/hf2vrUQf7y7m8we3q3UwVS+mw5GQ1MbzLS7h4E2ur97eFjNZFW+/4c44a/G+X7hgYSwlV4Z3QY2nALox8TG4D+n9xYicRHBDze3AWwS+0wdy281RR6shrIyGI316WA6HmqDlaaoD+PR3TttpRHYtYrnAMqSokQnCmLw2i2yHNO23ERWiYTIuyAkuMwhQkokIaWnPS7/BQAA//8DAFBLAwQUAAYACAAAACEAjqCrVJIBAACqAwAAFAAAAHhsL3NoYXJlZFN0cmluZ3MueG1sfJNRT+QgFIXfN9n/cNN3p6MPZmPaGmxxprEtDUWNjziDMyQFukDN+u+3xs0mwugj59zvcHIJ2fUfNcKrsE4anSfnq3UCQu/MXupDntyz27NfCTjP9Z6PRos8eRMuuS5+/sic87Cw2uXJ0fvpKk3d7igUdyszCb04L8Yq7pejPaRusoLv3VEIr8b0Yr2+TBWXOoGdmbXPk4vlllnL37Mo/wtF5mSR+YLVLR62GDM4g4FRxPCmLmEg97Ssu02W+iJL3yc/pntEWYdpKJeo3+C27moYFgj1hGLoGYaGVV8kAG77FXSoxeFASzq2DUU0WTlGt9bV1bfxlDRR/BNGUf3KvK8riiLQzepZ2BPG4Ln1UHEv4gYEsN6f9gby+C8zwm7qpkE3DYbtsvoBIn9rZuvCJtXyXLH2FG2vZPVDzSL9kdA7uKWkDYHPZUKX4hbRu1BlhKEGvidbo/0R+M7LV+mlcOBmpbh9C7M6ruLFlnw6CCW1BCrcZLSTz6MIyUEeNPezjXG8QafBdPlrxV8AAAD//wMAUEsDBBQABgAIAAAAIQDfLqcEcwEAANIFAAAQAAAAeGwvY2FsY0NoYWluLnhtbHSU627CIBiG/y/ZPRD+T+Qwd4jVBJJ5A9sFkIq2SUtNaZbt7scW+Zgc/phIH5/yva+w3X+NA/o0s+sn22C6WmNkbDsde3tu8Mf728MzRm7R9qiHyZoGfxuH97v7u22rh1Z1urfIG6xrcLcsl1dCXNuZUbvVdDHWPzlN86gX/3U+E3eZjT66zphlHAhbrzdk9AK827ZobvBBPGLU+01gNPx+kuu6pOK6HleA9C+O5IGyhFS05txkZP6Wp+J+FM1/WyYl9en9TRR3/lJxZiTzVRTSUDQYwMl8YiUSDJEM+dwmrMAQSMXKpGQ8nYiF3FJnSqoKKVnoKO4zJJw6c7LSEUs7kiwknDqDIc5eJiXLkueVjgpkpSMwwOy80hEPhkiGhJOJwAAT8TIpefaf5+XzogpkpSMwxLeXScnT5CWvdFQgK+cIDJCSqHTEszZFpSMwwEQVUors/oGV5E4T2TkSlXMU12/uOvXfQOA23v0AAAD//wMAUEsDBBQABgAIAAAAIQAwIYPQ9gAAAEUBAAAZAAAAZG9jTWV0YWRhdGEvTGFiZWxJbmZvLnhtbFTQ3WqEMBAF4FcJuY+JrjGuqAu9K2yhrxCTcQ3kZzFTaSl998a79u4wMB9zZrx9Bk8O2LNLcaJ1JSiBaJJ18THRD1xZT0lGHa32KcJEvyDT2zwav/jB6wX83WUkBYl5OIcT3RCfA+fZbBB0roIze8ppxcqkwNO6OgO8EY3gwT3vp/AGqK1GTf+yxNmJfoPqhGmvLeukFKxtl5b1jeyZUqZW0irVL+rnvFgvHspCTUkA3FKJ77s7nIcH2FLAIbyenup0o0Fqdl1F8dZuYVcJlkl7uRTNtLari2dSRIj44jBPtHxkh5CO0y+ZzyP/337+BQAA//8DAFBLAwQUAAYACAAAACEA0B2j6k4BAAB/AgAAEQAIAWRvY1Byb3BzL2NvcmUueG1sIKIEASigAAEAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAjJJfS8MwFMXfBb9DybNt2k2nlLYDlT05EJz45+2S3G1xTRqSzG7f3rTdamUKPibn3F/OuSSb7mQZfKKxolI5SaKYBKhYxYVa5eR5MQtvSGAdKA5lpTAne7RkWpyfZUynrDL4aCqNxgm0gScpmzKdk7VzOqXUsjVKsJF3KC8uKyPB+aNZUQ1sAyukozieUIkOODigDTDUPZEckJz1SL01ZQvgjGKJEpWzNIkS+u11aKT9daBVBk4p3F77Toe4QzZnndi7d1b0xrquo3rcxvD5E/o6f3hqq4ZCNbtiSIqMs5QZBFeZYg5lKTZgPrYKLoI3kBpK2GwzOvA0+yzBurlf/VIgv93/PXZq9a+15bonkQc+btqVOyov47v7xYwUo3g0CeObMIkXyXUaX6WX4/cmyY/5Jn53IQ95/k+ceOiAeAQUGT35MsUXAAAA//8DAFBLAwQUAAYACAAAACEAb3gbEpEBAAAaAwAAEAAIAWRvY1Byb3BzL2FwcC54bWwgogQBKKAAAQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACckkFv2zAMhe8D9h8M3Rs53VAMgayiSDf0sGEBkrZnVqZjobIkiKyR7NdPtpHU2XbajXx8ePpESd0eOlf0mMgGX4nlohQFehNq6/eVeNx9u/oiCmLwNbjgsRJHJHGrP35QmxQiJrZIRY7wVImWOa6kJNNiB7TIY58nTUgdcG7TXoamsQbvg3nr0LO8LssbiQdGX2N9Fc+BYkpc9fy/oXUwAx897Y4xA2t1F6OzBjjfUv+wJgUKDRdfDwadkvOhynRbNG/J8lGXSs5btTXgcJ2DdQOOUMl3QT0gDEvbgE2kVc+rHg2HVJD9ldd2LYoXIBxwKtFDsuA5Yw22qRlrF4mTfg7plVpEJiWzYRLHcu6d1/azXo6GXFwah4AJJA8uEXeWHdLPZgOJ/0G8nBOPDBPvhLMd+KYz53zjlfNJf2SvQxfBH/Ua4h47662SJ0l9t/6VHuMu3APjaa+Xotq2kLDOT3He+1lQD3mlyQ0h6xb8HuuT5+/B8Auepq+ulzeL8lOZH3imKfn+qfVvAAAA//8DAFBLAQItABQABgAIAAAAIQCVZpbtkAEAAIIFAAATAAAAAAAAAAAAAAAAAAAAAABbQ29udGVudF9UeXBlc10ueG1sUEsBAi0AFAAGAAgAAAAhADEdic0iAQAA3gIAAAsAAAAAAAAAAAAAAAAAyQMAAF9yZWxzLy5yZWxzUEsBAi0AFAAGAAgAAAAhAD4Ms3q3AwAAeQkAAA8AAAAAAAAAAAAAAAAAHAcAAHhsL3dvcmtib29rLnhtbFBLAQItABQABgAIAAAAIQCSB5TsBAEAAD8DAAAaAAAAAAAAAAAAAAAAAAALAAB4bC9fcmVscy93b3JrYm9vay54bWwucmVsc1BLAQItABQABgAIAAAAIQDDG6yZDw8AAG1IAAAYAAAAAAAAAAAAAAAAAEQNAAB4bC93b3Jrc2hlZXRzL3NoZWV0MS54bWxQSwECLQAUAAYACAAAACEA9mC0QbgHAAARIgAAEwAAAAAAAAAAAAAAAACJHAAAeGwvdGhlbWUvdGhlbWUxLnhtbFBLAQItABQABgAIAAAAIQBowXh5+wcAADNSAAANAAAAAAAAAAAAAAAAAHIkAAB4bC9zdHlsZXMueG1sUEsBAi0AFAAGAAgAAAAhAI6gq1SSAQAAqgMAABQAAAAAAAAAAAAAAAAAmCwAAHhsL3NoYXJlZFN0cmluZ3MueG1sUEsBAi0AFAAGAAgAAAAhAN8upwRzAQAA0gUAABAAAAAAAAAAAAAAAAAAXC4AAHhsL2NhbGNDaGFpbi54bWxQSwECLQAUAAYACAAAACEAMCGD0PYAAABFAQAAGQAAAAAAAAAAAAAAAAD9LwAAZG9jTWV0YWRhdGEvTGFiZWxJbmZvLnhtbFBLAQItABQABgAIAAAAIQDQHaPqTgEAAH8CAAARAAAAAAAAAAAAAAAAACoxAABkb2NQcm9wcy9jb3JlLnhtbFBLAQItABQABgAIAAAAIQBveBsSkQEAABoDAAAQAAAAAAAAAAAAAAAAAK8zAABkb2NQcm9wcy9hcHAueG1sUEsFBgAAAAAMAAwABQMAAHY2AAAAAA=="""

def get_default_template_bytes():
    """Decode the hardcoded default template back into raw .xlsx bytes."""
    return base64.b64decode(DEFAULT_TEMPLATE_B64.strip())


# =====================================================================
# 2. HELPER FUNCTIONS
# =====================================================================

def norm_name(name):
    """Normalize a name to a comparable token set regardless of 'First Last'
    vs 'Last, First' formatting differences between source files."""
    parts = re.split(r"[,\s]+", str(name).strip().lower())
    parts = [p for p in parts if p]
    return " ".join(sorted(parts))


def name_tokens(name):
    parts = re.split(r"[,\s]+", str(name).strip().lower())
    return {p for p in parts if p}


def find_matching_rep_key(sow_name, rep_name_keys):
    """Match a SOW resource name to a Replicon name key even when one side
    has an extra middle name (e.g. SOW 'Kalyani Ghaytadkar' vs Replicon
    'Ghaytadkar, Kalyani Popatrao'). A match is any rep key whose token set
    fully contains the SOW name's tokens, or vice versa."""
    sow_tok = name_tokens(sow_name)
    best_match = None
    for rep_key in rep_name_keys:
        rep_tok = set(rep_key.split())
        if sow_tok.issubset(rep_tok) or rep_tok.issubset(sow_tok):
            # Prefer the closest-length match if multiple candidates exist
            if best_match is None or abs(len(rep_tok) - len(sow_tok)) < abs(len(set(best_match.split())) - len(sow_tok)):
                best_match = rep_key
    return best_match

def get_field(row, candidates, default=""):
    """
    Flexible lookup of a value from a SOW row, regardless of small
    header-naming differences (spacing, trailing colon, casing) between
    what the code expects and what the uploaded SOW file actually has.

    `candidates` is a list of possible header names to try, in priority
    order. Returns `default` if none of them are found or all are blank.
    """
    def normalize(s):
        return re.sub(r"\s+", " ", str(s).strip().lower().rstrip(":"))

    normalized_map = {normalize(col): col for col in row.index}

    for cand in candidates:
        key = normalize(cand)
        if key in normalized_map:
            val = row[normalized_map[key]]
            if pd.notna(val) and str(val).strip() != "":
                return val
    return default


# def load_sow(file):
#     """Load the SOW & Resource Details file into a DataFrame."""
#     df = pd.read_excel(file)
#     df["name_key"] = df["Resource name"].apply(norm_name)
#     return df
def load_sow(file):
    df = pd.read_excel(file)
    name_col = None
    for candidate in ["Resource name", "Employee Name", "Name"]:
        for col in df.columns:
            if re.sub(r"\s+", " ", str(col).strip().lower()) == candidate.lower():
                name_col = col
                break
        if name_col:
            break
    if name_col is None:
        raise ValueError("Could not find a resource/employee name column in the SOW file.")
    df["name_key"] = df[name_col].apply(norm_name)
    return df

# def load_replicon(file):
#     """Load the Replicon dump (Expenditure Details Report) into a DataFrame.
#     Row 1 is an instructions banner, row 2 holds the real headers."""
#     df = pd.read_excel(file, sheet_name="Expenditure Details Report", header=1)
#     df = df.dropna(subset=["Employee Name/Supplier Name"])
#     df["name_key"] = df["Employee Name/Supplier Name"].apply(norm_name)
#     df["Item Date"] = pd.to_datetime(df["Item Date"], format="%d-%b-%Y")
#     df["Quantity"] = pd.to_numeric(df["Quantity"], errors="coerce").fillna(0)
#     return df

# def load_replicon(file):
#     """Load the Replicon dump into a DataFrame. Sheet is 'Sheet1' with
#     headers on row 1. Employee name, date, and hours columns are renamed
#     to the internal standard names used throughout the rest of the app."""
#     df = pd.read_excel(file, sheet_name="Sheet1")
#     df = df.dropna(subset=["Employee Name"])
#     df["name_key"] = df["Employee Name"].apply(norm_name)
#     df["Item Date"] = pd.to_datetime(df["Expenditure Item Date"])
#     df["Quantity"] = pd.to_numeric(df["QTT"], errors="coerce").fillna(0)
#     return df
def load_replicon(file):
    """Load the Replicon dump into a DataFrame. Sheet is 'Sheet1' with
    headers on row 1. Employee name, date, and hours columns are renamed
    to the internal standard names used throughout the rest of the app."""
    # df = pd.read_excel(file, sheet_name="Sheet1")
    # xls = pd.ExcelFile(file)
    # if "Sheet1" in xls.sheet_names:
    #     sheet_to_use = "Sheet1"
    # else:
    #     sheet_to_use = xls.sheet_names[0]
    #     st.warning(f"Replicon file has no sheet named 'Sheet1' — using the first sheet found: '{sheet_to_use}'.")
    # df = pd.read_excel(file, sheet_name=sheet_to_use)
    # xls = pd.ExcelFile(file)
    # if "Sheet1" in xls.sheet_names:
    #     sheet_to_use = "Sheet1"
    # else:
    #     sheet_to_use = xls.sheet_names[0]
    #     st.warning(f"Replicon file has no sheet named 'Sheet1' — using the first sheet found: '{sheet_to_use}'.")
    # file.seek(0)   # ← rewind before reading again
    # df = pd.read_excel(file, sheet_name=sheet_to_use)
    
    # df = df.dropna(subset=["Employee Name"])
    xls = pd.ExcelFile(file)
    if "Sheet1" in xls.sheet_names:
        sheet_to_use = "Sheet1"
    else:
        sheet_to_use = xls.sheet_names[0]
        st.warning(f"Replicon file has no sheet named 'Sheet1' — using the first sheet found: '{sheet_to_use}'.")
    df = xls.parse(sheet_to_use)

    # Drop exact duplicate line items: same Exp. Item Id AND same
    # Expd Line Num means the same row got exported twice (not a real
    # correction). Real corrections have the SAME Exp. Item Id but a
    # DIFFERENT Expd Line Num (e.g. line 2 = reversal, line 3 = rebooking),
    # so those are left alone.
    before = len(df)
    df = df.drop_duplicates(subset=["Exp. Item Id", "Expd Line Num"], keep="first")
    dropped = before - len(df)
    if dropped:
        st.info(f"Removed {dropped} duplicate Replicon line(s) (same Exp. Item Id and Line Num).")

    df["name_key"] = df["Employee Name"].apply(norm_name)
    df["Item Date"] = pd.to_datetime(df["Expenditure Item Date"])
    df["Quantity"] = pd.to_numeric(df["QTT"], errors="coerce").fillna(0)
    return df

# def build_daily_hours(rep_df):
#     """Sum hours per employee per date (nets out negative correction rows),
#     and track which (employee, date) pairs had a correction so we can flag
#     them in the Remark column."""
#     daily = rep_df.groupby(["name_key", "Item Date"])["Quantity"].sum().reset_index()
#     neg = rep_df[rep_df["Quantity"] < 0][["name_key", "Item Date"]]
#     neg_set = set(zip(neg["name_key"], neg["Item Date"]))
#     return daily, neg_set

def build_daily_hours(rep_df):
    """Sum hours per employee per PROJECT per date (nets out negative
    correction rows), and track which (employee, project, date) pairs
    had a correction so we can flag them in the Remark column."""
    daily = rep_df.groupby(["name_key", "Project Number", "Item Date"])["Quantity"].sum().reset_index()
    neg = rep_df[rep_df["Quantity"] < 0][["name_key", "Project Number", "Item Date"]]
    neg_set = set(zip(neg["name_key"], neg["Project Number"], neg["Item Date"]))
    return daily, neg_set


def determine_work_from(role):
    """Determine Work From value from the Partner Emp. Role string.

    Only fills "WFO" when the role string explicitly contains "onsite".
    Everything else (offshore, unrecognized, blank) is left empty —
    we don't guess WFH.
    """
    role_str = str(role).lower()

    if "onsite" in role_str:
        return "WFO"

    return ""  # not onsite -> leave blank, don't assume WFH
def is_valid_cid(cid):
    """CID must be 'C' followed by exactly 8 digits, e.g. C9000127."""
    return bool(re.fullmatch(r"C\d{8}", str(cid).strip()))
# def determine_work_from(role):
#     """Determine WFH vs WFO from the Partner Emp. Role string.

#     Rule (per business input): the role code contains a segment like
#     'T&M_13' — the number right after 'T&M_'. Only the LAST digit of that
#     number matters (the leading digit, e.g. the '1' in '13', is ignored):
#         - last digit 3  -> WFO (work from office)
#         - last digit 2  -> WFH (work from home)
#     Falls back to a keyword check ("onsite" -> WFO) and then defaults to
#     WFH if the pattern can't be parsed, so nothing breaks on unexpected
#     role formats.
#     """
#     role_str = str(role)

#     # Explicit keyword override, if present
#     if "onsite" in role_str.lower():
#         return "WFO"

#     match = re.search(r"T&M[_\s]*(\d+)", role_str)
#     if match:
#         last_digit = match.group(1)[-1]
#         if last_digit == "3":
#             return "WFO"
#         elif last_digit == "2":
#             return "WFH"

#     # Unrecognized pattern - default to WFH, but this should be reviewed
#     return "WFH"

# def fill_timesheet_for_employee(template_bytes, res_row, daily_hours, neg_set, month, year, matched_rep_key):
def fill_timesheet_for_employee(template_bytes, res_row, daily_hours, neg_set, month, year, matched_rep_key, hours_used_so_far=0):
    """Fill one copy of the timesheet template for a single employee and
    return it as an in-memory .xlsx (BytesIO)."""
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    # # ws = wb["Sheet1"]
    # if "Sheet1" in wb.sheetnames:
    #     ws = wb["Sheet1"]
    # else:
    #     ws = wb.active  # falls back to the workbook's active/default sheet
    # if len(wb.sheetnames) > 1:
    #     st.warning(f"Custom template has multiple sheets ({', '.join(wb.sheetnames)}). Using '{wb.sheetnames[0]}'.")
    # ws = wb[wb.sheetnames[0]]
    if "Actual template" in wb.sheetnames:
        ws = wb["Actual template"]
    elif len(wb.sheetnames) > 1:
        st.warning(f"Custom template has multiple sheets ({', '.join(wb.sheetnames)}) and no sheet named 'Actual template' — using '{wb.sheetnames[0]}'.")
        ws = wb[wb.sheetnames[0]]
    else:
        ws = wb[wb.sheetnames[0]]
    ws.column_dimensions["H"].width = 32

    name_key = matched_rep_key  # the Replicon-side key matched to this SOW resource

    # ---- Pull all resource/PO/SOW fields from the SOW & Resource Details
    # ---- file, tolerant of header-name variations in that file.
    emp_name = get_field(res_row, ["Resource name", "Employee Name", "Name"])
    cid = get_field(res_row, ["CID", "C.I.D", "C.I.D.", "Cid"])
    role = get_field(res_row, ["Partner Emp. role", "Partner Emp Role", "PARTNER EMP. ROLE", "Emp Role", "Role"])
    domain = get_field(res_row, ["Domain"])
    po_number = get_field(res_row, ["PO Number", "PO Number:", "PO No", "PO No."])
    po_start = get_field(res_row, ["PO Start Date", "PO Start Date:", "PO Start"])
    po_end = get_field(res_row, ["PO End Date", "PO End Date:", "PO End"])
    sow_number = get_field(res_row, ["SOW number", "SOW Number", "SOW Number:", "SOW No"])

    work_from_value = determine_work_from(role)
    planned_hours_raw = get_field(res_row, ["SOW planned duration hours", "SOW Planned Duration Hours", "Planned Hours"], None)
    planned_hours = float(planned_hours_raw) if pd.notna(planned_hours_raw) and str(planned_hours_raw).strip() != "" else None

    # ---- Header fields ----
    ws["D4"] = emp_name
    ws["D5"] = cid
    ws["D6"] = role
    ws["D7"] = domain
    if po_number != "":
        ws["D8"] = po_number
    if po_start != "":
        ws["D9"] = po_start
    if po_end != "":
        ws["D10"] = po_end
    ws["D11"] = sow_number
    ws["H4"] = calendar.month_name[month]
    ws["H6"] = year

    # ---- Daily rows (row 14 = day 1) ----
    # emp_hours = daily_hours[daily_hours["name_key"] == name_key].set_index("Item Date")["Quantity"]
    # sow_project_code = str(get_field(res_row, ["Project Code", "Project Number"], "")).strip()
    sow_project_code_raw = get_field(res_row, ["Project Code", "Project Number"], "")
    sow_project_code = re.sub(r"\.0$", "", str(sow_project_code_raw).strip())
    emp_hours = daily_hours[
        (daily_hours["name_key"] == name_key) &
        # (daily_hours["Project Number"].astype(str).str.strip() == sow_project_code)
        (daily_hours["Project Number"].astype(str).str.strip().str.replace(r"\.0$", "", regex=True) == sow_project_code)
    ].set_index("Item Date")["Quantity"]
    days_in_month = calendar.monthrange(year, month)[1]
    date_number_format = ws["B14"].number_format  # capture template's date format before overwriting

    proj_start_raw = get_field(res_row, ["Start date", "Start Date", "Project Start Date"], None)
    proj_end_raw = get_field(res_row, ["End date", "End Date", "Project End Date"], None)

    month_start = datetime(year, month, 1)
    month_end = datetime(year, month, days_in_month)

    proj_start = pd.to_datetime(proj_start_raw) if pd.notna(proj_start_raw) else month_start
    proj_end = pd.to_datetime(proj_end_raw) if pd.notna(proj_end_raw) else month_end

    effective_start = max(proj_start, month_start)
    effective_end = min(proj_end, month_end)

    for day in range(1, days_in_month + 1):
        row = 13 + day
        this_date = datetime(year, month, day)

        if this_date < effective_start or this_date > effective_end:
            for col in range(2, 9):
                ws.cell(row=row, column=col, value=None)
            continue

        hours = emp_hours.get(this_date, 0)
        # remark = "Correction adjusted" if (name_key, this_date) in neg_set else ""
        remark = "Correction adjusted" if (name_key, sow_project_code, this_date) in neg_set else ""
        is_weekend = this_date.weekday() >= 5  # 5 = Saturday, 6 = Sunday

        # Don't let total billed hours cross the SOW's planned duration hours
        if planned_hours is not None and hours > 0:
            remaining = planned_hours - hours_used_so_far
            if remaining <= 0:
                hours = 0
                remark = (remark + "; " if remark else "") + "Exceeds SOW planned hours - not billed"
            elif hours > remaining:
                hours = remaining
                remark = (remark + "; " if remark else "") + "Capped to SOW planned hours"
            hours_used_so_far += hours

        date_cell = ws.cell(row=row, column=2, value=this_date)
        date_cell.number_format = date_number_format
        ws.cell(row=row, column=3, value=this_date.strftime("%A").upper())

        if is_weekend and hours == 0:
            ws.cell(row=row, column=7, value=None)   # blank instead of 0
        else:
            ws.cell(row=row, column=7, value=float(hours))

        ws.cell(row=row, column=8, value=remark)
        remark_cell = ws.cell(row=row, column=8, value=remark)
        remark_cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left") 

        # Activity (D) and Work From (F) driven by whether hours were billed
        if hours == 0:
            if not is_weekend:
                ws.cell(row=row, column=4, value="On Leave")
                ws.cell(row=row, column=6, value="On Leave")
            # weekend + 0 hours -> leave D and F blank, same as G
        else:
            ws.cell(row=row, column=4, value="Project work")
            ws.cell(row=row, column=6, value=work_from_value)

    # ---- Dropdown for Work From column (F) — WFO / WFH ----
    # from openpyxl.worksheet.datavalidation import DataValidation
    
    # for day in range(1, days_in_month + 1):
    #     row = 13 + day
    #     this_date = datetime(year, month, day)

    #     if this_date < effective_start or this_date > effective_end:
    #         for col in range(2, 9):
    #             ws.cell(row=row, column=col, value=None)
    #         continue


        
    #     # hours = emp_hours.get(this_date, 0)
    #     # remark = "Correction adjusted" if (name_key, this_date) in neg_set else ""
    #     # is_weekend = this_date.weekday() >= 5  # 5 = Saturday, 6 = Sunday

    #     # date_cell = ws.cell(row=row, column=2, value=this_date)
    #     hours = emp_hours.get(this_date, 0)
    #     remark = "Correction adjusted" if (name_key, this_date) in neg_set else ""
    #     is_weekend = this_date.weekday() >= 5  # 5 = Saturday, 6 = Sunday

    #     # Don't let total billed hours cross the SOW's planned duration hours
    #     if planned_hours is not None and hours > 0:
    #         remaining = planned_hours - hours_used_so_far
    #         if remaining <= 0:
    #             hours = 0
    #             remark = (remark + "; " if remark else "") + "Exceeds SOW planned hours - not billed"
    #         elif hours > remaining:
    #             hours = remaining
    #             remark = (remark + "; " if remark else "") + "Capped to SOW planned hours"
    #         hours_used_so_far += hours

    #     date_cell = ws.cell(row=row, column=2, value=this_date)
    #     date_cell.number_format = date_number_format
    #     ws.cell(row=row, column=3, value=this_date.strftime("%A").upper())

    #     if is_weekend and hours == 0:
    #         ws.cell(row=row, column=7, value=None)   # blank instead of 0
    #     else:
    #         ws.cell(row=row, column=7, value=float(hours))

    #     ws.cell(row=row, column=8, value=remark)
    #     date_cell = ws.cell(row=row, column=2, value=this_date)
    #     date_cell.number_format = date_number_format
    #     ws.cell(row=row, column=3, value=this_date.strftime("%A").upper())

    #     if is_weekend and hours == 0:
    #         ws.cell(row=row, column=7, value=None)   # blank instead of 0
    #     else:
    #         ws.cell(row=row, column=7, value=float(hours))

    #     ws.cell(row=row, column=8, value=remark)

    #     # Activity (D) and Work From (F) driven by whether hours were billed
    #     if hours == 0:
    #         if not is_weekend:
    #             ws.cell(row=row, column=4, value="On Leave")
    #             ws.cell(row=row, column=6, value="On Leave")
    #         # weekend + 0 hours -> leave D and F blank, same as G
    #     else:
    #         ws.cell(row=row, column=4, value="Project work")
    #         ws.cell(row=row, column=6, value=work_from_value)
    #     # date_cell = ws.cell(row=row, column=2, value=this_date)
    #     # date_cell.number_format = date_number_format
    #     # ws.cell(row=row, column=3, value=this_date.strftime("%A").upper())

    #     # if is_weekend and hours == 0:
    #     #     ws.cell(row=row, column=7, value=None)   # blank instead of 0
    #     # else:
    #     #     ws.cell(row=row, column=7, value=float(hours))
    #     # if hours == 0:
    #     #     if not is_weekend:
    #     #         ws.cell(row=row, column=4, value="On Leave")
    #     #         ws.cell(row=row, column=6, value="On Leave")
    #     #     # weekend + 0 hours -> leave D and F blank, same as G
            
    #     # else:
    #     #     ws.cell(row=row, column=4, value="Project work")
    #     #     ws.cell(row=row, column=6, value=work_from_value)

    #     # ws.cell(row=row, column=8, value=remark)

    #     # # NEW: Activity (D) and Work From (F) driven by whether hours were billed
    #     # if hours == 0:
    #     #     ws.cell(row=row, column=4, value="On Leave")
    #     #     ws.cell(row=row, column=6, value="On Leave")
    #     # else:
    #     #     ws.cell(row=row, column=4, value="Project work")
    #     #     ws.cell(row=row, column=6, value=work_from_value)

    # # ---- Dropdown for Work From column (F) — WFO / WFH ----
    from openpyxl.worksheet.datavalidation import DataValidation
    # dv = DataValidation(type="list", formula1='"WFO,WFH"', allow_blank=True)
    dv = DataValidation(type="list", formula1='"WFO,WFH,On Leave"', allow_blank=True)
    dv.error = "Please select WFO or WFH"
    dv.errorTitle = "Invalid entry"
    dv.prompt = "Select WFO or WFH"
    dv.promptTitle = "Work From"
    ws.add_data_validation(dv)
    dv.add(f"F14:F{13 + days_in_month}")
    # ---- Dropdown for Activity column (D) — with custom entry allowed ----
    dv_activity = DataValidation(
        type="list",
        formula1='"On Leave,Sick Leave,Public Holiday,Project work"',
        allow_blank=True,
        showErrorMessage=False,  # allows typing a custom value beyond the 3 options
    )
    dv_activity.prompt = "Select an activity or type your own"
    dv_activity.promptTitle = "Activity"
    ws.add_data_validation(dv_activity)
    dv_activity.add(f"D14:D{13 + days_in_month}")
    # ---- Dropdown for Month (H4) ----
    dv_month = DataValidation(
        type="list",
        formula1='"January,February,March,April,May,June,July,August,September,October,November,December"',
        allow_blank=True,
    )
    dv_month.error = "Please select a valid month"
    dv_month.errorTitle = "Invalid entry"
    dv_month.prompt = "Select a month"
    dv_month.promptTitle = "Month"
    ws.add_data_validation(dv_month)
    dv_month.add("H4")

    # ---- Dropdown for Year (H6) ----
    dv_year = DataValidation(
        type="list",
        formula1='"2024,2025,2026,2027,2028,2029,2030,2031"',
        allow_blank=True,
    )
    dv_year.error = "Please select a valid year"
    dv_year.errorTitle = "Invalid entry"
    dv_year.prompt = "Select a year"
    dv_year.promptTitle = "Year"
    ws.add_data_validation(dv_year)
    dv_year.add("H6")
    # Blank out any leftover template rows beyond this month's day count
    for row in range(14 + days_in_month, 45):
        for col in range(2, 9):
            ws.cell(row=row, column=col, value=None)
    
    # ---- Sign-off block ----
    ws["C52"] = emp_name
    ws["F52"] = get_field(res_row, ["Capgemini Resposible", "Capgemini Responsible"])
    ws["F56"] = get_field(res_row, ["EGA Resposible", "EGA Responsible"])
    # ---- Lock only these specific cells; everything else stays editable ----
    # locked_cells = ["D6", "D7", "D8", "D9", "D10", "D11"]

    # # Unlock all cells first
    # for row in ws.iter_rows():
    #     for cell in row:
    #         cell.protection = Protection(locked=False)

    # # Lock only the required ones
    # for coord in locked_cells:
    #     ws[coord].protection = Protection(locked=True)

    # # Enable sheet protection so locking actually takes effect
    # ws.protection.sheet = True
    # ws.protection.password = "yourpassword"  # optional, remove this line if no password needed
    # ---- Freeze specific fields so employees can't edit them ----
    # from openpyxl.styles import Protection

    # locked_cells = ["D6", "D7", "D8", "D9", "D10", "D11"]

    # for row in ws.iter_rows():
    #     for cell in row:
    #         cell.protection = Protection(locked=False)

    # for coord in locked_cells:
    #     ws[coord].protection = Protection(locked=True)

    # ws.protection.sheet = True
    # ws.protection.enable()

    # out = io.BytesIO()
    # # wb.save(out)

    # # out = io.BytesIO()
    # ---- Lock everything EXCEPT Activity (D) and Work From (F) daily cells ----
    from openpyxl.styles import Protection

    # editable_cols = {4, 6, 7}  # column D = Activity, column F = Work From
    # editable_rows = set(range(14, 14 + days_in_month))  # daily table rows only

    # for row in ws.iter_rows():
    #     for cell in row:
    #         is_editable = (cell.row in editable_rows) and (cell.column in editable_cols)
    #         cell.protection = Protection(locked=not is_editable)
    # ws["D5"].protection = Protection(locked=False)   # ← new line: unlocks CID

    # ws.protection.sheet = True
    # ws.protection.enable()
    # editable_cols = {4, 6}  # column D = Activity, column F = Work From
    editable_cols = {4, 6, 7}  # column D = Activity, column F = Work From, column G = Billable Hours
    editable_rows = set(range(14, 14 + days_in_month))  # daily table rows only

    for row in ws.iter_rows():
        for cell in row:
            is_editable = (cell.row in editable_rows) and (cell.column in editable_cols)
            cell.protection = Protection(locked=not is_editable)

    # Explicitly unlock header cells that people should be able to fill in/edit
    # extra_unlocked_cells = ["D4", "D5", "H4" , "C52"]  # Name, CID, Month
    # for coord in extra_unlocked_cells:
    #     ws[coord].protection = Protection(locked=False)
    # extra_unlocked_cells = ["D4", "D5", "H4", "C52"]  # Name, CID, Month, Signature
    extra_unlocked_cells = ["D4", "D5", "C52", "F52"]  # Name, CID, Month, Signature, Capgemini Responsible
    for coord in extra_unlocked_cells:
        ws[coord].protection = Protection(locked=False)

    # Month activities summary is a merged block (B48:H50) — unlock every
    # cell in the merged range, since protection is per-cell even when merged
    # for row in ws["B48:H50"]:
    #     for cell in row:
    #         cell.protection = Protection(locked=False)
    # from openpyxl.styles import Alignment

    # Month activities summary — unlock AND align text/cursor to top-left
    for row in ws["B48:H50"]:
        for cell in row:
            cell.protection = Protection(locked=False)
            cell.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)

    ws.protection.sheet = True
    ws.protection.enable()

    # out = io.BytesIO()
    # # wb.save(out)
    # wb.save(out)
    # out.seek(0)
    # return out.getvalue()
    out = io.BytesIO()
    # wb.save(out)
    wb.save(out)
    out.seek(0)
    return out.getvalue(), hours_used_so_far


# =====================================================================
# 3. STREAMLIT UI
# =====================================================================
# =====================================================================
# 3. STREAMLIT UI
# =====================================================================

# st.title("📋 Timesheet Auto-Filler")
# st.write(
#     "Upload **SOW & Resource Details** and **Replicon Dump** files below. "
#     "The app will automatically generate a filled timesheet for **every employee** "
#     "found in the SOW file, and save each one with that employee's name."
# )
st.markdown("""
<div class="hero-banner">
    <div class="hero-icon">🗓️</div>
    <div class="hero-text">
        <h1>Timesheet Auto-Filler</h1>
        <p>Upload <b>SOW & Resource Details</b> and <b>Replicon Dump</b> files below.
        The app will automatically generate a filled timesheet for <b>every employee</b>
        found in the SOW file, and save each one with that employee's name.</p>
    </div>
</div>
""", unsafe_allow_html=True)
st.markdown("""
<div class="step-tracker">
    <div class="step active"><div class="step-num">1</div><div class="step-label">Upload SOW</div></div>
    <div class="step-line"></div>
    <div class="step active"><div class="step-num">2</div><div class="step-label">Upload Replicon</div></div>
    <div class="step-line"></div>
    <div class="step"><div class="step-num">3</div><div class="step-label">Generate</div></div>
    <div class="step-line"></div>
    <div class="step"><div class="step-num">4</div><div class="step-label">Download</div></div>
</div>
""", unsafe_allow_html=True)

col1, col2 = st.columns(2)
with col1:
    sow_file = st.file_uploader("1️⃣ Upload SOW & Resource Details (.xlsx)", type=["xlsx"])
with col2:
    replicon_file = st.file_uploader("2️⃣ Upload Replicon Dump (.xlsx)", type=["xlsx"])
stat_col1, stat_col2, stat_col3 = st.columns(3)
with stat_col1:
    st.markdown(f"""<div class="stat-chip">📄 SOW File<br><b>{'✅ Uploaded' if sow_file else '⏳ Pending'}</b></div>""", unsafe_allow_html=True)
with stat_col2:
    st.markdown(f"""<div class="stat-chip">🕒 Replicon Dump<br><b>{'✅ Uploaded' if replicon_file else '⏳ Pending'}</b></div>""", unsafe_allow_html=True)
with stat_col3:
    st.markdown(f"""<div class="stat-chip">📋 Template<br><b>Default</b></div>""", unsafe_allow_html=True)

# st.subheader("3️⃣ Timesheet Template")
# template_option = st.radio(
#     "Choose which timesheet template to fill:",
#     ["Use built-in default template (hardcoded in app)", "Upload a custom template"],
#     index=0,
# )

# custom_template_file = None
# if template_option == "Upload a custom template":
#     custom_template_file = st.file_uploader(
#         "Upload Timesheet Template (.xlsx)", type=["xlsx"], key="template_upload"
#     )
# else:
#     st.caption("✅ Using the built-in default timesheet template embedded in this app.")
template_option = "Use built-in default template (hardcoded in app)"
custom_template_file = None
st.divider()
generate_btn = st.button("🚀 Generate Timesheets", type="primary", use_container_width=True)

if generate_btn:
    if not sow_file or not replicon_file:
        st.error("Please upload both the SOW & Resource Details and Replicon Dump files.")
    elif template_option == "Upload a custom template" and not custom_template_file:
        st.error("Please upload a custom timesheet template, or switch to the built-in default.")
    else:
        try:
            with st.spinner("Reading files..."):
                sow_df = load_sow(sow_file)
                rep_df = load_replicon(replicon_file)
                daily_hours, neg_set = build_daily_hours(rep_df)

                # Auto-detect month/year from the Replicon dump's dates
                # month = int(rep_df["Item Date"].dt.month.mode()[0])
                # year = int(rep_df["Item Date"].dt.year.mode()[0])
                # Detect ALL distinct months present in the Replicon dump
                # (not just the single most-common one) so no billable
                # hours from any month get silently dropped.
                month_year_pairs = sorted(
                    set(zip(rep_df["Item Date"].dt.year, rep_df["Item Date"].dt.month))
                )

                template_bytes = (
                    custom_template_file.read() if custom_template_file else get_default_template_bytes()
                )

            # st.success(
            #     f"Detected period: **{calendar.month_name[month]} {year}**. "
            #     f"Found **{len(sow_df)}** resource(s) in the SOW file."
            # )

            # generated_files = {}
            # unmatched_employees = []
            # rep_name_keys = daily_hours["name_key"].unique().tolist()
            # progress = st.progress(0.0)
            # for i, (_, res_row) in enumerate(sow_df.iterrows()):
            #     # matched_key = find_matching_rep_key(res_row["Resource name"], rep_name_keys)
            #     matched_key = find_matching_rep_key(get_field(res_row, ["Resource name", "Employee Name", "Name"]), rep_name_keys)
            #     # if matched_key is None:
            #     #     unmatched_employees.append(res_row["Resource name"])
            #     if matched_key is None:
            #         unmatched_employees.append(get_field(res_row, ["Resource name", "Employee Name", "Name"]))
            #     data = fill_timesheet_for_employee(
            #         template_bytes, res_row, daily_hours, neg_set, month, year, matched_key
            #     )
            #     # emp_name_clean = str(res_row.get("Resource name", f"Employee_{i+1}")).strip().replace(" ", "_")
            #     emp_name_clean = str(get_field(res_row, ["Resource name", "Employee Name", "Name"], f"Employee_{i+1}")).strip().replace(" ", "_")
            #     fname = f"Timesheet_{calendar.month_name[month]}_{year}_{emp_name_clean}.xlsx"
            #     generated_files[fname] = data
            #     progress.progress((i + 1) / len(sow_df))
            months_label = ", ".join(f"{calendar.month_name[m]} {y}" for y, m in month_year_pairs)
            st.success(
                f"Detected **{len(month_year_pairs)}** month(s) in the Replicon dump: **{months_label}**. "
                f"Found **{len(sow_df)}** resource(s) in the SOW file."
            )

            # generated_files = {}
            # unmatched_employees = set()
            # rep_name_keys = daily_hours["name_key"].unique().tolist()
            generated_files = {}
            unmatched_employees = set()
            invalid_cid_employees = []
            rep_name_keys = daily_hours["name_key"].unique().tolist()

            total_steps = len(sow_df) * len(month_year_pairs)
            progress = st.progress(0.0)
            step = 0

            for _, res_row in sow_df.iterrows():
                emp_display_name = get_field(res_row, ["Resource name", "Employee Name", "Name"], "Employee")
                emp_cid = get_field(res_row, ["CID", "C.I.D", "C.I.D.", "Cid"], "")
                if emp_cid and not is_valid_cid(emp_cid):
                    invalid_cid_employees.append(f"{emp_display_name} ({emp_cid})")
                # matched_key = find_matching_rep_key(emp_display_name, rep_name_keys)
                # if matched_key is None:
                #     unmatched_employees.add(emp_display_name)
                #     hours_used_so_far = 0
                matched_key = find_matching_rep_key(emp_display_name, rep_name_keys)
                hours_used_so_far = 0
                if matched_key is None:
                    unmatched_employees.add(emp_display_name)

                for year, month in month_year_pairs:
                    step += 1
                    progress.progress(step / total_steps)

                    # Skip generating a file for months with zero overlap
                    # between this employee's SOW range and the month
                    proj_start_raw = get_field(res_row, ["Start date", "Start Date", "Project Start Date"], None)
                    proj_end_raw = get_field(res_row, ["End date", "End Date", "Project End Date"], None)
                    days_in_month = calendar.monthrange(year, month)[1]
                    month_start = datetime(year, month, 1)
                    month_end = datetime(year, month, days_in_month)
                    proj_start = pd.to_datetime(proj_start_raw) if pd.notna(proj_start_raw) else month_start
                    proj_end = pd.to_datetime(proj_end_raw) if pd.notna(proj_end_raw) else month_end
                    effective_start = max(proj_start, month_start)
                    effective_end = min(proj_end, month_end)
                    if effective_start > effective_end:
                        continue  # no overlap with SOW range -> don't generate a blank file

                    # data = fill_timesheet_for_employee(
                    #     template_bytes, res_row, daily_hours, neg_set, month, year, matched_key
                    # )
                    data, hours_used_so_far = fill_timesheet_for_employee(
                        template_bytes, res_row, daily_hours, neg_set, month, year, matched_key, hours_used_so_far
                    )
                    # emp_name_clean = str(emp_display_name).strip().replace(" ", "_")
                    emp_name_clean = re.sub(r"[^\w\-]", "_", str(emp_display_name).strip())
                    fname = f"Timesheet_{calendar.month_name[month]}_{year}_{emp_name_clean}.xlsx"
                    generated_files[fname] = data

            unmatched_employees = list(unmatched_employees)
            if unmatched_employees:
                st.warning(
                    "⚠️ No matching hours found in the Replicon dump for: "
                    + ", ".join(unmatched_employees)
                    + ". Their timesheet was generated with all days marked as Leave — "
                    "double check the name spelling in both files."
                )
            if invalid_cid_employees:
                st.warning(
                "⚠️ CID format looks incorrect (should be 'C' followed by 8 digits) for: "
                + ", ".join(invalid_cid_employees)
                + ". Timesheets were still generated — please verify these CIDs."
            )


            # st.subheader("✅ Generated Timesheets")
            # for fname, data in generated_files.items():
            #     st.download_button(
            #         label=f"⬇️ Download {fname}",
            #         data=data,
            #         file_name=fname,
            #         mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            #         key=fname,
            #     )

            # # Zip download of all files together
            # zip_buffer = io.BytesIO()
            # with zipfile.ZipFile(zip_buffer, "w") as zf:
            #     for fname, data in generated_files.items():
            #         zf.writestr(fname, data)
            # zip_buffer.seek(0)

            # st.download_button(
            #     "📦 Download All Timesheets as ZIP",
            #     data=zip_buffer.getvalue(),
            #     # file_name=f"All_Timesheets_{calendar.month_name[month]}_{year}.zip",
            #     file_name="All_Timesheets.zip",
            #     mime="application/zip",
            #     use_container_width=True,
            # )
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w") as zf:
                for fname, data in generated_files.items():
                    zf.writestr(fname, data)
            zip_buffer.seek(0)

            st.session_state["generated_files"] = generated_files
            st.session_state["zip_bytes"] = zip_buffer.getvalue()

        except Exception as e:
            st.error(f"Something went wrong: {e}")
            st.exception(e)
if "generated_files" in st.session_state:
    st.divider()
    st.subheader("✅ Generated Timesheets")

    st.download_button(
        "📦 Download All Timesheets as ZIP",
        data=st.session_state["zip_bytes"],
        file_name="All_Timesheets.zip",
        mime="application/zip",
        use_container_width=True,
        type="primary",
    )

    # # with st.expander(f"Or download individual files ({len(st.session_state['generated_files'])})"):
    # with st.expander(f"Or download individual files ({len(st.session_state['generated_files'])})", expanded=True):
    #     for fname, data in st.session_state["generated_files"].items():
    #         st.download_button(
    #             label=f"⬇️ {fname}",
    #             data=data,
    #             file_name=fname,
    #             mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    #             key=fname,
    #         )
    with st.expander(f"Or download individual files ({len(st.session_state['generated_files'])})", expanded=True):
        for fname, data in st.session_state["generated_files"].items():
            display_name = fname.replace("Timesheet_", "").replace(".xlsx", "").replace("_", " ")
            card_col1, card_col2 = st.columns([4, 1])
            with card_col1:
                st.markdown(f"""
                    <div class="file-card">
                        <div class="file-icon">📊</div>
                        <div class="file-name">{display_name}</div>
                    </div>
                """, unsafe_allow_html=True)
            with card_col2:
                st.download_button("⬇️", data=data, file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=fname)
    st.markdown("""
<div class="app-footer">
    <span>⚡ Powered by Capgemini Automation</span>
</div>
""", unsafe_allow_html=True)

# st.divider()
# with st.expander("ℹ️ How field mapping works"):
#     st.markdown(
#         """
#         - **Partner Emp. Name / CID / Role / Domain / PO Number / SOW Number** ← from SOW & Resource Details file
#         - **Daily Billable Hours** ← summed per employee per date from the Replicon dump
#           (negative correction rows are netted automatically)
#         - **Activity** ← SOW name, filled only on days with hours > 0
#         - **Work From** ← "WFO" if role contains "onsite", otherwise "WFH"
#         - **Days with 0 hours** ← marked as "Leave"
#         - **Capgemini/EGA Responsible, sign-off Name** ← from SOW & Resource Details file 
#         """
#     )
st.divider()
with st.expander("ℹ️ How field mapping works"):
    st.markdown(
        """
        - **Partner Emp. Name / CID / Role / Domain / PO Number / SOW Number** ← from SOW & Resource Details file
        - **Daily Billable Hours** ← summed per employee per date from the Replicon dump
          (negative correction rows are netted automatically)
        - **Activity** ← "Project work" on days with billable hours, "On Leave" on 0-hour weekdays
        - **Work From** ← "WFO" only if role contains "onsite"; left blank otherwise (not auto-filled as WFH)
        - **Days with 0 hours (weekdays)** ← marked as "On Leave"
        - **Weekends with 0 hours** ← left blank
        - **Capgemini/EGA Responsible, sign-off Name** ← from SOW & Resource Details file 
        """
    )











